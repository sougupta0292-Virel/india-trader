"""
India Trader v3 — Replica of US Stocks Bot
White/Blue theme · Angel One · NSE/BSE · Nifty · BankNifty · Sensex
"""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))
import json, glob as _glob_mod
import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from datetime import datetime, date, time, timezone, timedelta
from collections import defaultdict

# ── Backend imports ────────────────────────────────────────────────────────────
try:
    from strategy_engine import GlobalCuesEngine, OIAnalyser, FibonacciEngine, SignalGenerator
    STRATEGY_ENGINE_OK = True
except Exception:
    STRATEGY_ENGINE_OK = False

try:
    from backtester import Backtester, FibOIBacktester, IntraHunterBacktester, PerIndexBacktester, SubasishBacktester, IntraDay5MinBacktester, BtResult
    BT_OK = True
except Exception:
    BT_OK = False

try:
    from india_all_strategies_engine import (
        MultiTFScanner, MultiTFBacktester,
        TF_STRATEGIES, STRATEGY_DESCRIPTIONS, INDEX_UNIVERSE,
        StratBtResult, StratTrade
    )
    MULTI_TF_OK = True
except Exception as _mte:
    MULTI_TF_OK = False
    TF_STRATEGIES = {"5M": [], "15M": [], "1H": [], "4H": []}
    STRATEGY_DESCRIPTIONS = {}

try:
    from pos_auto_trader import POSAutoTrader, POSState, POSTrade
    POS_OK = True
except Exception:
    POS_OK = False

try:
    from options_auto_trader import OptionsAutoTrader, OptionsState, OptionTrade
    OPT_OK = True
except Exception:
    OPT_OK = False

try:
    from options_futures_strategies import (
        scan_all as _ofs_scan_all,
        get_straddle_strangle as _ofs_straddle,
        get_two_expiries as _ofs_expiries,
        get_futures_expiry as _ofs_fut_expiry,
        get_all_expiries as _ofs_all_expiries,
        get_all_option_strategies as _ofs_all_strategies,
        get_three_option_expiries as _ofs_three_opt_exp,
        get_two_futures_expiries as _ofs_two_fut_exp,
        get_past_n_option_expiries as _ofs_past_opt,
        get_past_n_futures_expiries as _ofs_past_fut,
        backtest_option_strategies_period as _ofs_bt_opt_strat,
        StrategyBacktester as _OFSBacktester,
        OptionsFuturesAutoTrader as _OFSAutoTrader,
        IST as _OFS_IST,
        STRATEGIES as _OFS_STRATEGIES,
    )
    OFS_OK = True
except Exception as _ofs_err:
    OFS_OK = False
    _ofs_err_msg = str(_ofs_err)
    _OFS_STRATEGIES = ["CHANDELIER", "TUX_SUPERTREND", "CE_REGIME"]

try:
    from angel_one import AngelOneClient
    ANGEL_OK = True
except Exception:
    ANGEL_OK = True
    class AngelOneClient:
        def __init__(self, **kw): pass
        def connect(self): return False, "pip install smartapi-python pyotp"
        def get_funds(self): return {}
        def get_quote(self, s): return {}
        def get_option_chain(self, s): return {}

try:
    from vwap_cpr import CPRCalculator, VWAPCalculator
    VWAP_OK = True
except Exception:
    VWAP_OK = False

_IST = timezone(timedelta(hours=5, minutes=30))
_APP_DIR = os.path.dirname(os.path.abspath(__file__))
_RESULTS_DIR = os.path.join(_APP_DIR, "results")

# ── NSE instruments ────────────────────────────────────────────────────────────
NSE_INDICES   = ["NIFTY", "BANKNIFTY", "SENSEX"]
NSE_LOT_SIZE  = {"NIFTY": 75, "BANKNIFTY": 35, "SENSEX": 20}
NSE_YF_MAP    = {"NIFTY": "^NSEI", "BANKNIFTY": "^NSEBANK", "SENSEX": "^BSESN", "INDIAVIX": "^INDIAVIX"}
NSE_TOP_STOCKS = [
    "RELIANCE","TCS","INFY","HDFCBANK","ICICIBANK","SBIN","WIPRO","BHARTIARTL",
    "ITC","KOTAKBANK","LT","AXISBANK","BAJFINANCE","HINDUNILVR","MARUTI",
    "SUNPHARMA","TECHM","ASIANPAINT","TITAN","ULTRACEMCO",
    "ONGC","POWERGRID","NTPC","COALINDIA","BAJAJFINSV","ADANIPORTS","GRASIM",
    "M&M","HCLTECH","TATAMOTORS","TATASTEEL","JSWSTEEL","INDUSINDBK","HINDALCO",
    "DRREDDY","CIPLA","DIVISLAB","EICHERMOT","HEROMOTOCO","APOLLOHOSP",
    "BRITANNIA","NESTLEIND","TATACONSUM","BPCL","LTIM","HDFCLIFE","SBILIFE",
    "BAJAJ-AUTO","UPL","ADANIENT"
]
# Add .NS suffix mapping for yfinance
for _s in NSE_TOP_STOCKS:
    if _s not in NSE_YF_MAP:
        NSE_YF_MAP[_s] = _s + ".NS"
TV_MAP = {"NIFTY": "NSE:NIFTY", "BANKNIFTY": "NSE:BANKNIFTY", "SENSEX": "BSE:SENSEX"}

# ── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(page_title="India Trader v3", page_icon="🇮🇳", layout="wide", initial_sidebar_state="expanded")

# ── White/Blue theme (same as US app) ─────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@400;600;700&family=Inter:wght@400;600;700&display=swap');
html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
.stApp { background-color: #f0f6ff; color: #0f172a; }
.main .block-container { padding: 1.2rem 2rem; max-width: 1500px; }

/* ── Sidebar ────────────────────────────────────────────────────── */
[data-testid="stSidebar"] { background: #e8efff; border-right: 2px solid #c5d5f0; }
[data-testid="stSidebar"] .block-container { padding: 1rem; }
[data-testid="stSidebar"] * { color: #1e293b !important; }

/* ── Tabs ───────────────────────────────────────────────────────── */
.stTabs [data-baseweb="tab-list"] { background: #1e3a5f; border-radius: 10px; padding: 4px; gap: 4px; }
.stTabs [data-baseweb="tab"] { color: #fde68a !important; font-family:'JetBrains Mono'; font-size:12px; font-weight:600; border-radius:8px; padding:6px 16px; }
.stTabs [data-baseweb="tab"] p { color: #fde68a !important; }
.stTabs [data-baseweb="tab"]:hover { color:#ffffff !important; background:#2d5a9e; }
.stTabs [data-baseweb="tab"]:hover p { color:#ffffff !important; }
.stTabs [aria-selected="true"] { color:#0f172a !important; background:#ffffff; font-weight:700; box-shadow:0 2px 8px rgba(0,0,0,0.15); }
.stTabs [aria-selected="true"] p { color:#0f172a !important; }

/* ── Selectbox / Dropdown ───────────────────────────────────────── */
[data-baseweb="select"] > div { background:#ffffff !important; border:1px solid #c5d5f0 !important; border-radius:8px !important; }
[data-baseweb="select"] > div:focus-within { border-color:#3b82f6 !important; box-shadow:0 0 0 2px rgba(59,130,246,0.2) !important; }
[data-baseweb="select"] span,
[data-baseweb="select"] div { color:#0f172a !important; }
[data-baseweb="select"] svg { fill:#3b82f6 !important; }
/* Dropdown popover/menu */
[data-baseweb="popover"] { background:#ffffff !important; border:1px solid #c5d5f0 !important; border-radius:8px !important; }
[data-baseweb="menu"] { background:#ffffff !important; }
[data-baseweb="option"] { background:#ffffff !important; color:#0f172a !important; }
[data-baseweb="option"]:hover,
[data-baseweb="option"][aria-selected="true"] { background:#e8efff !important; color:#1d4ed8 !important; }
/* Number input / text input */
[data-baseweb="input"] { background:#ffffff !important; border:1px solid #c5d5f0 !important; border-radius:8px !important; }
[data-baseweb="input"] input { background:#ffffff !important; color:#0f172a !important; }
[data-baseweb="textarea"] { background:#ffffff !important; border:1px solid #c5d5f0 !important; }
[data-baseweb="textarea"] textarea { background:#ffffff !important; color:#0f172a !important; }
/* Multiselect tags */
[data-baseweb="tag"] { background:#dbeafe !important; color:#1d4ed8 !important; border-radius:4px !important; }

/* ── Buttons ────────────────────────────────────────────────────── */
.stButton > button { font-family:'JetBrains Mono'; border-radius:8px; color:#1e293b; background:#ffffff; border:1px solid #c5d5f0; }
.stButton > button:hover { border-color:#3b82f6; color:#1d4ed8; }
.stButton > button[kind="primary"] { background: linear-gradient(135deg,#1d4ed8,#3b82f6); border:none; color:#fff !important; font-weight:700; }

/* ── Metrics / cards ────────────────────────────────────────────── */
div[data-testid="stMetricValue"] { font-family:'JetBrains Mono'; color:#0f172a; }
.mcard { background:#ffffff; border:1px solid #dde9ff; border-radius:10px; padding:14px; box-shadow:0 2px 8px rgba(29,78,216,0.08); }
.mlbl  { color:#5a72a0; font-size:10px; letter-spacing:1.5px; text-transform:uppercase; font-family:'JetBrains Mono'; }
.mval  { font-size:22px; font-weight:700; font-family:'JetBrains Mono'; color:#1e293b; margin-top:4px; }

/* ── Misc ───────────────────────────────────────────────────────── */
div.stAlert { border-radius:8px; }
hr { border-color:#e0e9ff !important; }
.stMarkdown p { color:#0f172a; }
input, textarea, select { background:#ffffff !important; color:#0f172a !important; border:1px solid #c5d5f0 !important; }
/* Expander headers */
[data-testid="stExpander"] summary { color:#0f172a !important; background:#ffffff; border:1px solid #c5d5f0; border-radius:8px; }
/* Checkbox / radio */
[data-testid="stCheckbox"] label,
[data-testid="stRadio"] label { color:#0f172a !important; }
/* Slider — track=black, filled portion+thumb=yellow */
[data-testid="stSlider"] [data-baseweb="slider"] div { background:#1f2937 !important; }
[data-testid="stSlider"] [data-baseweb="slider"] div[role="progressbar"] > div { background:#fbbf24 !important; }
[data-testid="stSlider"] [data-baseweb="slider"] div[data-baseweb="thumb"] > div { background:#fbbf24 !important; }
</style>
""", unsafe_allow_html=True)

# ── Load env ───────────────────────────────────────────────────────────────────
def _load_env():
    for name in ['.env', '.env.txt']:
        ep = os.path.join(_APP_DIR, name)
        if os.path.exists(ep):
            with open(ep) as f:
                for line in f:
                    line = line.strip()
                    if line and '=' in line and not line.startswith('#'):
                        k, _, v = line.partition('=')
                        os.environ[k.strip()] = v.strip()
_load_env()

# ── Session state ──────────────────────────────────────────────────────────────
_SS_DEFAULTS = {
    'angel_client': None, 'angel_connected': False, 'angel_funds': {},
    'spot': 0.0, 'prev_high': 0.0, 'prev_low': 0.0, 'prev_close': 24480.0,
    'day_high': 0.0, 'day_low': 0.0,
    'global_bias': 'NEUTRAL', 'global_score': 5.0,
    'oi_df': None, 'oi_signal': None, 'oi_source': 'Simulated',
    'cpr': None, 'vwap': None, 'signal': None, 'cues': None,
    'pos_trader': None, 'pos_state': None,
    'opt_trader': None,
    'bt_result_pos': None, 'bt_5min': {}, 'bt_per_index': {},
    'ticker_prices': {},
}
for _k, _v in _SS_DEFAULTS.items():
    if _k not in st.session_state:
        st.session_state[_k] = _v

# ── Helpers ────────────────────────────────────────────────────────────────────
def _ist_now(): return datetime.now(_IST)
def _market_open():
    t = _ist_now().time()
    return time(9, 15) <= t <= time(15, 30)

def fetch_nse_live_data(symbol="NIFTY"):
    import yfinance as yf
    ticker = NSE_YF_MAP.get(symbol, "^NSEI")
    try:
        tk   = yf.Ticker(ticker)
        hist = tk.history(period="5d", interval="1d")
        if len(hist) >= 2:
            prev = hist.iloc[-2]; curr = hist.iloc[-1]
            return {
                'spot':       round(float(curr['Close']), 2),
                'prev_high':  round(float(prev['High']), 2),
                'prev_low':   round(float(prev['Low']), 2),
                'prev_close': round(float(prev['Close']), 2),
                'prev_open':  round(float(prev['Open']), 2),
                'change_pct': round((float(curr['Close'])-float(prev['Close']))/float(prev['Close'])*100, 2),
                'day_high':   round(float(curr['High']), 2),
                'day_low':    round(float(curr['Low']), 2),
            }
    except Exception:
        pass
    return None

@st.cache_data(ttl=60, show_spinner=False)
def _fetch_nse_ticker_prices(symbols):
    import yfinance as _yf
    result = {}
    for sym in symbols:
        ticker = NSE_YF_MAP.get(sym, sym)
        try:
            fi = _yf.Ticker(ticker).fast_info
            lp = float(fi.last_price or 0)
            pc = float(fi.regular_market_previous_close or lp)
            chg = ((lp - pc) / pc * 100) if pc > 0 else 0.0
            result[sym] = (lp, chg)
        except Exception:
            result[sym] = (0.0, 0.0)
    return result

def _all_trades():
    trades  = []
    seen_ids = set()

    # 1. Existing results/ JSON files (backtest exports, manual saves)
    for pf in sorted(_glob_mod.glob(os.path.join(_RESULTS_DIR, "trades_*.json"))):
        try:
            with open(pf, encoding='utf-8') as fh:
                for t in json.load(fh):
                    tid = t.get('id', '')
                    if tid and tid in seen_ids: continue
                    if tid: seen_ids.add(tid)
                    trades.append(t)
        except Exception:
            pass

    # 2. Live equity/futures bot trades from pos_state_SYMBOL_TF.json
    _idx_set = {'NIFTY', 'BANKNIFTY', 'SENSEX'}
    for pf in sorted(_glob_mod.glob(os.path.join(_APP_DIR, "pos_state_*.json"))):
        try:
            fname = os.path.basename(pf)
            inner = fname[len('pos_state_'):-len('.json')]   # e.g. "TCS_5M"
            parts = inner.rsplit('_', 1)
            sym = parts[0] if len(parts) == 2 else inner
            tf  = parts[1].lower() if len(parts) == 2 else '5m'
            seg = 'Futures' if sym in _idx_set else 'Equity'
            with open(pf, encoding='utf-8') as fh:
                data = json.load(fh)
            for t in data.get('history', []):
                tid = t.get('id', '')
                if tid and tid in seen_ids: continue
                if tid: seen_ids.add(tid)
                entry_px = t.get('entry', 0)
                pnl_pts  = t.get('pnl_pts', 0)
                t2 = dict(t)
                t2.setdefault('Symbol',      sym)
                t2.setdefault('Segment',     seg)
                t2.setdefault('TF',          tf)
                t2.setdefault('Strategy',    t.get('strategy', ''))
                t2.setdefault('Direction',   t.get('direction', ''))
                t2.setdefault('Entry ₹',     entry_px)
                t2.setdefault('Exit ₹',      t.get('exit_price', 0))
                t2.setdefault('Net PnL ₹',   t.get('pnl_rs', 0))
                t2.setdefault('PnL %',       round(pnl_pts / max(abs(entry_px), 1) * 100, 2))
                t2.setdefault('Charges ₹',   0)
                t2.setdefault('Date',        t.get('date', ''))
                t2.setdefault('Exit Reason', t.get('exit_reason', ''))
                t2.setdefault('Qty',         t.get('qty', 1))
                trades.append(t2)
        except Exception:
            pass

    # 3. Live options bot trades from options_state*.json (bare or options_state_INSTRUMENT_TF.json)
    for pf in sorted(_glob_mod.glob(os.path.join(_APP_DIR, "options_state*.json"))):
        try:
            fname = os.path.basename(pf)
            if fname == "options_state.json":
                sym, tf = "NIFTY", "5m"
            else:
                inner = fname[len('options_state_'):-len('.json')]
                parts = inner.rsplit('_', 1)
                sym = parts[0] if len(parts) == 2 else inner
                tf  = parts[1].lower() if len(parts) == 2 else '5m'
            with open(pf, encoding='utf-8') as fh:
                data = json.load(fh)
            for t in data.get('history', []):
                tid = t.get('id', '')
                if tid and tid in seen_ids: continue
                if tid: seen_ids.add(tid)
                entry_px = t.get('entry_premium', t.get('entry', 0))
                pnl_pts  = t.get('pnl_pts', 0)
                t2 = dict(t)
                t2.setdefault('Symbol',      sym)
                t2.setdefault('Segment',     'Options')
                t2.setdefault('TF',          tf)
                t2.setdefault('Strategy',    t.get('strategy', ''))
                t2.setdefault('Direction',   t.get('direction', ''))
                t2.setdefault('Entry ₹',     entry_px)
                t2.setdefault('Exit ₹',      t.get('exit_premium', t.get('exit_price', 0)))
                t2.setdefault('Net PnL ₹',   t.get('pnl_rs', 0))
                t2.setdefault('PnL %',       round(pnl_pts / max(abs(entry_px), 1) * 100, 2) if entry_px else 0)
                t2.setdefault('Charges ₹',   0)
                t2.setdefault('Date',        t.get('date', ''))
                t2.setdefault('Exit Reason', t.get('exit_reason', ''))
                t2.setdefault('Qty',         t.get('qty', 1))
                trades.append(t2)
        except Exception:
            pass

    # 4. GitHub Actions paper trades from data/paper_trades.json
    _gh_paper_file = os.path.join(_APP_DIR, "data", "paper_trades.json")
    if os.path.exists(_gh_paper_file):
        try:
            with open(_gh_paper_file, encoding="utf-8") as fh:
                _gh_data = json.load(fh)
            for t in _gh_data.get("closed_trades", []):
                tid = t.get("id", "")
                if tid and tid in seen_ids: continue
                if tid: seen_ids.add(tid)
                trades.append(t)
        except Exception:
            pass

    return trades

def _trade_entry(t):
    return float(t.get('Entry ₹', t.get('entry_price', t.get('entry', 0))) or 0)

def _trade_exit(t):
    return float(t.get('Exit ₹', t.get('exit_price', t.get('exit', 0))) or 0)

def _trade_net_pnl(t):
    return float(t.get('Net PnL ₹', t.get('net_pnl', 0)) or 0)

def _trade_charges(t):
    return float(t.get('Charges ₹', t.get('charges', 0)) or 0)

def _trade_gross_pnl(t):
    return _trade_net_pnl(t) + _trade_charges(t)

def _trade_pnl_pct(t):
    return float(t.get('PnL %', t.get('pnl_pct', 0)) or 0)

def _trade_tf(t):
    seg = t.get('Segment', '')
    for tf in ['4h','1h','15m','5m','4H','1H','15M','5M']:
        if tf.lower() in seg.lower():
            return tf.upper().replace('M','m').replace('H','h')
    raw = t.get('TF', t.get('tf', t.get('timeframe', '5m')))
    return str(raw).lower() if raw else '5m'

def _tv_link(sym):
    mapped = TV_MAP.get(sym, f"NSE:{sym}")
    tf_int = "5"
    return f"https://www.tradingview.com/chart/?symbol={mapped}&interval={tf_int}"

def _india_charges_estimate(seg, qty, entry):
    """Estimate India charges if not already in trade record."""
    seg_lo = (seg or '').lower()
    broker = 20.0  # flat ₹20 per side (Angel One)
    if 'opt' in seg_lo:
        # Options: ₹20 × 2 + STT 0.05% of sell premium + exchange charges
        return round(broker * 2 + entry * qty * 0.0005 + entry * qty * 0.0005, 2)
    elif 'fut' in seg_lo:
        # Futures: ₹20 × 2 + STT 0.01% sell side + exchange 0.002%
        notional = entry * qty
        return round(broker * 2 + notional * 0.0001 + notional * 0.00002, 2)
    else:
        # Equity intraday: ₹20 × 2 + STT 0.025% sell
        return round(broker * 2 + entry * qty * 0.00025, 2)

# ── Strategy map for banners ───────────────────────────────────────────────────
_IND_SM = {
    "Equity": {
        "5m":  ["PPT Fibonacci","PPT Gap Fill","PPT Tail Rev","9/20 EMA Pullback",
                "5 EMA Cross","VWAP Reclaim","Round Number","Gap and Go","ORB Breakout"],
        "15m": ["Bull Flag","Bear Flag","Ascending Triangle","Double Bottom","Double Top",
                "Head & Shoulders","Cup & Handle","PPT Fibonacci","PPT Gap Fill",
                "9/20 EMA Pullback","VWAP CPR","RSI Divergence","Bollinger Squeeze","Inside Candle"],
        "1h":  ["Trend Following","Mean Reversion","Breakout Momentum","Momentum Swing","Swing High/Low",
                "PPT Fibonacci","VWAP CPR","EMA Stack","RSI Divergence","MACD Histogram",
                "Bollinger Bands","Volume Surge","Support Bounce","Earnings Momentum",
                "Fabio Daily","Supertrend Signal","Volume Profile POC","AI Narrative","Relative Strength","MA Crossover","Mean Reversion"],
        "4h":  ["Trend Following","Mean Reversion","Breakout Momentum","Momentum Swing","Swing High/Low",
                "PPT Fibonacci","VWAP CPR","EMA Stack","RSI Divergence","MACD Histogram",
                "Bollinger Bands","Volume Surge","Support Bounce","Fabio Daily",
                "Supertrend Signal","Volume Profile POC","AI Narrative","Relative Strength","MA Crossover","Inside Day","Mean Reversion"],
    },
    "Options": {
        "5m":  ["ATM Call Momentum","ATM Put Momentum","ORB Option Play","VWAP CE/PE"],
        "15m": ["Bull Call Spread","Bear Put Spread","Vertical Credit Spread",
                "Synthetic Long","Gamma Scalp","Calendar Spread"],
        "1h":  ["Iron Condor","Iron Butterfly","Long Butterfly","Broken Wing Butterfly",
                "Long Straddle","Short Straddle","Long Strangle","Short Strangle",
                "Jade Lizard","Bull Call Spread","Bear Put Spread","Collar Strategy",
                "Protective Put","Ratio Spread","Risk Reversal","LEAPS Bull Call","Wheel Strategy"],
        "4h":  ["Iron Condor","Iron Butterfly","Short Straddle","Short Strangle",
                "Jade Lizard","Covered Call","Cash Secured Put","Calendar Spread",
                "Diagonal Spread","Put Skew Trade","Wheel Strategy","Poor Man Covered Call",
                "Risk Reversal","Broken Wing Butterfly","Long Butterfly","LEAPS Bull Call","Ratio Spread"],
    },
    "Futures": {
        "5m":  ["ORB Nifty","Gift Nifty Break","Nifty Momentum Scalp","PPT Fibonacci",
                "VWAP Reclaim","Gap Fill","Reversal at VWAP","RTH Open Drive"],
        "15m": ["ORB 15M","EMA Pullback","Bollinger Squeeze","VWAP CPR","Inside Candle",
                "Nifty Gap Fill","Lunch Hour Fade","Power Hour Breakout",
                "BankNifty Momentum","Sensex Divergence","Cumulative Delta Divergence"],
        "1h":  ["EMA Trend","Momentum Swing","RSI Divergence","Breakout Momentum",
                "Market Profile TPO","Nifty 50 EMA","BankNifty Relative Strength",
                "Volume Node Rejection","Fibonacci Cluster","Nifty Trend Follow","India VIX Reversal"],
        "4h":  ["Swing EMA","Fib Swing","RSI Reversal","Trend Continuation","Nifty Inside Day",
                "MNF/MBNF Swing","FII DII Flow","India VIX Spike","Budget Event Play",
                "Volume Profile TPO","Fibonacci Cluster"],
    },
}

def _make_tf_banner(segment, color=None, label=None):
    _colors = {"Equity":"#16a34a","Options":"#d97706","Futures":"#7c3aed","Stocks":"#1d4ed8"}
    if color is None: color = _colors.get(segment, "#1d4ed8")
    if label is None: label = segment
    sm = _IND_SM.get(segment, {})
    _tf_labels = {"5m":"5-Min (Scalp)","15m":"15-Min (Intraday)","1h":"1-Hour (Swing)","4h":"4-Hour (Position)"}
    _tf_bg     = {"5m":"#fef9c3","15m":"#fde68a","1h":"#fcd34d","4h":"#f59e0b"}
    cells = ""
    for _btf in ["5m","15m","1h","4h"]:
        _strats = sm.get(_btf, [])
        _cnt = len(_strats)
        if _cnt == 0: continue
        _chips = "".join(
            f'<span style="background:#fff8;color:{color};border:1px solid {color}44;'
            f'padding:1px 6px;border-radius:8px;margin:1px 2px;font-size:9px">{_s}</span>'
            for _s in _strats[:6]
        )
        if _cnt > 6:
            _chips += (f'<span style="background:{color}22;color:{color};'
                       f'padding:1px 6px;border-radius:8px;margin:1px 2px;font-size:9px">'
                       f'+{_cnt-6} more</span>')
        cells += (
            f'<div style="flex:1;background:{_tf_bg[_btf]}20;border:1px solid {color}33;'
            f'border-radius:6px;padding:6px 8px;min-width:0">'
            f'<div style="display:flex;align-items:center;gap:6px;margin-bottom:4px">'
            f'<span style="background:{color};color:#fff;padding:1px 8px;border-radius:10px;'
            f'font-size:10px;font-weight:700">{_cnt}</span>'
            f'<span style="font-size:9px;font-weight:700;color:{color}">{_tf_labels[_btf]}</span>'
            f'</div><div style="display:flex;flex-wrap:wrap;gap:2px">{_chips}</div></div>'
        )
    return (
        f'<div style="background:linear-gradient(135deg,{color}11,{color}22);'
        f'border:2px solid {color}55;border-radius:10px;padding:10px 12px;'
        f'margin-bottom:12px;font-family:JetBrains Mono">'
        f'<div style="font-size:11px;font-weight:700;color:{color};margin-bottom:8px">'
        f'{label} — Strategies active across all timeframes</div>'
        f'<div style="display:flex;gap:8px;flex-wrap:wrap">{cells}</div></div>'
    )


# ── 16-column P&L table renderer ───────────────────────────────────────────────
def _render_pnl_table(trades, key_prefix="ind"):
    if not trades:
        st.info("No trades to display.")
        return
    _sorted = sorted(trades, key=lambda t: (t.get('Date',''), t.get('Entry Time','')), reverse=True)
    rows_html = ""
    for t in _sorted[:500]:
        net   = _trade_net_pnl(t)
        gross = _trade_gross_pnl(t)
        chg   = _trade_charges(t)
        pct   = _trade_pnl_pct(t)
        ent   = _trade_entry(t)
        ext   = _trade_exit(t)
        sym   = t.get('Symbol', '—')
        seg   = t.get('Segment', '—')
        tf    = _trade_tf(t)
        d     = t.get('Direction', '—')
        strat = t.get('Strategy', '—')
        reason= t.get('Exit Reason', '—')
        entry_t = t.get('Entry Time', '—')
        exit_t  = t.get('Exit Time', '—')
        date_s  = t.get('Date', '—')
        tv_url  = _tv_link(sym)
        win = net >= 0
        bg  = "#f0fdf4" if win else "#fff1f2"
        nc  = "#166534" if win else "#991b1b"
        dc  = "#1d4ed8" if d == "LONG" else "#dc2626"
        try: date_lbl = datetime.strptime(date_s, '%Y-%m-%d').strftime('%d %b')
        except: date_lbl = date_s
        rows_html += (
            f'<tr style="background:{bg};border-bottom:1px solid #e0e9ff;font-size:10px">'
            f'<td style="padding:5px 8px;color:#1e293b;white-space:nowrap;font-weight:600">{date_lbl}</td>'
            f'<td style="padding:5px 8px;color:#5a72a0">{entry_t}</td>'
            f'<td style="padding:5px 8px;color:#5a72a0">{exit_t}</td>'
            f'<td style="padding:5px 8px;color:#1d4ed8;font-weight:700">{sym}</td>'
            f'<td style="padding:5px 8px;color:#5a72a0">{seg}</td>'
            f'<td style="padding:5px 8px;color:#5a72a0">{tf}</td>'
            f'<td style="padding:5px 8px;color:{dc};font-weight:700">{d}</td>'
            f'<td style="padding:5px 8px;color:#1e293b">{strat}</td>'
            f'<td style="padding:5px 8px;font-family:JetBrains Mono">{"₹{:,.0f}".format(ent) if ent else "—"}</td>'
            f'<td style="padding:5px 8px;font-family:JetBrains Mono">{"₹{:,.0f}".format(ext) if ext else "—"}</td>'
            f'<td style="padding:5px 8px;color:{nc};font-weight:600">{pct:+.2f}%</td>'
            f'<td style="padding:5px 8px;color:{nc};font-weight:600">₹{gross:+,.0f}</td>'
            f'<td style="padding:5px 8px;color:#64748b">₹{chg:,.0f}</td>'
            f'<td style="padding:5px 8px;color:{nc};font-weight:800;font-size:11px">₹{net:+,.0f}</td>'
            f'<td style="padding:5px 8px;color:#5a72a0;font-size:9px">{reason}</td>'
            f'<td style="padding:5px 8px"><a href="{tv_url}" target="_blank" style="color:#1d4ed8;font-size:9px;text-decoration:none">📈 TV</a></td>'
            f'</tr>'
        )
    st.markdown(
        '<div style="overflow-x:auto">'
        '<table style="width:100%;border-collapse:collapse;font-family:JetBrains Mono">'
        '<thead><tr style="background:#dde9ff;border-bottom:2px solid #1d4ed8">'
        '<th style="padding:6px 8px;color:#1d4ed8;text-align:left;font-size:10px;white-space:nowrap">DATE</th>'
        '<th style="padding:6px 8px;color:#1d4ed8;text-align:left;font-size:10px;white-space:nowrap">ENTRY</th>'
        '<th style="padding:6px 8px;color:#1d4ed8;text-align:left;font-size:10px;white-space:nowrap">EXIT</th>'
        '<th style="padding:6px 8px;color:#1d4ed8;text-align:left;font-size:10px">SYMBOL</th>'
        '<th style="padding:6px 8px;color:#1d4ed8;text-align:left;font-size:10px">SEGMENT</th>'
        '<th style="padding:6px 8px;color:#1d4ed8;text-align:left;font-size:10px">TF</th>'
        '<th style="padding:6px 8px;color:#1d4ed8;text-align:left;font-size:10px">DIR</th>'
        '<th style="padding:6px 8px;color:#1d4ed8;text-align:left;font-size:10px">STRATEGY</th>'
        '<th style="padding:6px 8px;color:#1d4ed8;text-align:left;font-size:10px">ENTRY ₹</th>'
        '<th style="padding:6px 8px;color:#1d4ed8;text-align:left;font-size:10px">EXIT ₹</th>'
        '<th style="padding:6px 8px;color:#1d4ed8;text-align:left;font-size:10px">PNL%</th>'
        '<th style="padding:6px 8px;color:#1d4ed8;text-align:left;font-size:10px">GROSS ₹</th>'
        '<th style="padding:6px 8px;color:#1d4ed8;text-align:left;font-size:10px">CHARGES</th>'
        '<th style="padding:6px 8px;color:#1d4ed8;text-align:left;font-size:10px;font-weight:800">NET P/L</th>'
        '<th style="padding:6px 8px;color:#1d4ed8;text-align:left;font-size:10px">EXIT REASON</th>'
        '<th style="padding:6px 8px;color:#1d4ed8;text-align:left;font-size:10px">CHART</th>'
        '</tr></thead><tbody>' + rows_html + '</tbody></table></div>',
        unsafe_allow_html=True)

def _render_daily_summary(trades, key_prefix="ind_ds"):
    if not trades:
        st.info("No trades to summarise.")
        return
    daily = defaultdict(lambda: {"trades":0,"invested":0.0,"gross":0.0,"charges":0.0,"net":0.0})
    for t in trades:
        d    = t.get('Date','-')
        ent  = _trade_entry(t)
        net  = _trade_net_pnl(t)
        chg  = _trade_charges(t)
        gross= _trade_gross_pnl(t)
        seg  = t.get('Segment','')
        seg_lo = seg.lower()
        qty  = float(t.get('qty', t.get('lot_size', NSE_LOT_SIZE.get(t.get('Symbol','NIFTY'), 75))) or 75)
        inv  = ent * qty
        daily[d]["trades"]   += 1
        daily[d]["invested"] += inv
        daily[d]["gross"]    += gross
        daily[d]["charges"]  += chg
        daily[d]["net"]      += net

    sorted_days = sorted(daily.items(), key=lambda x: x[0], reverse=True)
    gt_inv  = sum(v["invested"] for v in daily.values())
    gt_net  = sum(v["net"]      for v in daily.values())
    gt_chg  = sum(v["charges"]  for v in daily.values())
    gt_ret  = (gt_net / gt_inv * 100) if gt_inv > 0 else 0.0
    gt_gross= sum(v["gross"]    for v in daily.values())

    st.markdown(
        f'<div style="background:#fef9c3;border:2px solid #eab308;border-radius:8px;'
        f'padding:10px 16px;margin-bottom:10px;font-family:JetBrains Mono;'
        f'display:flex;gap:24px;flex-wrap:wrap">'
        f'<div><div style="font-size:9px;color:#78350f;font-weight:700">TOTAL INVESTED</div>'
        f'<div style="font-size:20px;font-weight:700;color:#1e293b">₹{gt_inv:,.0f}</div></div>'
        f'<div><div style="font-size:9px;color:#78350f;font-weight:700">GROSS P&L</div>'
        f'<div style="font-size:20px;font-weight:700;color:{"#166534" if gt_gross>=0 else "#991b1b"}">₹{gt_gross:+,.0f}</div></div>'
        f'<div><div style="font-size:9px;color:#78350f;font-weight:700">TOTAL CHARGES</div>'
        f'<div style="font-size:20px;font-weight:700;color:#64748b">₹{gt_chg:,.0f}</div></div>'
        f'<div><div style="font-size:9px;color:#78350f;font-weight:700">NET P&L</div>'
        f'<div style="font-size:20px;font-weight:700;color:{"#166534" if gt_net>=0 else "#991b1b"}">₹{gt_net:+,.0f}</div></div>'
        f'<div><div style="font-size:9px;color:#78350f;font-weight:700">NET RETURN</div>'
        f'<div style="font-size:20px;font-weight:700;color:{"#166534" if gt_ret>=0 else "#991b1b"}">{gt_ret:+.3f}%</div></div>'
        f'</div>',
        unsafe_allow_html=True)

    rows_html = ""
    for date_s, dv in sorted_days:
        try: dlbl = datetime.strptime(date_s,'%Y-%m-%d').strftime('%d %b %Y')
        except: dlbl = date_s
        ret  = (dv["net"] / dv["invested"] * 100) if dv["invested"] > 0 else 0.0
        win  = dv["net"] >= 0
        rbg  = "#f0fdf4" if win else "#fff1f2"
        rc   = "#166534" if win else "#991b1b"
        rows_html += (
            f'<tr style="background:{rbg};border-bottom:1px solid #fde68a">'
            f'<td style="padding:6px 10px;font-weight:700;color:#78350f;white-space:nowrap">{dlbl}</td>'
            f'<td style="padding:6px 10px;text-align:center;color:#1e293b">{dv["trades"]}</td>'
            f'<td style="padding:6px 10px;color:#1d4ed8;font-weight:600">₹{dv["invested"]:,.0f}</td>'
            f'<td style="padding:6px 10px;color:{rc};font-weight:600">₹{dv["gross"]:+,.0f}</td>'
            f'<td style="padding:6px 10px;color:#64748b">₹{dv["charges"]:,.2f}</td>'
            f'<td style="padding:6px 10px;color:{rc};font-weight:800;font-size:13px">₹{dv["net"]:+,.0f}</td>'
            f'<td style="padding:6px 10px;color:{rc};font-weight:700">{ret:+.3f}%</td>'
            f'</tr>'
        )
    st.markdown(
        '<div style="overflow-x:auto">'
        '<table style="width:100%;border-collapse:collapse;font-family:JetBrains Mono;font-size:11px">'
        '<thead><tr style="background:#fef08a;border-bottom:2px solid #d97706">'
        '<th style="padding:7px 10px;color:#78350f;text-align:left;font-size:10px">DATE</th>'
        '<th style="padding:7px 10px;color:#78350f;text-align:center;font-size:10px">TRADES</th>'
        '<th style="padding:7px 10px;color:#1d4ed8;text-align:left;font-size:10px">INVESTED ₹</th>'
        '<th style="padding:7px 10px;color:#78350f;text-align:left;font-size:10px">GROSS P&L</th>'
        '<th style="padding:7px 10px;color:#64748b;text-align:left;font-size:10px">CHARGES</th>'
        '<th style="padding:7px 10px;color:#78350f;text-align:left;font-size:10px;font-weight:800">NET P/L</th>'
        '<th style="padding:7px 10px;color:#78350f;text-align:left;font-size:10px">RETURN %</th>'
        '</tr></thead><tbody>' + rows_html + '</tbody></table></div>',
        unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# SIDEBAR — Angel One Connection
# ─────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    # ── Bot count for title badge ─────────────────────────────────────────────
    _n_eq_bots  = len(st.session_state.get('ind_eq_bots',  {}))
    _n_fut_bots = len(st.session_state.get('ind_fut_bots', {}))
    _n_opt_bots = len(st.session_state.get('ind_opt_bots', {}))
    _n_all_bots = _n_eq_bots + _n_fut_bots + _n_opt_bots

    # ── Index ticker cards (Nifty / BankNifty / Sensex / VIX) ────────────────
    _sb_idx_syms   = ["NIFTY","BANKNIFTY","SENSEX","INDIAVIX"]
    _sb_idx_labels = {"NIFTY":"Nifty 50","BANKNIFTY":"Bank Nifty","SENSEX":"Sensex","INDIAVIX":"India VIX"}
    _sb_idx_prices = _fetch_nse_ticker_prices(tuple(_sb_idx_syms))

    _sb_idx_html = '<div style="display:flex;gap:6px;flex-wrap:nowrap;overflow-x:auto;padding:4px 0;margin-bottom:4px">'
    for _si in _sb_idx_syms:
        _sip, _sic = _sb_idx_prices.get(_si, (0.0, 0.0))
        _sicc = "#16a34a" if _sic >= 0 else "#dc2626"
        _sibg = "#f0fdf4" if _sic >= 0 else "#fff1f2"
        _siarr = "+" if _sic >= 0 else ""
        _silbl = _sb_idx_labels.get(_si, _si)
        _sb_idx_html += (
            f'<div style="background:{_sibg};border:1px solid {_sicc}44;border-radius:8px;'
            f'padding:4px 8px;white-space:nowrap;font-family:JetBrains Mono;min-width:90px;flex-shrink:0">'
            f'<div style="font-size:8px;color:#64748b;font-weight:600">{_silbl}</div>'
            f'<div style="font-size:12px;font-weight:700;color:#1e293b">'
            f'{"" if _sip == 0 else ("Rs{:,.0f}".format(_sip) if _si != "INDIAVIX" else "{:.2f}".format(_sip))}</div>'
            f'<div style="font-size:9px;font-weight:700;color:{_sicc}">{_siarr}{_sic:.2f}%</div>'
            f'</div>'
        )
    _sb_idx_html += '</div>'
    st.markdown(_sb_idx_html, unsafe_allow_html=True)

    # ── Scrolling marquee: SCANNING Nifty 50 ─────────────────────────────────
    _sb_eq_prices  = _fetch_nse_ticker_prices(tuple(NSE_TOP_STOCKS))
    _sb_mq_items   = []
    for _smq in NSE_TOP_STOCKS:
        _smqp, _smqc = _sb_eq_prices.get(_smq, (0.0, 0.0))
        _smqcc = "#16a34a" if _smqc >= 0 else "#dc2626"
        _smqarr = "+" if _smqc >= 0 else ""
        _smqprice = "Rs{:,.0f}".format(_smqp) if _smqp > 0 else "--"
        _sb_mq_items.append(
            f'<span style="margin:0 14px;white-space:nowrap">'
            f'<span style="color:#60a5fa;font-weight:700">{_smq}</span>'
            f'&nbsp;<span style="color:#e2e8f0">{_smqprice}</span>'
            f'&nbsp;<span style="color:{_smqcc};font-size:10px">{_smqarr}{_smqc:.2f}%</span>'
            f'</span><span style="color:#475569;margin:0 2px">|</span>'
        )
    _sb_mq_content = ''.join(_sb_mq_items)
    _sb_mq_full    = _sb_mq_content + _sb_mq_content
    _sb_scroll_dur = max(60, len(NSE_TOP_STOCKS) * 3)
    st.markdown(
        f'<div style="background:#1e293b;border-radius:8px;padding:5px 10px;'
        f'margin-bottom:8px;overflow:hidden">'
        f'<div style="display:flex;align-items:center;gap:6px">'
        f'<span style="background:#1d4ed8;color:#fff;font-size:8px;font-weight:700;'
        f'padding:2px 6px;border-radius:4px;white-space:nowrap;flex-shrink:0;'
        f'font-family:JetBrains Mono">SCANNING {len(NSE_TOP_STOCKS)} STOCKS</span>'
        f'<div style="overflow:hidden;flex:1">'
        f'<div style="display:inline-flex;animation:marquee_sb {_sb_scroll_dur}s linear infinite;'
        f'font-family:JetBrains Mono;font-size:10px">'
        f'{_sb_mq_full}'
        f'</div></div></div></div>'
        f'<style>@keyframes marquee_sb{{0%{{transform:translateX(0)}}100%{{transform:translateX(-50%)}}}}</style>',
        unsafe_allow_html=True)

    # ── Heading + live badge ──────────────────────────────────────────────────
    if _n_all_bots > 0:
        _live_badge = f'  <span style="background:#16a34a;color:#fff;border-radius:10px;padding:2px 8px;font-size:11px"> {_n_all_bots} LIVE * RUNNING</span>'
    else:
        _live_badge = ""
    st.markdown(f'<div style="color:#1d4ed8;font-size:20px;font-weight:700;font-family:JetBrains Mono">INDIA TRADER {_live_badge}</div>', unsafe_allow_html=True)
    st.markdown('<div style="color:#64748b;font-size:11px;margin-bottom:8px">India Trader | PPT * EMA * Bollinger * VWAP * 4 Timeframes</div>', unsafe_allow_html=True)

    # ── Market status ─────────────────────────────────────────────────────────
    _ist_now2 = _ist_now()
    _mkt_open = _market_open()
    _ist_ts   = _ist_now2.strftime('%d %b %H:%M IST')
    _mkt_c    = "#166534" if _mkt_open else "#d97706"
    _mkt_lbl  = "MARKET OPEN" if _mkt_open else ("PRE-OPEN" if _ist_now2.time() < __import__('datetime').time(9,15) else "AFTER-HOURS")
    st.markdown(
        f'<div style="background:#ffffff;border:1px solid #dde9ff;border-radius:8px;'
        f'padding:8px 12px;margin-bottom:10px;font-family:JetBrains Mono">'
        f'<span style="color:{_mkt_c};font-size:12px;font-weight:700"> - {_mkt_lbl}</span>'
        f'<span style="color:#64748b;font-size:10px;float:right">{_ist_ts}</span></div>',
        unsafe_allow_html=True)

    # ── Angel One connection ──────────────────────────────────────────────────
    _ao_connected = st.session_state.angel_connected
    _ao_c   = "#166534" if _ao_connected else "#dc2626"
    _ao_lbl = "CONNECTED" if _ao_connected else "DISCONNECTED"
    st.markdown(f'<div style="font-family:JetBrains Mono;font-size:11px;color:{_ao_c};margin-bottom:4px">Angel One: {_ao_lbl}</div>', unsafe_allow_html=True)

    if not _ao_connected:
        if st.button("Connect Angel One", type="primary", width='stretch'):
            try:
                _c = AngelOneClient(
                    api_key=os.getenv('ANGEL_API_KEY',''),
                    client_id=os.getenv('ANGEL_CLIENT_ID',''),
                    mpin=os.getenv('ANGEL_MPIN',''),
                    totp_key=os.getenv('ANGEL_TOTP_KEY','')
                )
                _ok, _msg = _c.connect()
                if _ok:
                    st.session_state.angel_client    = _c
                    st.session_state.angel_connected = True
                    st.session_state.angel_funds     = _c.get_funds()
                    st.success("Connected!")
                    st.rerun()
                else:
                    st.error(f"Failed: {_msg}")
            except Exception as _ae:
                st.error(str(_ae))
    else:
        if st.button("Disconnect", width='stretch'):
            st.session_state.angel_client    = None
            st.session_state.angel_connected = False
            st.session_state.angel_funds     = {}
            st.rerun()
        _funds = st.session_state.angel_funds
        if _funds:
            st.markdown(f'<div style="font-family:JetBrains Mono;font-size:11px;color:#5a72a0">Available: Rs{float(_funds.get("availablecash",0)):,.0f}</div>', unsafe_allow_html=True)

    # ── Bot status card ───────────────────────────────────────────────────────
    if _n_all_bots > 0:
        _eq_alive  = sum(1 for b in st.session_state.get('ind_eq_bots',{}).values()  if b is not None and getattr(b,'_running',False))
        _fut_alive = sum(1 for b in st.session_state.get('ind_fut_bots',{}).values() if b is not None and getattr(b,'_running',False))
        _opt_alive = sum(1 for b in st.session_state.get('ind_opt_bots',{}).values() if b is not None and getattr(b,'_running',False))
        _total_alive = _eq_alive + _fut_alive + _opt_alive
        def _running_tfs(bot_dict, key_sep="_"):
            _tfs = sorted(set(
                k.split(key_sep)[-1]
                for k, b in bot_dict.items()
                if b is not None and getattr(b, '_running', False)
            ), key=lambda t: ["5M","15M","1H","4H"].index(t) if t in ["5M","15M","1H","4H"] else 99)
            return " + ".join(t.lower() for t in _tfs) if _tfs else "none"
        _eq_tf_str  = _running_tfs(st.session_state.get('ind_eq_bots', {}))
        _fut_tf_str = _running_tfs(st.session_state.get('ind_fut_bots', {}))
        _opt_tf_str = _running_tfs(st.session_state.get('ind_opt_bots', {}))
        st.markdown(
            f'<div style="background:#f0fdf4;border:2px solid #16a34a;border-radius:10px;'
            f'padding:10px 14px;margin-bottom:8px;font-family:JetBrains Mono;font-size:11px">'
            f'<div style="font-size:13px;font-weight:700;color:#15803d"> LIVE AND RUNNING  *  {_total_alive} active threads</div>'
            f'<div style="margin-top:6px;display:flex;gap:6px;flex-wrap:wrap">'
            f'<span style="background:#dcfce7;color:#166534;padding:2px 8px;border-radius:8px;font-weight:700"> Equity (Nifty 50): {_eq_alive} bots ({_eq_tf_str})</span>'
            f'<span style="background:#dcfce7;color:#166534;padding:2px 8px;border-radius:8px;font-weight:700"> Futures: {_fut_alive} bots ({_fut_tf_str})</span>'
            f'<span style="background:#dcfce7;color:#166534;padding:2px 8px;border-radius:8px;font-weight:700"> Options: {_opt_alive} bots ({_opt_tf_str})</span>'
            f'</div>'
            f'<div style="margin-top:5px;color:#64748b;font-size:10px">Each bot = 1 symbol × 1 timeframe · scanning all strategies every interval</div>'
            f'</div>', unsafe_allow_html=True)
    else:
        st.markdown(
            '<div style="background:#f8faff;border:2px solid #e2e8f0;border-radius:10px;'
            'padding:10px 14px;margin-bottom:8px;font-family:JetBrains Mono;font-size:11px">'
            '<div style="font-size:13px;font-weight:700;color:#dc2626"> NOT RUNNING</div>'
            '<div style="color:#64748b;margin-top:4px">Click START below to launch all bots.</div>'
            '<div style="color:#64748b;font-size:10px;margin-top:4px">Auto-starts at 9:15 AM IST if app is open</div>'
            '</div>', unsafe_allow_html=True)

    # ── Per-segment strategy map (display name → bot code) ────────────────────
    _SB_STRAT_MAP = {
        "ORB Breakout":    "ORB",
        "5 EMA Cross":     "5EMA",
        "VWAP Bounce":     "VWAP_TREND",
        "Gap & Go":        "GAP_AND_GO",
        "Gap Fill":        "GAP_FILL",
        "Premarket Break": "PREMARKET_BREAK",
        "Momentum Scalp":  "MOM_SCALP",
        "HOD Break":       "HOD_BREAK",
        "Tail Reversal":   "TAIL_REVERSAL",
        "BB Squeeze":      "BB_SQUEEZE",
        "Inside Candle":   "INSIDE",
        "Support Bounce":  "SUPPORT_BOUNCE",
        "Mean Reversion":  "MEAN_REV",
        "Supertrend":      "SUPERTREND",
        "RSI Divergence":  "RSI_DIV",
        "MACD Reversal":   "MACD_REV",
        "Bull Flag":       "BULL_FLAG",
        "Breakout Mom":    "BREAKOUT_MOM",
        "Trend Follow":    "TREND_FOLLOW",
        "Fib Cluster":     "FIB_CLUSTER",
        "EMA 50 Support":  "EMA50_SUPPORT",
    }
    _SB_ALL_STRATS = list(_SB_STRAT_MAP.keys())
    _SB_DEFAULT_EQ  = ["ORB Breakout","5 EMA Cross","VWAP Bounce","Premarket Break","Gap & Go"]
    _SB_DEFAULT_FUT = ["ORB Breakout","5 EMA Cross","VWAP Bounce","Momentum Scalp","HOD Break"]
    _SB_DEFAULT_OPT = ["ORB Breakout","5 EMA Cross","VWAP Bounce","Gap & Go","Premarket Break"]

    st.markdown("---")
    st.markdown('<div style="color:#1d4ed8;font-size:13px;font-weight:700;font-family:JetBrains Mono;margin-bottom:8px"> SEGMENT CONTROLS</div>', unsafe_allow_html=True)

    # ── EQUITY ────────────────────────────────────────────────────────────────
    _sb_eq_hdr = st.columns([3, 2])
    with _sb_eq_hdr[0]:
        _sb_eq_on = st.toggle("Equity", value=st.session_state.get('sb_eq_on', True), key="sb_eq_on")
    with _sb_eq_hdr[1]:
        _sb_eq_live = st.toggle("LIVE", value=st.session_state.get('sb_eq_live', False), key="sb_eq_live",
                                 disabled=not _sb_eq_on, help="OFF = Paper trade")
    if _sb_eq_on:
        _sb_eq_mode_c = "#dc2626" if _sb_eq_live else "#166534"
        _sb_eq_mode_l = "⚡ LIVE ORDERS" if _sb_eq_live else "📋 PAPER"
        st.markdown(f'<div style="font-family:JetBrains Mono;font-size:9px;color:{_sb_eq_mode_c};margin:-6px 0 4px 0;font-weight:700">{_sb_eq_mode_l}</div>', unsafe_allow_html=True)
        _sb_eq_tfs = st.multiselect("TF", ["5M","15M","1H","4H"],
            default=st.session_state.get('sb_eq_tfs_val', ["15M","1H"]),
            key="sb_eq_tfs_val", label_visibility="collapsed",
            placeholder="Pick timeframes…")
        _sb_eq_strats = st.multiselect("Strategies",_SB_ALL_STRATS,
            default=st.session_state.get('sb_eq_strats_val', _SB_DEFAULT_EQ),
            key="sb_eq_strats_val", label_visibility="collapsed",
            placeholder="Pick strategies…")
    else:
        _sb_eq_tfs = []; _sb_eq_strats = []
        st.markdown('<div style="font-family:JetBrains Mono;font-size:9px;color:#94a3b8;margin:-4px 0 4px 0">Equity bots disabled</div>', unsafe_allow_html=True)

    st.markdown('<div style="border-top:1px solid #e2e8f0;margin:6px 0"></div>', unsafe_allow_html=True)

    # ── FUTURES ───────────────────────────────────────────────────────────────
    _sb_fut_hdr = st.columns([3, 2])
    with _sb_fut_hdr[0]:
        _sb_fut_on = st.toggle("Futures", value=st.session_state.get('sb_fut_on', True), key="sb_fut_on")
    with _sb_fut_hdr[1]:
        _sb_fut_live = st.toggle("LIVE", value=st.session_state.get('sb_fut_live', False), key="sb_fut_live",
                                  disabled=not _sb_fut_on, help="OFF = Paper trade")
    if _sb_fut_on:
        _sb_fut_mode_c = "#dc2626" if _sb_fut_live else "#166534"
        _sb_fut_mode_l = "⚡ LIVE ORDERS" if _sb_fut_live else "📋 PAPER"
        st.markdown(f'<div style="font-family:JetBrains Mono;font-size:9px;color:{_sb_fut_mode_c};margin:-6px 0 4px 0;font-weight:700">{_sb_fut_mode_l}</div>', unsafe_allow_html=True)
        _sb_fut_tfs = st.multiselect("TF", ["5M","15M","1H","4H"],
            default=st.session_state.get('sb_fut_tfs_val', ["15M","1H"]),
            key="sb_fut_tfs_val", label_visibility="collapsed",
            placeholder="Pick timeframes…")
        _sb_fut_strats = st.multiselect("Strategies", _SB_ALL_STRATS,
            default=st.session_state.get('sb_fut_strats_val', _SB_DEFAULT_FUT),
            key="sb_fut_strats_val", label_visibility="collapsed",
            placeholder="Pick strategies…")
    else:
        _sb_fut_tfs = []; _sb_fut_strats = []
        st.markdown('<div style="font-family:JetBrains Mono;font-size:9px;color:#94a3b8;margin:-4px 0 4px 0">Futures bots disabled</div>', unsafe_allow_html=True)

    st.markdown('<div style="border-top:1px solid #e2e8f0;margin:6px 0"></div>', unsafe_allow_html=True)

    # ── OPTIONS (always paper) ────────────────────────────────────────────────
    _sb_opt_on = st.toggle("Options (Paper only)", value=st.session_state.get('sb_opt_on', False), key="sb_opt_on")
    st.markdown('<div style="font-family:JetBrains Mono;font-size:9px;color:#d97706;margin:-6px 0 4px 0;font-weight:700">📋 PAPER ONLY — buys CE/PE on signal</div>', unsafe_allow_html=True)
    if _sb_opt_on:
        _sb_opt_tfs = st.multiselect("TF", ["5M","15M","1H","4H"],
            default=st.session_state.get('sb_opt_tfs_val', ["5M","15M"]),
            key="sb_opt_tfs_val", label_visibility="collapsed",
            placeholder="Pick timeframes…")
        _sb_opt_strats = st.multiselect("Strategies", _SB_ALL_STRATS,
            default=st.session_state.get('sb_opt_strats_val', _SB_DEFAULT_OPT),
            key="sb_opt_strats_val", label_visibility="collapsed",
            placeholder="Pick strategies…")
    else:
        _sb_opt_tfs = []; _sb_opt_strats = []

    st.markdown('<div style="border-top:1px solid #e2e8f0;margin:6px 0"></div>', unsafe_allow_html=True)

    # ── OPT/FUT STRATEGIES ENGINE (Chandelier / TUX / CE_REGIME) ─────────────
    _sb_oat_on = st.toggle("OPT/FUT Strategies Engine",
                            value=st.session_state.get('sb_oat_on', True), key="sb_oat_on")
    st.markdown('<div style="font-family:JetBrains Mono;font-size:9px;color:#7c3aed;'
                'margin:-6px 0 4px 0;font-weight:700">'
                '⚡ CHANDELIER · TUX+ST · CE_REGIME — All 3 Indices</div>',
                unsafe_allow_html=True)
    if _sb_oat_on and OFS_OK:
        _sb_oat_mode  = st.selectbox("Mode", ["both", "options", "futures"],
                                      index=["both", "options", "futures"].index(
                                          st.session_state.get('oat_mode', 'both')),
                                      key="oat_mode")
        _sb_oat_paper = st.toggle("Paper Mode",
                                   value=st.session_state.get('oat_paper', True), key="oat_paper")
        _sb_oat_qty   = st.number_input("Lots", 1, 10,
                                         int(st.session_state.get('oat_qty', 1)), key="oat_qty")
    else:
        _sb_oat_mode  = st.session_state.get('oat_mode', 'both')
        _sb_oat_paper = st.session_state.get('oat_paper', True)
        _sb_oat_qty   = int(st.session_state.get('oat_qty', 1))

    st.markdown("---")

    # ── Auto-start at 9:00 AM IST (once per day, if laptop is on) ────────────
    _today_key = f"_daily_autostart_{_ist_now2.strftime('%Y-%m-%d')}"
    _auto_t = _ist_now2.time()
    if (time(9, 0) <= _auto_t <= time(9, 10) and
            not st.session_state.get(_today_key, False) and
            _n_all_bots == 0 and
            not st.session_state.get('_oat_user_stopped', False)):
        st.session_state[_today_key] = True
        # Equity bots
        if _sb_eq_on and _sb_eq_tfs and POS_OK:
            try:
                from nse_stocks_engine import NIFTY50_STOCKS as _N50s
            except Exception:
                _N50s = {}
            _eq_codes = [_SB_STRAT_MAP[s] for s in _sb_eq_strats if s in _SB_STRAT_MAP] or None
            _ebd = dict(st.session_state.get('ind_eq_bots', {}))
            for _esym in _N50s:
                for _etf in _sb_eq_tfs:
                    _ek = f"{_esym}_{_etf}"
                    if _ek not in _ebd or not getattr(_ebd.get(_ek), '_running', False):
                        try:
                            _eb = POSAutoTrader(instrument=_esym, timeframe=_etf,
                                qty=1, strategies=_eq_codes, paper_trade=not _sb_eq_live)
                            _eb.start(); _ebd[_ek] = _eb
                        except Exception: pass
            st.session_state.ind_eq_bots = _ebd
        # Futures bots
        if _sb_fut_on and _sb_fut_tfs and POS_OK:
            _fut_codes = [_SB_STRAT_MAP[s] for s in _sb_fut_strats if s in _SB_STRAT_MAP] or None
            _fbd = dict(st.session_state.get('ind_fut_bots', {}))
            for _fsym in NSE_INDICES:
                for _ftf in _sb_fut_tfs:
                    _fk = f"{_fsym}_FUT_{_ftf}"
                    if _fk not in _fbd or not getattr(_fbd.get(_fk), '_running', False):
                        try:
                            _fb = POSAutoTrader(instrument=_fsym, timeframe=_ftf,
                                qty=NSE_LOT_SIZE.get(_fsym, 75), strategies=_fut_codes,
                                paper_trade=not _sb_fut_live)
                            _fb.start(); _fbd[_fk] = _fb
                        except Exception: pass
            st.session_state.ind_fut_bots = _fbd
        # Options bots
        if _sb_opt_on and _sb_opt_tfs and OPT_OK:
            _opt_codes = [_SB_STRAT_MAP[s] for s in _sb_opt_strats if s in _SB_STRAT_MAP] or None
            _obd = dict(st.session_state.get('ind_opt_bots', {}))
            for _osym in NSE_INDICES:
                for _otf in _sb_opt_tfs:
                    _ok2 = f"{_osym}_{_otf}"
                    if _ok2 not in _obd or not getattr(_obd.get(_ok2), '_running', False):
                        try:
                            _ob = OptionsAutoTrader(instrument=_osym, timeframe=_otf,
                                qty=1, paper_trade=True, strategies=_opt_codes)
                            _ob.start(); _obd[_ok2] = _ob
                        except Exception: pass
            st.session_state.ind_opt_bots = _obd
        # OPT/FUT Strategies Engine
        if _sb_oat_on and OFS_OK:
            _oat_key_auto = "_oat_trader_ALL"
            if not getattr(st.session_state.get(_oat_key_auto), '_running', False):
                try:
                    _oat_tfs_auto = st.session_state.get('oat_tfs', ["5m","15m","1h"]) or ["5m","15m","1h"]
                    _oat_strats_auto = st.session_state.get('oat_strats', _OFS_STRATEGIES) or _OFS_STRATEGIES
                    _oat_tr_auto = _OFSAutoTrader(
                        instruments=["NIFTY", "BANKNIFTY", "SENSEX"],
                        strategies=_oat_strats_auto, timeframes=_oat_tfs_auto,
                        mode=_sb_oat_mode, qty=_sb_oat_qty, paper=_sb_oat_paper, rr_mult=2.0,
                        angel_client=st.session_state.get('angel_client'))
                    _oat_tr_auto.start()
                    st.session_state[_oat_key_auto] = _oat_tr_auto
                    st.session_state['_oat_auto_started'] = True
                    st.session_state['_oat_user_stopped'] = False
                except Exception: pass
        # Run unified OPT/Fut signal scan
        if OFS_OK:
            try:
                _auto_sigs = {_idx: _ofs_scan_all(_idx, ["5m","15m","1h"]) for _idx in NSE_INDICES}
                st.session_state["ofs_all_signals"] = _auto_sigs
                st.session_state["ofs_scan_time"]   = _ist_now2.strftime('%H:%M:%S IST')
                st.session_state["ofs_all_exp"]      = {_idx: _ofs_all_expiries(_idx) for _idx in NSE_INDICES}
            except Exception: pass
        st.toast("9 AM Auto-start: All bots launched!", icon="🚀")
        st.rerun()

    # ── START / STOP buttons ──────────────────────────────────────────────────
    _sc1, _sc2 = st.columns(2)
    with _sc1:
        if st.button("▶ START ALL BOTS", type="primary", width='stretch', key="sb_start_sel"):
            with st.spinner("Launching selected bots…"):
                # Equity bots
                if _sb_eq_on and _sb_eq_tfs and POS_OK:
                    try:
                        from nse_stocks_engine import NIFTY50_STOCKS as _N50s
                    except Exception:
                        _N50s = {}
                    _eq_codes = [_SB_STRAT_MAP[s] for s in _sb_eq_strats if s in _SB_STRAT_MAP] or None
                    _ebd = dict(st.session_state.get('ind_eq_bots', {}))
                    for _esym in _N50s:
                        for _etf in _sb_eq_tfs:
                            _ek = f"{_esym}_{_etf}"
                            if _ek not in _ebd or not getattr(_ebd.get(_ek), '_running', False):
                                try:
                                    _eb = POSAutoTrader(instrument=_esym, timeframe=_etf,
                                        qty=1, strategies=_eq_codes,
                                        paper_trade=not _sb_eq_live)
                                    _eb.start(); _ebd[_ek] = _eb
                                except Exception: pass
                    st.session_state.ind_eq_bots = _ebd
                # Futures bots
                if _sb_fut_on and _sb_fut_tfs and POS_OK:
                    _fut_codes = [_SB_STRAT_MAP[s] for s in _sb_fut_strats if s in _SB_STRAT_MAP] or None
                    _fbd = dict(st.session_state.get('ind_fut_bots', {}))
                    for _fsym in NSE_INDICES:
                        for _ftf in _sb_fut_tfs:
                            _fk = f"{_fsym}_FUT_{_ftf}"
                            if _fk not in _fbd or not getattr(_fbd.get(_fk), '_running', False):
                                try:
                                    _fb = POSAutoTrader(instrument=_fsym, timeframe=_ftf,
                                        qty=NSE_LOT_SIZE.get(_fsym, 75),
                                        strategies=_fut_codes,
                                        paper_trade=not _sb_fut_live)
                                    _fb.start(); _fbd[_fk] = _fb
                                except Exception: pass
                    st.session_state.ind_fut_bots = _fbd
                # Options bots
                if _sb_opt_on and _sb_opt_tfs and OPT_OK:
                    _opt_codes = [_SB_STRAT_MAP[s] for s in _sb_opt_strats if s in _SB_STRAT_MAP] or None
                    _obd = dict(st.session_state.get('ind_opt_bots', {}))
                    for _osym in NSE_INDICES:
                        for _otf in _sb_opt_tfs:
                            _ok2 = f"{_osym}_{_otf}"
                            if _ok2 not in _obd or not getattr(_obd.get(_ok2), '_running', False):
                                try:
                                    _ob = OptionsAutoTrader(instrument=_osym, timeframe=_otf,
                                        qty=1, paper_trade=True, strategies=_opt_codes)
                                    _ob.start(); _obd[_ok2] = _ob
                                except Exception: pass
                    st.session_state.ind_opt_bots = _obd
                # OPT/FUT Strategies Engine (OptionsFuturesAutoTrader)
                if _sb_oat_on and OFS_OK:
                    _oat_key_sb = "_oat_trader_ALL"
                    _existing_oat = st.session_state.get(_oat_key_sb)
                    if _existing_oat is None or not getattr(_existing_oat, '_running', False):
                        try:
                            _oat_tfs_sb    = st.session_state.get('oat_tfs', ["5m","15m","1h"]) or ["5m","15m","1h"]
                            _oat_strats_sb = st.session_state.get('oat_strats', _OFS_STRATEGIES) or _OFS_STRATEGIES
                            _oat_tr = _OFSAutoTrader(
                                instruments=["NIFTY", "BANKNIFTY", "SENSEX"],
                                strategies=_oat_strats_sb, timeframes=_oat_tfs_sb,
                                mode=_sb_oat_mode, qty=_sb_oat_qty,
                                paper=_sb_oat_paper, rr_mult=2.0,
                                angel_client=st.session_state.get('angel_client'))
                            _oat_tr.start()
                            st.session_state[_oat_key_sb] = _oat_tr
                        except Exception as _oat_ex:
                            st.warning(f"OPT/FUT engine failed: {_oat_ex}")
                _oat_alive = 1 if (st.session_state.get("_oat_trader_ALL") and
                                    getattr(st.session_state["_oat_trader_ALL"], '_running', False)) else 0
                _n_s = (len(st.session_state.get('ind_eq_bots', {})) +
                        len(st.session_state.get('ind_fut_bots', {})) +
                        len(st.session_state.get('ind_opt_bots', {})))
                # Run unified OPT/Fut signal scan alongside bot start
                if OFS_OK:
                    try:
                        _now_ist = _ist_now()
                        _sigs_sb = {_idx: _ofs_scan_all(_idx, ["5m","15m","1h"]) for _idx in NSE_INDICES}
                        st.session_state["ofs_all_signals"] = _sigs_sb
                        st.session_state["ofs_scan_time"]   = _now_ist.strftime('%H:%M:%S IST')
                        st.session_state["ofs_all_exp"]      = {_idx: _ofs_all_expiries(_idx) for _idx in NSE_INDICES}
                    except Exception: pass
                st.session_state['_oat_user_stopped'] = False
                _oat_suffix = f" + OPT/FUT engine" if _oat_alive else ""
                st.success(f"{_n_s} bots{_oat_suffix} running"); st.rerun()
    with _sc2:
        if st.button("⏹ STOP ALL", width='stretch', key="sb_stop_all"):
            for _bdict_key in ['ind_eq_bots','ind_fut_bots','ind_opt_bots']:
                for _bv in st.session_state.get(_bdict_key, {}).values():
                    try:
                        if _bv and getattr(_bv, '_running', False): _bv.stop()
                    except Exception: pass
            st.session_state.ind_eq_bots  = {}
            st.session_state.ind_fut_bots = {}
            st.session_state.ind_opt_bots = {}
            _oat_obj_sb = st.session_state.get("_oat_trader_ALL")
            if _oat_obj_sb and getattr(_oat_obj_sb, '_running', False):
                try: _oat_obj_sb.stop()
                except Exception: pass
            st.session_state["_oat_trader_ALL"] = None
            st.session_state['_oat_user_stopped'] = True
            st.session_state['_oat_auto_started'] = False
            st.success("All bots stopped."); st.rerun()

    st.markdown("---")
    if st.button("Refresh All Data", type="primary", width='stretch'):
        with st.spinner("Fetching NSE data..."):
            for _sym in NSE_INDICES:
                _ld = fetch_nse_live_data(_sym)
                if _ld and _sym == "NIFTY":
                    st.session_state.spot       = _ld['spot']
                    st.session_state.prev_high  = _ld['prev_high']
                    st.session_state.prev_low   = _ld['prev_low']
                    st.session_state.prev_close = _ld['prev_close']
                    st.session_state.day_high   = _ld.get('day_high', _ld['spot']*1.01)
                    st.session_state.day_low    = _ld.get('day_low',  _ld['spot']*0.99)
        if STRATEGY_ENGINE_OK:
            try:
                cues = GlobalCuesEngine.fetch_global_cues()
                gb, gs, _ = GlobalCuesEngine.compute_global_bias(cues)
                st.session_state.global_bias  = gb
                st.session_state.global_score = gs
                st.session_state.cues         = cues
            except Exception: pass
        st.success(f"Refreshed! Nifty: Rs{st.session_state.spot:,.0f}")
        st.rerun()

    # Auto-connect Angel One if market open and env keys set
    if _mkt_open and not st.session_state.angel_connected and os.getenv('ANGEL_API_KEY'):
        try:
            _ac = AngelOneClient(
                api_key=os.getenv('ANGEL_API_KEY',''),
                client_id=os.getenv('ANGEL_CLIENT_ID',''),
                mpin=os.getenv('ANGEL_MPIN',''),
                totp_key=os.getenv('ANGEL_TOTP_KEY','')
            )
            _aok, _ = _ac.connect()
            if _aok:
                st.session_state.angel_client    = _ac
                st.session_state.angel_connected = True
                st.session_state.angel_funds     = _ac.get_funds()
        except Exception: pass


# ─────────────────────────────────────────────────────────────────────────────
# MAIN TABS
# ─────────────────────────────────────────────────────────────────────────────
_tabs = st.tabs(["Signals", "Auto Trade", "All Strategies",
                 "Opt/Fut Strategies", "Opt/Fut Auto Trade",
                 "P&L Summary", "Settings", "Testing"])

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 0 — SIGNALS
# ═══════════════════════════════════════════════════════════════════════════════
with _tabs[0]:
    st.markdown(f'<div style="color:#1d4ed8;font-size:20px;font-weight:700;font-family:JetBrains Mono">India Signals</div>', unsafe_allow_html=True)
    instrument = st.radio("Index", NSE_INDICES, horizontal=True, key="sig_instrument")

    # Live index cards
    _idx_prices = _fetch_nse_ticker_prices(tuple(NSE_INDICES))
    _sc1, _sc2, _sc3 = st.columns(3)
    for _scol, _ssym in zip([_sc1, _sc2, _sc3], NSE_INDICES):
        _sp, _spc = _idx_prices.get(_ssym, (0.0, 0.0))
        _scc = "#166534" if _spc >= 0 else "#991b1b"
        _sbg = "#f0fdf4" if _spc >= 0 else "#fff1f2"
        _lot = NSE_LOT_SIZE.get(_ssym, 75)
        with _scol:
            st.markdown(
                f'<div class="mcard" style="text-align:center;border-color:{_scc}44">'
                f'<div class="mlbl">{_ssym}</div>'
                f'<div class="mval" style="color:#1d4ed8">{"₹{:,.2f}".format(_sp) if _sp else "—"}</div>'
                f'<div style="color:{_scc};font-size:12px;font-weight:700;font-family:JetBrains Mono">{"+" if _spc>=0 else ""}{_spc:.2f}%</div>'
                f'<div style="color:#5a72a0;font-size:9px;font-family:JetBrains Mono;margin-top:4px">Lot: {_lot} | <a href="{_tv_link(_ssym)}" target="_blank" style="color:#1d4ed8">TradingView</a></div>'
                f'</div>',
                unsafe_allow_html=True)

    st.markdown("---")

    # Signal generator
    if STRATEGY_ENGINE_OK:
        _spot = st.session_state.spot
        _ph   = st.session_state.prev_high
        _pl   = st.session_state.prev_low
        _pc2  = st.session_state.prev_close
        _gb   = st.session_state.global_bias
        _gs   = st.session_state.global_score

        if _spot > 0 and _ph > 0:
            try:
                _sig = SignalGenerator.generate_signal(
                    instrument=instrument,
                    spot_price=_spot,
                    prev_high=_ph, prev_low=_pl,
                    first_candle={'open':_pc2,'high':st.session_state.day_high or _spot*1.01,
                                  'low':st.session_state.day_low or _spot*0.99,'close':_spot,'volume':1000000},
                    oi_signal=st.session_state.oi_signal,
                    global_bias=_gb, global_score=_gs,
                    vix=14.0, is_expiry=False
                )
                st.session_state.signal = _sig
            except Exception: _sig = st.session_state.signal
        else:
            _sig = st.session_state.signal

        if _sig:
            _sig_c  = "#16a34a" if getattr(_sig,'direction','') == "LONG" else "#dc2626"
            _sig_bg = "#f0fdf4" if getattr(_sig,'direction','') == "LONG" else "#fff1f2"
            _sig_lbl= "LONG — BUY CALL" if getattr(_sig,'direction','') == "LONG" else "SHORT — BUY PUT"
            st.markdown(
                f'<div style="background:{_sig_bg};border:2px solid {_sig_c};border-radius:10px;'
                f'padding:18px;font-family:JetBrains Mono;margin-bottom:10px">'
                f'<div style="color:{_sig_c};font-size:18px;font-weight:700">{_sig_lbl} — {instrument}</div>'
                f'<div style="color:#5a72a0;font-size:11px;margin-top:4px">'
                f'Score: {getattr(_sig,"confluence_score",0)}/10 | Bias: {_gb} ({_gs:.1f})</div>'
                f'<hr style="border-color:#e0e9ff;margin:10px 0">'
                f'<div style="display:flex;gap:24px;flex-wrap:wrap">'
                f'<div><div style="color:#5a72a0;font-size:9px">ENTRY</div><div style="font-size:16px;font-weight:700">₹{getattr(_sig,"entry_price",0):,.0f}</div></div>'
                f'<div><div style="color:#5a72a0;font-size:9px">TARGET 1</div><div style="font-size:16px;font-weight:700;color:#166534">₹{getattr(_sig,"target_1",0):,.0f}</div></div>'
                f'<div><div style="color:#5a72a0;font-size:9px">STOP LOSS</div><div style="font-size:16px;font-weight:700;color:#dc2626">₹{getattr(_sig,"stop_loss",0):,.0f}</div></div>'
                f'</div>'
                f'<div style="color:#5a72a0;font-size:11px;margin-top:8px">{getattr(_sig,"reason","")[:120]}</div>'
                f'</div>',
                unsafe_allow_html=True)
        else:
            st.info("No signal yet — click 'Refresh All Data' in the sidebar to fetch live data.")
    else:
        st.info("Strategy engine not available. Check strategy_engine.py.")

    st.markdown("---")

    # Global cues summary
    _cues = st.session_state.get('cues')
    _col_gc1, _col_gc2 = st.columns([2,1])
    with _col_gc1:
        st.markdown("**Global Cues**")
        if _cues:
            _cue_html = '<div style="display:flex;gap:8px;flex-wrap:wrap">'
            for _cue_item in (_cues if isinstance(_cues, list) else []):
                _name  = getattr(_cue_item,'name','') if hasattr(_cue_item,'name') else str(_cue_item)[:20]
                _cv    = getattr(_cue_item,'value',None) or (dict(_cue_item).get('change_pct',0) if isinstance(_cue_item,dict) else 0)
                try: _cv = float(_cv)
                except: _cv = 0.0
                _cc    = "#16a34a" if _cv >= 0 else "#dc2626"
                _cbg   = "#f0fdf4" if _cv >= 0 else "#fff1f2"
                _cue_html += (
                    f'<div style="background:{_cbg};border:1px solid {_cc}44;border-radius:6px;padding:4px 10px;font-family:JetBrains Mono">'
                    f'<div style="font-size:9px;color:#5a72a0">{_name}</div>'
                    f'<div style="font-size:12px;font-weight:700;color:{_cc}">{"+" if _cv>=0 else ""}{_cv:.2f}%</div>'
                    f'</div>'
                )
            _cue_html += '</div>'
            st.markdown(_cue_html, unsafe_allow_html=True)
        else:
            st.caption("No cues data — click Refresh.")

    with _col_gc2:
        _gb2 = st.session_state.global_bias
        _gs2 = st.session_state.global_score
        _gbc = "#16a34a" if _gb2 == "BULLISH" else "#dc2626" if _gb2 == "BEARISH" else "#64748b"
        st.markdown(
            f'<div class="mcard" style="text-align:center;border-color:{_gbc}44">'
            f'<div class="mlbl">Global Bias</div>'
            f'<div style="font-size:18px;font-weight:700;color:{_gbc};font-family:JetBrains Mono">{_gb2}</div>'
            f'<div style="font-size:12px;color:#5a72a0;font-family:JetBrains Mono">Score: {_gs2:.1f}/10</div>'
            f'</div>',
            unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 1 — AUTO TRADE
# ═══════════════════════════════════════════════════════════════════════════════
with _tabs[1]:
    _at_tabs = st.tabs(["P&L Dashboard", "Futures (Nifty/BN/Sensex)", "Options (Nifty/BN/Sensex)", "Nifty 50 Stocks"])

    # ── Shared segment P&L banner (shown at top of every sub-tab) ────────────
    def _render_segment_pnl_banner(highlight_seg: str = ""):
        """Compact 4-column banner: Equity | Futures | Options | Total + today's deployed."""
        _bt = _all_trades()
        _seg_net  = {"Equity": 0.0, "Futures": 0.0, "Options": 0.0}
        _seg_cnt  = {"Equity": 0,   "Futures": 0,   "Options": 0}
        _seg_wins = {"Equity": 0,   "Futures": 0,   "Options": 0}
        _today    = str(date.today())
        for _t in _bt:
            _s = str(_t.get('Segment', _t.get('segment', 'Equity')))
            if any(x in _s.upper() for x in ('OPT','CE','PE')): _s = 'Options'
            elif any(x in _s.upper() for x in ('FUT','FUTURE')): _s = 'Futures'
            else: _s = 'Equity'
            _pnl = float(_t.get('Net PnL ₹', _t.get('pnl_rs', _t.get('net_pnl', 0))) or 0)
            _gross = float(_t.get('pnl_pts', 0) or 0) * float(_t.get('qty', 1) or 1)
            _charges = float(_t.get('Charges ₹', 0) or 0)
            _net = _pnl if _pnl != 0 else round(_gross - _charges, 2)
            _seg_net[_s]  += _net
            _seg_cnt[_s]  += 1
            _seg_wins[_s] += 1 if _net > 0 else 0

        _total = sum(_seg_net.values())
        _total_cnt = sum(_seg_cnt.values())

        # Try to get today's deployed cap info
        try:
            from pos_auto_trader import _get_today_deployed_rs, DAILY_CAP_RS
            _deployed = _get_today_deployed_rs()
            _cap_pct  = min(_deployed / DAILY_CAP_RS * 100, 100)
            _cap_html = (
                f'<div style="display:flex;align-items:center;gap:8px;margin-top:6px;font-size:10px;color:#64748b">'
                f'<span>Daily Cap:</span>'
                f'<div style="flex:1;background:#e2e8f0;border-radius:4px;height:6px;max-width:120px">'
                f'<div style="width:{_cap_pct:.0f}%;background:{"#ef4444" if _cap_pct>80 else "#f59e0b" if _cap_pct>50 else "#22c55e"};height:100%;border-radius:4px"></div>'
                f'</div>'
                f'<span>₹{_deployed/1e5:.1f}L / ₹{DAILY_CAP_RS/1e5:.0f}L ({_cap_pct:.0f}%)</span>'
                f'</div>'
            )
        except Exception:
            _cap_html = ''

        _cols_data = [
            ("Equity",  "#1d4ed8", "#dbeafe", _seg_net["Equity"],  _seg_cnt["Equity"],  _seg_wins["Equity"]),
            ("Futures", "#7c3aed", "#ede9fe", _seg_net["Futures"], _seg_cnt["Futures"], _seg_wins["Futures"]),
            ("Options", "#d97706", "#fef9c3", _seg_net["Options"], _seg_cnt["Options"], _seg_wins["Options"]),
            ("Total",   "#0f172a", "#f0f6ff", _total,              _total_cnt,          sum(_seg_wins.values())),
        ]

        _html = '<div style="display:flex;gap:8px;margin-bottom:14px;flex-wrap:wrap">'
        for _label, _tc, _bg, _net, _cnt, _wins in _cols_data:
            _is_hl  = highlight_seg and _label == highlight_seg
            _nc     = "#166534" if _net >= 0 else "#991b1b"
            _border = f"3px solid {_tc}" if _is_hl else f"1px solid {_tc}33"
            _wr     = f"{_wins/_cnt*100:.0f}%" if _cnt else "—"
            _html += (
                f'<div style="background:{_bg};border:{_border};border-radius:10px;'
                f'padding:10px 16px;min-width:140px;flex:1;font-family:JetBrains Mono">'
                f'<div style="font-size:10px;font-weight:700;color:{_tc};letter-spacing:1px;text-transform:uppercase">{_label}</div>'
                f'<div style="font-size:20px;font-weight:800;color:{_nc};margin:4px 0">{"+" if _net>=0 else ""}₹{_net:,.0f}</div>'
                f'<div style="font-size:10px;color:#64748b">{_cnt} trades · {_wr} WR</div>'
                f'</div>'
            )
        _html += '</div>'
        if _cap_html:
            _html = _html[:-6] + _cap_html + '</div>'
        st.markdown(_html, unsafe_allow_html=True)

    # ── AT Sub-tab 0: P&L Dashboard ───────────────────────────────────────────
    with _at_tabs[0]:
        _render_segment_pnl_banner()
        import glob as _dp_glob, io as _dp_io, csv as _dp_csv, os as _dp_os
        from collections import defaultdict as _dp_dd

        # ── Reports folder: auto-save ─────────────────────────────────────────
        _REPORTS_DIR2 = _RESULTS_DIR
        os.makedirs(_REPORTS_DIR2, exist_ok=True)

        def _autosave_trades(trades):
            if not trades: return
            try:
                _today2 = date.today().isoformat()
                _jpath  = os.path.join(_REPORTS_DIR2, f"trades_{_today2}.json")
                _cpath  = os.path.join(_REPORTS_DIR2, f"trades_{_today2}.csv")
                with open(_jpath, 'w', encoding='utf-8') as _jf:
                    import json as _jmod; _jmod.dump(trades, _jf, indent=2, ensure_ascii=False, default=str)
                df_save = pd.DataFrame(trades)
                df_save.to_csv(_cpath, index=False)
            except Exception: pass

        # ── Charge calculator (Angel One India) ───────────────────────────────
        def _dp_ind_charges(seg, qty, entry_price, exit_price=0, symbol=''):
            """
            Angel One India charges — corrected formula including lot sizes.
            Brokerage : Rs20 flat per executed order = Rs40 per trade (entry+exit)
            STT       : Futures = 0.01% on sell turnover (lot-size adjusted)
                        Options = 0.05% on sell premium
                        Equity  = 0.025% on sell side (intraday)
            NSE Trans : 0.00193% of total turnover (Futures/Options)
                        0.00325% of total turnover (Equity)
            SEBI      : 0.0001% of turnover
            GST       : 18% on (brokerage + transaction + SEBI)
            Stamp duty: 0.002% on buy-side (Equity) / 0.002% Futures / 0.003% Options
            """
            seg = str(seg).upper()
            qty  = max(int(qty or 1), 1)
            ep   = float(entry_price or 0)
            xp   = float(exit_price or ep)

            # qty is already total contract units (lot size baked in by the bot)
            _sym_up   = str(symbol).upper().replace(' ','')
            buy_to    = ep * qty
            sell_to   = xp * qty
            turnover  = buy_to + sell_to
            brokerage = 40.0  # Rs20 × 2

            if 'OPT' in seg or _sym_up.endswith('CE') or _sym_up.endswith('PE'):
                stt       = sell_to * 0.0005                  # 0.05% sell premium
                trans     = turnover * 0.00053                 # NSE Options: 0.053% of premium
                sebi      = turnover * 0.000001
                stamp     = buy_to  * 0.00003                  # 0.003% buy side
                gst_base  = brokerage + trans + sebi
                charges   = brokerage + stt + trans + sebi + stamp + gst_base * 0.18
            elif 'FUT' in seg or 'FUTURE' in seg:
                stt       = sell_to * 0.0001                   # 0.01% sell side
                trans     = turnover * 0.0000193               # NSE Futures transaction
                sebi      = turnover * 0.000001
                stamp     = buy_to  * 0.00002                  # 0.002% buy side
                gst_base  = brokerage + trans + sebi
                charges   = brokerage + stt + trans + sebi + stamp + gst_base * 0.18
            else:
                # Equity intraday delivery
                stt       = sell_to * 0.00025                  # 0.025% sell (intraday)
                trans     = turnover * 0.0000325               # NSE Equity transaction
                sebi      = turnover * 0.000001
                stamp     = buy_to  * 0.00015                  # 0.015% buy side
                gst_base  = brokerage + trans + sebi
                charges   = brokerage + stt + trans + sebi + stamp + gst_base * 0.18
            return round(charges, 2)

        def _dp_get_charges(t):
            """Use stored charge if reasonable, else recalculate with lot-size-aware formula."""
            stored = float(t.get('Charges Rs', t.get('Charges ₹', t.get('charges', -1))) or -1)
            seg = t.get('Segment', t.get('segment', 'Equity'))
            sym = t.get('Symbol', t.get('symbol', ''))
            qty = float(t.get('Qty', t.get('qty', t.get('contracts', 1))) or 1)
            ep  = _trade_entry(t)
            xp  = _trade_exit(t)
            calc = _dp_ind_charges(seg, qty, ep, xp, symbol=sym)
            # Accept stored only if non-zero and within 10x of calculated (sanity check)
            if stored > 0 and stored < calc * 10:
                return stored
            return calc

        def _dp_get_net(t):
            gross = _trade_gross_pnl(t)
            chg   = _dp_get_charges(t)
            stored_net = float(t.get('Net PnL Rs', t.get('Net PnL ₹', t.get('net_pnl', None)) or 0)  )
            # Recalculate net = gross - correct charges
            return round(gross - chg, 2)

        def _dp_get_invested(t):
            seg = str(t.get('Segment', t.get('segment', 'Equity'))).upper()
            qty = float(t.get('Qty', t.get('qty', t.get('contracts', 1))) or 1)
            ep  = _trade_entry(t)
            # qty already includes lot_size for index instruments — don't double-multiply
            if 'FUT' in seg or 'FUTURE' in seg:
                return round(ep * qty * 0.1, 2)  # ~10% margin; qty has lot_size baked in
            if 'OPT' in seg or seg in ('CE', 'PE', 'OPTIONS'):
                return round(ep * qty, 2)  # premium × qty (qty = lot_size × num_lots)
            return round(ep * qty, 2)  # Equity: price × shares

        def _dp_exit_reason(t):
            er = t.get('Exit Reason', t.get('exit_reason', ''))
            if er: return str(er).upper()[:4]
            net = _dp_get_net(t)
            return 'TP' if net > 0 else 'SL'

        def _dp_seg_clean(t):
            seg = str(t.get('Segment', t.get('segment', 'Equity')))
            if any(x in seg.upper() for x in ('OPT','CE','PE')): return 'Options'
            if any(x in seg.upper() for x in ('FUT','FUTURE')): return 'Futures'
            return 'Equity'

        # ── Load all trades ───────────────────────────────────────────────────
        _dp_all_raw = _all_trades()
        _autosave_trades(_dp_all_raw)  # auto-save on every dashboard load

        _dp_all_trades = []
        _dp_syms = {"Equity": set(), "Futures": set(), "Options": set()}
        for _t in _dp_all_raw:
            _seg2 = _dp_seg_clean(_t)
            _sym2 = _t.get('Symbol', _t.get('symbol', '?'))
            _tf2  = _trade_tf(_t)
            _t2   = dict(_t)
            _t2['_seg'] = _seg2; _t2['_sym'] = _sym2; _t2['_tf'] = _tf2
            _t2['_chg'] = _dp_get_charges(_t2)
            _t2['_net'] = _dp_get_net(_t2)
            _t2['_inv'] = _dp_get_invested(_t2)
            _dp_all_trades.append(_t2)
            _dp_syms[_seg2].add(_sym2)

        # ── Segment totals ────────────────────────────────────────────────────
        _dp_eq_pnl  = sum(t['_net'] for t in _dp_all_trades if t['_seg']=='Equity')
        _dp_ft_pnl  = sum(t['_net'] for t in _dp_all_trades if t['_seg']=='Futures')
        _dp_op_pnl  = sum(t['_net'] for t in _dp_all_trades if t['_seg']=='Options')
        _dp_net_tot = _dp_eq_pnl + _dp_ft_pnl + _dp_op_pnl
        _dp_n       = len(_dp_all_trades)
        _dp_wins    = sum(1 for t in _dp_all_trades if t['_net'] > 0)
        _dp_wr      = _dp_wins / _dp_n * 100 if _dp_n else 0
        _dp_tc      = "#166534" if _dp_net_tot >= 0 else "#991b1b"
        _dp_tbbg    = "#f0fdf4" if _dp_net_tot >= 0 else "#fff1f2"

        # ── Reset button ──────────────────────────────────────────────────────
        _rst_c1, _rst_c2 = st.columns([1, 5])
        with _rst_c1:
            if st.button("Reset All", type="secondary", key="ind_pnl_reset",
                         help="Clears ALL trade history and P&L"):
                for _rf in _glob_mod.glob(os.path.join(_APP_DIR, "pos_state_*.json")):
                    try:
                        import json as _jrst; _ss_r = _jrst.load(open(_rf, encoding='utf-8'))
                        _ss_r['history'] = []; _ss_r['total_trades'] = 0
                        _ss_r['winning'] = 0; _ss_r['total_pnl'] = 0.0; _ss_r['open_trade'] = None
                        _jrst.dump(_ss_r, open(_rf,'w', encoding='utf-8'), indent=2)
                    except: pass
                st.success("All data reset!"); st.rerun()

        # ── Big net banner ────────────────────────────────────────────────────
        st.markdown(
            f'<div style="background:{_dp_tbbg};border:2px solid {_dp_tc}33;border-radius:14px;'
            f'padding:16px 22px;margin-bottom:14px;font-family:JetBrains Mono">'
            f'<div style="font-size:11px;font-weight:700;color:{_dp_tc};margin-bottom:4px">'
            f'P&L DASHBOARD - EQUITY * OPTIONS * FUTURES</div>'
            f'<div style="font-size:36px;font-weight:700;color:{_dp_tc};line-height:1">'
            f'Rs{_dp_net_tot:+,.2f} <span style="font-size:13px;color:#78350f">NET</span></div>'
            f'<div style="color:#78350f;font-size:10px;margin-top:3px">'
            f'{_dp_n} trades &nbsp;|&nbsp; {_dp_wins}W/{_dp_n-_dp_wins}L &nbsp;|&nbsp;'
            f'<span style="color:{"#166534" if _dp_wr>=50 else "#991b1b"}">{_dp_wr:.1f}% win</span>'
            f'</div></div>', unsafe_allow_html=True)

        _dm1, _dm2, _dm3, _dm4 = st.columns(4)
        _dm1.metric("Equity Net",  f"Rs{_dp_eq_pnl:+,.2f}")
        _dm2.metric("Futures Net", f"Rs{_dp_ft_pnl:+,.2f}")
        _dm3.metric("Options Net", f"Rs{_dp_op_pnl:+,.2f}")
        _dm4.metric("Total Net",   f"Rs{_dp_net_tot:+,.2f}")

        # ── Monitored symbol chips ─────────────────────────────────────────────
        _all_eq_syms  = sorted(_dp_syms['Equity']  | set(NSE_TOP_STOCKS))
        _all_fut_syms = sorted(_dp_syms['Futures'] | set(NSE_INDICES))
        _all_opt_syms = sorted(_dp_syms['Options'] | {"NIFTY CE","NIFTY PE","BANKNIFTY CE","BANKNIFTY PE","SENSEX CE","SENSEX PE"})

        def _dp_sym_chips(syms, bg, tc):
            return ''.join(
                f'<span style="background:{bg};color:{tc};padding:2px 8px;border-radius:8px;'
                f'font-size:9px;font-weight:700;margin:2px;display:inline-block">{s}</span>'
                for s in syms)

        st.markdown(
            f'<div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;'
            f'padding:8px 12px;margin-bottom:8px;font-family:JetBrains Mono">'
            f'<div style="font-size:9px;font-weight:700;color:#64748b;margin-bottom:4px">MONITORED SYMBOLS</div>'
            f'<div style="margin-bottom:4px"><span style="font-size:9px;color:#1d4ed8;font-weight:700;margin-right:6px">EQUITY({len(_all_eq_syms)})</span>'
            f'{_dp_sym_chips(_all_eq_syms,"#dbeafe","#1d4ed8")}</div>'
            f'<div style="margin-bottom:4px"><span style="font-size:9px;color:#7c3aed;font-weight:700;margin-right:6px">FUTURES({len(_all_fut_syms)})</span>'
            f'{_dp_sym_chips(_all_fut_syms,"#ede9fe","#7c3aed")}</div>'
            f'<div><span style="font-size:9px;color:#854d0e;font-weight:700;margin-right:6px">OPTIONS({len(_all_opt_syms)})</span>'
            f'{_dp_sym_chips(_all_opt_syms,"#fef9c3","#854d0e")}</div>'
            f'</div>', unsafe_allow_html=True)

        # ── Open positions ─────────────────────────────────────────────────────
        _st_files = _glob_mod.glob(os.path.join(_APP_DIR, "pos_state_*.json"))
        _open_trades_list = []
        if POS_OK:
            _ist_t  = _ist_now()
            _past_e = _ist_t.time() >= time(15, 15)
            for _sf in _st_files:
                _skey = os.path.basename(_sf).replace("pos_state_","").replace(".json","")
                try:
                    _sst = POSState.load(_skey)
                    if _sst.open_trade:
                        if _past_e:
                            _sst.open_trade = None; _sst.status = "STOPPED"
                            _sst.log_msg("EOD auto-close"); _sst.save(_skey)
                        else:
                            _open_trades_list.append((_skey, _sst))
                except: pass
        if _open_trades_list:
            st.markdown("**Open Positions**")
            for _okey, _ost2 in _open_trades_list:
                _otr = _ost2.open_trade
                _odc = "#16a34a" if _otr.direction == "LONG" else "#dc2626"
                st.markdown(
                    f'<div style="background:#fff;border:2px solid {_odc};border-radius:8px;'
                    f'padding:10px 14px;margin:4px 0;font-family:JetBrains Mono;font-size:11px">'
                    f'<span style="color:{_odc};font-weight:700">{_okey} — {_otr.direction} | {_otr.strategy}</span>'
                    f'&nbsp;&nbsp;<span style="color:#5a72a0">Entry: Rs{_otr.entry:,.0f} | Target: Rs{_otr.target:,.0f} | SL: Rs{_otr.stop_loss:,.0f} | Qty: {_otr.qty}</span>'
                    f'</div>', unsafe_allow_html=True)
                if st.button(f"Force Close {_okey}", key=f"fc_{_okey}"):
                    _ost2.open_trade = None; _ost2.status = "STOPPED"
                    _ost2.log_msg("Force-closed"); _ost2.save(_okey)
                    st.success(f"Closed {_okey}!"); st.rerun()
            st.markdown("---")

        # ── View toggle ───────────────────────────────────────────────────────
        _dp_view = st.radio("View", ["Trade by Trade","Daily Summary"],
                            horizontal=True, key="dp_view_ind", label_visibility="collapsed")

        # ── Daily Summary ─────────────────────────────────────────────────────
        if _dp_view == "Daily Summary":
            _daily = _dp_dd(lambda: {"trades":0,"invested":0.0,"gross":0.0,"charges":0.0,"net":0.0,"wins":0})
            for _t3 in _dp_all_trades:
                _d3 = _t3.get('Date', _t3.get('date', date.today().isoformat()))
                _daily[_d3]["trades"]   += 1
                _daily[_d3]["invested"] += _t3['_inv']
                _daily[_d3]["gross"]    += _trade_gross_pnl(_t3)
                _daily[_d3]["charges"]  += _t3['_chg']
                _daily[_d3]["net"]      += _t3['_net']
                if _t3['_net'] > 0: _daily[_d3]["wins"] += 1

            if not _daily:
                st.info("No trades yet.")
            else:
                _ds_sorted = sorted(_daily.items(), key=lambda x: x[0], reverse=True)
                _ds_html = (
                    '<div style="overflow-x:auto"><table style="width:100%;border-collapse:collapse;'
                    'font-family:JetBrains Mono;font-size:11px">'
                    '<thead><tr style="background:#1e293b;color:#fff">'
                    '<th style="padding:8px 10px;text-align:left">Date</th>'
                    '<th style="padding:8px 10px;text-align:center">Trades</th>'
                    '<th style="padding:8px 10px;text-align:right">Invested Rs</th>'
                    '<th style="padding:8px 10px;text-align:right">Gross P&L</th>'
                    '<th style="padding:8px 10px;text-align:right">Charges</th>'
                    '<th style="padding:8px 10px;text-align:right">Net P&L</th>'
                    '<th style="padding:8px 10px;text-align:center">Win%</th>'
                    '<th style="padding:8px 10px;text-align:right">Return%</th>'
                    '</tr></thead><tbody>'
                )
                _ds_tot_inv = _ds_tot_net = _ds_tot_gross = _ds_tot_chg = _ds_tot_tr = _ds_tot_wins = 0
                for _ddate, _dd in _ds_sorted:
                    _dwin = _dd['wins']; _dtr = _dd['trades']
                    _dwr  = _dwin / _dtr * 100 if _dtr else 0
                    _dnet = _dd['net']; _dinv = _dd['invested']
                    _dret = _dnet / _dinv * 100 if _dinv > 0 else 0.0
                    _dnc  = "#166534" if _dnet >= 0 else "#991b1b"
                    _drnc = "#166534" if _dret >= 0 else "#991b1b"
                    try: _dlbl = datetime.strptime(_ddate, '%Y-%m-%d').strftime('%d %b %Y')
                    except: _dlbl = _ddate
                    _ds_html += (
                        f'<tr style="border-bottom:1px solid #e2e8f0">'
                        f'<td style="padding:6px 10px;font-weight:700">{_dlbl}</td>'
                        f'<td style="padding:6px 10px;text-align:center">{_dtr}</td>'
                        f'<td style="padding:6px 10px;text-align:right;color:#5a72a0">Rs{_dinv:,.0f}</td>'
                        f'<td style="padding:6px 10px;text-align:right">Rs{_dd["gross"]:+,.2f}</td>'
                        f'<td style="padding:6px 10px;text-align:right;color:#64748b">Rs{_dd["charges"]:,.2f}</td>'
                        f'<td style="padding:6px 10px;text-align:right;font-weight:700;color:{_dnc}">Rs{_dnet:+,.2f}</td>'
                        f'<td style="padding:6px 10px;text-align:center">{_dwr:.0f}%</td>'
                        f'<td style="padding:6px 10px;text-align:right;font-weight:700;color:{_drnc}">{_dret:+.2f}%</td>'
                        f'</tr>'
                    )
                    _ds_tot_inv += _dinv; _ds_tot_net += _dnet; _ds_tot_gross += _dd['gross']
                    _ds_tot_chg += _dd['charges']; _ds_tot_tr += _dtr; _ds_tot_wins += _dwin
                _ds_tot_nc  = "#166534" if _ds_tot_net >= 0 else "#991b1b"
                _ds_tot_ret = _ds_tot_net / _ds_tot_inv * 100 if _ds_tot_inv > 0 else 0
                _ds_html += (
                    f'<tr style="background:#1e293b;color:#fff;font-weight:700">'
                    f'<td style="padding:8px 10px">TOTAL</td>'
                    f'<td style="padding:8px 10px;text-align:center">{_ds_tot_tr}</td>'
                    f'<td style="padding:8px 10px;text-align:right">Rs{_ds_tot_inv:,.0f}</td>'
                    f'<td style="padding:8px 10px;text-align:right">Rs{_ds_tot_gross:+,.2f}</td>'
                    f'<td style="padding:8px 10px;text-align:right">Rs{_ds_tot_chg:,.2f}</td>'
                    f'<td style="padding:8px 10px;text-align:right;color:{"#86efac" if _ds_tot_net>=0 else "#f87171"}">Rs{_ds_tot_net:+,.2f}</td>'
                    f'<td style="padding:8px 10px;text-align:center">{_ds_tot_wins/_ds_tot_tr*100:.0f}% </td>'
                    f'<td style="padding:8px 10px;text-align:right;color:{"#86efac" if _ds_tot_ret>=0 else "#f87171"}">{_ds_tot_ret:+.2f}%</td>'
                    f'</tr></tbody></table></div>'
                )
                st.markdown(_ds_html, unsafe_allow_html=True)

                # Export daily summary
                st.markdown("")
                _ds_exp_df = pd.DataFrame([
                    {"Date":d,"Trades":v["trades"],"Invested Rs":round(v["invested"],2),
                     "Gross P&L":round(v["gross"],2),"Charges Rs":round(v["charges"],2),
                     "Net P&L":round(v["net"],2),"Win%":round(v["wins"]/v["trades"]*100,1) if v["trades"] else 0}
                    for d,v in _ds_sorted
                ])
                _ds_csv = _ds_exp_df.to_csv(index=False)
                st.download_button("Export Daily Summary CSV", _ds_csv,
                    f"india_daily_summary_{date.today().isoformat()}.csv", "text/csv", key="dp_ds_exp")
            st.stop()

        # ── Filters ───────────────────────────────────────────────────────────
        _dpf1, _dpf2, _dpf3, _dpf4, _dpf5, _dpf6 = st.columns(6)
        with _dpf1: _dp_seg  = st.selectbox("Segment",   ["All","Equity","Futures","Options"], key="dp_seg_ind")
        with _dpf2: _dp_tf   = st.selectbox("Timeframe", ["All","5m","15m","1h","4h"],         key="dp_tf_ind")
        with _dpf3: _dp_dir  = st.selectbox("Direction", ["All","LONG","SHORT"],               key="dp_dir_ind")
        with _dpf4: _dp_res  = st.selectbox("Result",    ["All","WIN","LOSS"],                 key="dp_res_ind")
        with _dpf5: _dp_er   = st.selectbox("Exit",      ["All","TP","SL","EOD","MANUAL"],     key="dp_er_ind")
        with _dpf6: _dp_syminp = st.text_input("Symbol", placeholder="NIFTY,RELIANCE...",     key="dp_sym_ind")

        _dp_shown = _dp_all_trades[:]
        if _dp_seg  != "All": _dp_shown = [t for t in _dp_shown if t['_seg'] == _dp_seg]
        if _dp_tf   != "All": _dp_shown = [t for t in _dp_shown if t['_tf'].lower() == _dp_tf.lower()]
        if _dp_dir  != "All": _dp_shown = [t for t in _dp_shown if t.get('Direction', t.get('direction','')) == _dp_dir]
        if _dp_res == "WIN":  _dp_shown = [t for t in _dp_shown if t['_net'] > 0]
        elif _dp_res == "LOSS": _dp_shown = [t for t in _dp_shown if t['_net'] <= 0]
        if _dp_er  != "All": _dp_shown = [t for t in _dp_shown if _dp_exit_reason(t) == _dp_er]
        if _dp_syminp.strip():
            _dp_sym_set = {s.strip().upper() for s in _dp_syminp.split(',')}
            _dp_shown = [t for t in _dp_shown if str(t.get('Symbol',t.get('symbol',''))).upper() in _dp_sym_set]

        st.caption(f"Showing {len(_dp_shown)} of {len(_dp_all_trades)} trades")

        # ── Trade table ───────────────────────────────────────────────────────
        _dp_rows_html = ""
        _dp_xl        = []

        for _t in _dp_shown[:500]:
            _seg3   = _t['_seg']; _sym3 = _t['_sym']; _tf3 = _t['_tf']
            _dir3   = _t.get('Direction', _t.get('direction', '-'))
            _strat3 = _t.get('Strategy', _t.get('strategy', '-')) or '-'
            _ep3    = _trade_entry(_t); _xp3 = _trade_exit(_t)
            _LOT3   = {'NIFTY':75,'BANKNIFTY':35,'SENSEX':20}
            qty3    = float(_t.get('Qty', _t.get('qty', _t.get('contracts', 1))) or 1)
            _sym3u  = str(_sym3).upper()
            lot3    = _LOT3.get(_sym3u, 1)
            actual_qty3 = qty3 * lot3 if (_ep3 > 1000 and lot3 > 1) else qty3
            _gross3 = _trade_gross_pnl(_t)
            _chg3   = _t['_chg']; _net3 = _t['_net']; _inv3 = _t['_inv']
            _er3    = _dp_exit_reason(_t)
            _pct3   = _trade_pnl_pct(_t)
            _entry_t3 = str(_t.get('Entry Time', _t.get('time', '-')) or '-')[:5]
            _exit_t3  = str(_t.get('Exit Time', _t.get('exit_time', '-')) or '-')[:5]
            _date_s3  = _t.get('Date', _t.get('date', '-'))
            try: _date_lbl3 = datetime.strptime(_date_s3, '%Y-%m-%d').strftime('%d %b')
            except: _date_lbl3 = _date_s3

            _is_win3  = _net3 > 0
            _row_bg3  = "#f0fdf4" if _is_win3 else "#fff1f2"
            _pnl_c3   = "#166534" if _is_win3 else "#991b1b"
            _dir_bg3  = "#dcfce7" if _dir3=="LONG" else ("#fee2e2" if _dir3=="SHORT" else "#f1f5f9")
            _dir_c3   = "#166534" if _dir3=="LONG" else ("#991b1b" if _dir3=="SHORT" else "#64748b")
            _seg_bg3  = {"Equity":"#dbeafe","Futures":"#ede9fe","Options":"#fef9c3"}.get(_seg3,"#f1f5f9")
            _seg_c3   = {"Equity":"#1d4ed8","Futures":"#7c3aed","Options":"#854d0e"}.get(_seg3,"#64748b")
            _er_bg3   = "#dcfce7" if _er3=="TP" else "#fee2e2" if _er3=="SL" else "#fef9c3"
            _er_c3    = "#166534" if _er3=="TP" else "#991b1b" if _er3=="SL" else "#854d0e"
            _tv_url3  = _tv_link(_sym3)

            _prem3 = _t.get('entry_premium', '') if _seg3 == 'Options' else ''
            _xprem3 = _t.get('exit_premium', '') if _seg3 == 'Options' else ''
            _prem_cell3 = (f'Rs{float(_prem3):,.2f}→Rs{float(_xprem3):,.2f}'
                           if _prem3 and _xprem3 else
                           f'Rs{float(_prem3):,.2f}' if _prem3 else '—')
            _dp_rows_html += (
                f'<tr style="background:{_row_bg3};border-bottom:1px solid #e0e9ff;font-size:10px">'
                f'<td style="padding:5px 8px;white-space:nowrap;font-weight:600">{_date_lbl3}</td>'
                f'<td style="padding:5px 8px;white-space:nowrap">{_entry_t3}</td>'
                f'<td style="padding:5px 8px;white-space:nowrap">{_exit_t3}</td>'
                f'<td style="padding:5px 8px"><span style="background:{_seg_bg3};color:{_seg_c3};padding:2px 6px;border-radius:4px;font-size:9px;font-weight:700">{_seg3[:3]}</span></td>'
                f'<td style="padding:5px 8px;font-weight:700;color:#1d4ed8">'
                f'<a href="{_tv_url3}" target="_blank" style="color:#1d4ed8;text-decoration:none">{_sym3}</a></td>'
                f'<td style="padding:5px 8px;color:#64748b">{_tf3}</td>'
                f'<td style="padding:5px 8px"><span style="background:{_dir_bg3};color:{_dir_c3};padding:2px 6px;border-radius:4px;font-size:9px;font-weight:700">{_dir3}</span></td>'
                f'<td style="padding:5px 8px;color:#475569;max-width:140px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap" title="{_strat3}">{_strat3[:22]}</td>'
                f'<td style="padding:5px 8px;text-align:right;color:#1d4ed8;font-weight:700">{int(qty3)}</td>'
                f'<td style="padding:5px 8px;text-align:right">{int(actual_qty3)}</td>'
                f'<td style="padding:5px 8px;text-align:right">Rs{_ep3:,.2f}</td>'
                f'<td style="padding:5px 8px;text-align:right">Rs{_xp3:,.2f}</td>'
                f'<td style="padding:5px 8px;text-align:right;color:#854d0e;font-style:italic">{_prem_cell3}</td>'
                f'<td style="padding:5px 8px;text-align:right;color:#5a72a0">Rs{_inv3:,.0f}</td>'
                f'<td style="padding:5px 8px;text-align:right">{_pct3:+.2f}%</td>'
                f'<td style="padding:5px 8px;text-align:right">Rs{_gross3:+,.2f}</td>'
                f'<td style="padding:5px 8px;text-align:right;color:#dc2626">Rs{_chg3:,.2f}</td>'
                f'<td style="padding:5px 8px;text-align:right;font-weight:700;color:{_pnl_c3}">Rs{_net3:+,.2f}</td>'
                f'<td style="padding:5px 8px"><span style="background:{_er_bg3};color:{_er_c3};padding:2px 6px;border-radius:4px;font-size:9px;font-weight:700">{_er3}</span></td>'
                f'</tr>'
            )
            _dp_xl.append({
                "Date":_date_s3,"Entry Time":_entry_t3,"Exit Time":_exit_t3,
                "Segment":_seg3,"Symbol":_sym3,"TF":_tf3,"Direction":_dir3,"Strategy":_strat3,
                "Lots":int(qty3),"Qty":int(actual_qty3),"Entry Rs":_ep3,"Exit Rs":_xp3,
                "Premium Rs":float(_prem3) if _prem3 else "","Invested Rs":_inv3,"PnL%":_pct3,
                "Gross P&L":_gross3,"Charges Rs":_chg3,"Net P&L":_net3,"Exit":_er3
            })

        _dp_table_html = (
            '<div style="overflow-x:auto;max-height:520px;overflow-y:auto">'
            '<table style="width:100%;border-collapse:collapse;font-family:JetBrains Mono">'
            '<thead><tr style="background:#1e293b;color:#fff;font-size:10px;position:sticky;top:0">'
            '<th style="padding:6px 8px;text-align:left;white-space:nowrap">Date</th>'
            '<th style="padding:6px 8px;text-align:left;white-space:nowrap">Entry</th>'
            '<th style="padding:6px 8px;text-align:left;white-space:nowrap">Exit</th>'
            '<th style="padding:6px 8px">Seg</th>'
            '<th style="padding:6px 8px">Symbol</th>'
            '<th style="padding:6px 8px">TF</th>'
            '<th style="padding:6px 8px">Dir</th>'
            '<th style="padding:6px 8px;min-width:120px">Strategy</th>'
            '<th style="padding:6px 8px;text-align:right">Lots</th>'
            '<th style="padding:6px 8px;text-align:right">Qty</th>'
            '<th style="padding:6px 8px;text-align:right">Entry Rs</th>'
            '<th style="padding:6px 8px;text-align:right">Exit Rs</th>'
            '<th style="padding:6px 8px;text-align:right;color:#a16207">Premium ₹</th>'
            '<th style="padding:6px 8px;text-align:right">Invested</th>'
            '<th style="padding:6px 8px;text-align:right">PnL%</th>'
            '<th style="padding:6px 8px;text-align:right">Gross</th>'
            '<th style="padding:6px 8px;text-align:right">Charges</th>'
            '<th style="padding:6px 8px;text-align:right">Net P&L</th>'
            '<th style="padding:6px 8px">Exit</th>'
            '</tr></thead><tbody>'
            + _dp_rows_html +
            '</tbody></table></div>'
        )
        st.markdown(_dp_table_html, unsafe_allow_html=True)

        # ── Export buttons ────────────────────────────────────────────────────
        st.markdown("")
        _exp1, _exp2, _exp3 = st.columns(3)
        with _exp1:
            if _dp_xl:
                _csv_filt = pd.DataFrame(_dp_xl).to_csv(index=False)
                st.download_button("Export Filtered CSV", _csv_filt,
                    f"india_trades_filtered_{date.today().isoformat()}.csv", "text/csv", key="dp_dl_fil")
        with _exp2:
            if _dp_all_trades:
                _csv_all = pd.DataFrame([{k:v for k,v in t.items() if not k.startswith('_')} for t in _dp_all_trades]).to_csv(index=False)
                st.download_button("Export ALL Trades CSV", _csv_all,
                    f"india_trades_all_{date.today().isoformat()}.csv", "text/csv", key="dp_dl_all")
        with _exp3:
            st.markdown(
                f'<div style="background:#f0fdf4;border:1px solid #86efac;border-radius:8px;'
                f'padding:8px 12px;font-family:JetBrains Mono;font-size:10px;color:#166534">'
                f'Auto-saved to: reports/trades_{date.today().isoformat()}.json + .csv</div>',
                unsafe_allow_html=True)


    # ── AT Sub-tab 1: Futures ──────────────────────────────────────────────────
    with _at_tabs[1]:
        _render_segment_pnl_banner("Futures")
        # ── Strategy banner ───────────────────────────────────────────────────
        st.markdown(_make_tf_banner("Futures", "#7c3aed", "Futures"), unsafe_allow_html=True)

        # ── Bot status banner ─────────────────────────────────────────────────
        _fut_alive2 = sum(1 for b in st.session_state.get('ind_fut_bots',{}).values() if b and getattr(b,'_running',False))
        if _fut_alive2 > 0:
            st.markdown(
                f'<div style="background:linear-gradient(135deg,#3b0764,#7c3aed);color:#fff;'
                f'border-radius:10px;padding:12px 18px;margin-bottom:12px;font-family:JetBrains Mono">'
                f'<div style="font-size:15px;font-weight:700"> FUTURES LIVE — {_fut_alive2} bots active</div>'
                f'<div style="font-size:11px;margin-top:4px;opacity:0.9">Nifty * BankNifty * Sensex | 15m + 1H | Paper: {"ON" if st.session_state.get("sb_paper_mode",True) else "LIVE"}</div>'
                f'</div>', unsafe_allow_html=True)
        else:
            st.markdown(
                '<div style="background:#faf5ff;border:2px solid #7c3aed;border-radius:10px;'
                'padding:12px 18px;margin-bottom:12px;font-family:JetBrains Mono">'
                '<div style="font-size:15px;font-weight:700;color:#7c3aed"> Futures Bots Not Running</div>'
                '<div style="font-size:11px;color:#64748b;margin-top:4px">Click START in sidebar to launch all bots.</div>'
                '</div>', unsafe_allow_html=True)

        # ── Bot status (controlled from sidebar) ─────────────────────────────
        st.info("Futures bots are controlled from the **sidebar** — use the Futures toggle + ▶ START SELECTED.", icon="ℹ️")

        if POS_OK:
            _fat_c1, _fat_c2 = st.columns(2)
            with _fat_c1: _fat_inst = st.selectbox("View status for index", NSE_INDICES, key="fut_inst")
            with _fat_c2: _fat_tf   = st.selectbox("Timeframe", ["5M","15M","1H","4H"], key="fut_tf")

            _fst_key = f"{_fat_inst}_{_fat_tf}"
            try:   _fst = POSState.load(_fst_key)
            except: _fst = POSState()
            _fsc = {"RUNNING":"#7c3aed","STOPPED":"#dc2626","PAUSED":"#d97706"}.get(_fst.status,"#64748b")
            st.markdown(
                f'<div style="background:#fff;border:2px solid {_fsc};border-radius:10px;'
                f'padding:14px;font-family:JetBrains Mono;margin-bottom:10px">'
                f'<div style="color:{_fsc};font-size:16px;font-weight:700">{_fst.status} — {_fst_key}</div>'
                f'<div style="color:#5a72a0;font-size:11px;margin-top:4px">'
                f'Trades today: {_fst.trades_today} | Last scan: {_fst.last_scan or "—"} | Signal: {_fst.last_signal}</div>'
                f'</div>', unsafe_allow_html=True)

        st.markdown("---")

        # ── Activity Log — purple Futures theme ───────────────────────────────
        _FUT_SYMS = set(NSE_INDICES)  # NIFTY, BANKNIFTY, SENSEX
        st.markdown('<div style="color:#7c3aed;font-weight:700;font-family:JetBrains Mono;font-size:13px;margin-bottom:6px">Futures Activity Log — Nifty · BankNifty · Sensex</div>', unsafe_allow_html=True)
        _fat_tf_tabs = st.tabs(["5m","15m","1H","4H","All Trades"])
        for _fat_ti, (_fat_tt, _fat_tfk) in enumerate(zip(_fat_tf_tabs[:4], ["5M","15M","1H","4H"])):
            with _fat_tt:
                _fat_log_entries = []
                for _fat_sf in _glob_mod.glob(os.path.join(_APP_DIR, f"pos_state_*_{_fat_tfk}.json")):
                    try:
                        _fat_sk = os.path.basename(_fat_sf).replace("pos_state_","").replace(".json","")
                        # Only include futures instruments (Nifty/BankNifty/Sensex)
                        _fat_base = _fat_sk.replace(f"_{_fat_tfk}","").split("_FUT")[0]
                        if not any(idx in _fat_sk.upper() for idx in _FUT_SYMS): continue
                        _fat_ss = POSState.load(_fat_sk)
                        for _fat_le in getattr(_fat_ss,'log',[]):
                            _fat_log_entries.append((_fat_sk, _fat_le))
                    except: pass
                if _fat_log_entries:
                    _fat_log_html = '<div style="background:#faf5ff;border:1px solid #7c3aed;border-radius:8px;padding:10px;max-height:220px;overflow-y:auto;font-family:JetBrains Mono;font-size:10px">'
                    for _fat_lk, _fat_ll in reversed(_fat_log_entries[-40:]):
                        _fat_lc = "#7c3aed" if "SIGNAL" in str(_fat_ll).upper() or "ENTRY" in str(_fat_ll).upper() else "#166534" if "EXIT" in str(_fat_ll).upper() or "PROFIT" in str(_fat_ll).upper() else "#1e293b"
                        _fat_log_html += f'<div style="color:{_fat_lc};padding:1px 0;border-bottom:1px solid #ede9fe"><span style="color:#7c3aed;font-weight:700;margin-right:6px">[{_fat_lk}]</span>{_fat_ll}</div>'
                    _fat_log_html += '</div>'
                    st.markdown(_fat_log_html, unsafe_allow_html=True)
                else:
                    st.markdown('<div style="background:#faf5ff;border:1px solid #ddd6fe;border-radius:8px;padding:10px;font-family:JetBrains Mono;font-size:10px;color:#64748b">No futures activity on this timeframe yet.</div>', unsafe_allow_html=True)
        with _fat_tf_tabs[4]:
            # Filter to futures trades only (Segment contains FUT/FUTURE, or symbol is an index)
            _all_fut_trades = [
                t for t in _all_trades()
                if any(x in str(t.get('Segment', t.get('segment', ''))).upper() for x in ('FUT','FUTURE'))
                or str(t.get('Symbol', t.get('symbol', ''))).upper() in _FUT_SYMS
            ]
            if _all_fut_trades:
                _fat_all_html = '<div style="background:#faf5ff;border:1px solid #7c3aed;border-radius:8px;padding:10px;max-height:280px;overflow-y:auto;font-family:JetBrains Mono;font-size:10px">'
                for _fat_at in reversed(_all_fut_trades[-50:]):
                    _fat_at_pnl = _trade_net_pnl(_fat_at)
                    _fat_at_pc  = "#166534" if _fat_at_pnl >= 0 else "#dc2626"
                    _fat_all_html += (
                        f'<div style="color:#1e293b;padding:2px 0;border-bottom:1px solid #ede9fe">'
                        f'<span style="color:#7c3aed;font-weight:700">{_fat_at.get("Symbol","?")}</span>'
                        f' <span style="color:#475569">{_fat_at.get("Direction","?")}</span>'
                        f' <span style="color:#64748b">{_fat_at.get("Strategy","?")}</span>'
                        f' <span style="color:{_fat_at_pc};font-weight:700">₹{_fat_at_pnl:+,.0f}</span>'
                        f'</div>'
                    )
                _fat_all_html += '</div>'
                st.markdown(_fat_all_html, unsafe_allow_html=True)
            else:
                st.markdown('<div style="background:#faf5ff;border:1px solid #ddd6fe;border-radius:8px;padding:10px;font-family:JetBrains Mono;font-size:10px;color:#64748b">No futures trade history yet.</div>', unsafe_allow_html=True)


    with _at_tabs[2]:
        _render_segment_pnl_banner("Options")
        # ── Strategy banner ───────────────────────────────────────────────────
        st.markdown(_make_tf_banner("Options", "#d97706", "Options"), unsafe_allow_html=True)

        # ── Bot status banner ─────────────────────────────────────────────────
        _opt_alive2 = sum(1 for b in st.session_state.get('ind_opt_bots',{}).values() if b and getattr(b,'_running',False))
        if _opt_alive2 > 0:
            st.markdown(
                f'<div style="background:linear-gradient(135deg,#78350f,#d97706);color:#fff;'
                f'border-radius:10px;padding:12px 18px;margin-bottom:12px;font-family:JetBrains Mono">'
                f'<div style="font-size:15px;font-weight:700"> OPTIONS LIVE — {_opt_alive2} bots active</div>'
                f'<div style="font-size:11px;margin-top:4px;opacity:0.9">Nifty CE/PE * BankNifty CE/PE * Sensex CE/PE | Paper Trade always ON</div>'
                f'</div>', unsafe_allow_html=True)
        else:
            st.markdown(
                '<div style="background:#fffbeb;border:2px solid #d97706;border-radius:10px;'
                'padding:12px 18px;margin-bottom:12px;font-family:JetBrains Mono">'
                '<div style="font-size:15px;font-weight:700;color:#d97706"> Options Bots Not Running</div>'
                '<div style="font-size:11px;color:#64748b;margin-top:4px">Options always paper trade. Click START in sidebar to launch.</div>'
                '</div>', unsafe_allow_html=True)

        # ── Bot status (controlled from sidebar) ─────────────────────────────
        st.info("Options bots are controlled from the **sidebar** — use the Options toggle + ▶ START SELECTED.", icon="ℹ️")

        _oat_c1, _oat_c2 = st.columns(2)
        with _oat_c1: _oat_inst = st.selectbox("View status for index", NSE_INDICES, key="opt_inst")
        with _oat_c2: _oat_tf   = st.selectbox("Timeframe", ["5M","15M","1H","4H"], key="opt_tf")

        if OPT_OK:
            _oat_key = f"{_oat_inst}_{_oat_tf}"
            try:   _ost2 = OptionsState.load(_oat_key)
            except: _ost2 = OptionsState()
            _oat_live_bot = st.session_state.get('ind_opt_bots', {}).get(_oat_key)
            if _oat_live_bot and getattr(_oat_live_bot, '_running', False):
                _ost2.status = "RUNNING"

            _osc = {"RUNNING":"#d97706","STOPPED":"#dc2626","PAUSED":"#d97706"}.get(_ost2.status,"#64748b")
            st.markdown(
                f'<div style="background:#fff;border:2px solid {_osc};border-radius:10px;'
                f'padding:14px;font-family:JetBrains Mono;margin-bottom:10px">'
                f'<div style="color:{_osc};font-size:16px;font-weight:700">{_ost2.status} — {_oat_inst} {_oat_tf} Options</div>'
                f'<div style="color:#5a72a0;font-size:11px;margin-top:4px">Trades today: {_ost2.trades_today} | Last scan: {_ost2.last_scan or "—"} | Signal: {_ost2.last_signal}</div>'
                f'</div>', unsafe_allow_html=True)

            if _ost2.open_trade:
                _otr2 = _ost2.open_trade
                st.markdown(
                    f'<div style="background:#fffbeb;border:2px solid #d97706;border-radius:8px;'
                    f'padding:10px;font-family:JetBrains Mono;font-size:11px">'
                    f'<span style="color:#92400e;font-weight:700">OPEN: {getattr(_otr2,"instrument","")} '
                    f'{getattr(_otr2,"strike","")} {getattr(_otr2,"option_type","")}</span><br>'
                    f'<span style="color:#5a72a0">Entry: Rs{getattr(_otr2,"entry_premium",0):.2f} | '
                    f'Target: Rs{getattr(_otr2,"target_premium",0):.2f} | '
                    f'SL: Rs{getattr(_otr2,"sl_premium",0):.2f} | Qty: {getattr(_otr2,"qty",0)}</span>'
                    f'</div>', unsafe_allow_html=True)
        else:
            st.warning("options_auto_trader.py not found. Install to enable live options.")

        st.markdown("---")

        # ── Activity Log — yellow/black theme ────────────────────────────────
        st.markdown('<div style="color:#92400e;font-weight:700;font-family:JetBrains Mono;font-size:13px;margin-bottom:6px">Options Activity Log — All Timeframes (5m | 15m | 1H | 4H)</div>', unsafe_allow_html=True)
        _oat_tf_tabs = st.tabs(["5m","15m","1H","4H","All"])
        for _oat_ti, (_oat_tt, _oat_tfk) in enumerate(zip(_oat_tf_tabs[:4], ["5M","15M","1H","4H"])):
            with _oat_tt:
                _oat_log = []
                if OPT_OK:
                    for _ons in NSE_INDICES:
                        try:
                            _oss2 = OptionsState.load(_ons)
                            for _ole in getattr(_oss2,'log',[]):
                                _oat_log.append((_ons, _ole))
                        except: pass
                if _oat_log:
                    _oat_log_html = '<div style="background:#f8fafc;border:1px solid #1d4ed8;border-radius:8px;padding:10px;max-height:220px;overflow-y:auto;font-family:JetBrains Mono;font-size:10px">'
                    for _oat_lk, _oat_ll in reversed(_oat_log[-40:]):
                        _oat_lc = "#1d4ed8" if "SIGNAL" in str(_oat_ll).upper() else "#166534" if "EXIT" in str(_oat_ll).upper() else "#1e293b"
                        _oat_log_html += f'<div style="color:{_oat_lc};padding:1px 0;border-bottom:1px solid #e2e8f0"><span style="color:#1d4ed8;font-weight:700;margin-right:6px">[{_oat_lk}]</span>{_oat_ll}</div>'
                    _oat_log_html += '</div>'
                    st.markdown(_oat_log_html, unsafe_allow_html=True)
                else:
                    st.markdown('<div style="background:#f8fafc;border:1px solid #dde9ff;border-radius:8px;padding:10px;font-family:JetBrains Mono;font-size:10px;color:#64748b">No options activity on this timeframe yet.</div>', unsafe_allow_html=True)
        with _oat_tf_tabs[4]:
            _all_opt_trades = [t for t in _all_trades() if 'OPT' in str(t.get('Segment',''))]
            if _all_opt_trades:
                _oat_all_html = '<div style="background:#f8fafc;border:1px solid #1d4ed8;border-radius:8px;padding:10px;max-height:280px;overflow-y:auto;font-family:JetBrains Mono;font-size:10px">'
                for _oat_at in reversed(_all_opt_trades[-50:]):
                    _oat_pnl = _trade_net_pnl(_oat_at)
                    _oat_pc  = "#166534" if _oat_pnl >= 0 else "#dc2626"
                    _oat_all_html += (
                        f'<div style="color:#e2e8f0;padding:2px 0;border-bottom:1px solid #e2e8f0">'
                        f'<span style="color:#fbbf24">{_oat_at.get("Symbol","?")} {_oat_at.get("Direction","?")}</span>'
                        f' <span style="color:#475569">{_oat_at.get("Strategy","?")}</span>'
                        f' <span style="color:{_oat_pc};font-weight:700">Rs{_oat_pnl:+,.0f}</span>'
                        f'</div>'
                    )
                _oat_all_html += '</div>'
                st.markdown(_oat_all_html, unsafe_allow_html=True)
            else:
                st.markdown('<div style="background:#f8fafc;border:1px solid #dde9ff;border-radius:8px;padding:10px;font-family:JetBrains Mono;font-size:10px;color:#64748b">No options trade history yet.</div>', unsafe_allow_html=True)


    with _at_tabs[3]:
        _render_segment_pnl_banner("Equity")
        # ── Strategy banner ───────────────────────────────────────────────────
        st.markdown(_make_tf_banner("Equity", "#1d4ed8", "Nifty 50 Equity"), unsafe_allow_html=True)

        # ── Bot status ────────────────────────────────────────────────────────
        _n50_alive2 = sum(1 for b in st.session_state.get('ind_eq_bots',{}).values() if b and getattr(b,'_running',False))
        if _n50_alive2 > 0:
            st.markdown(
                f'<div style="background:linear-gradient(135deg,#1e3a8a,#1d4ed8);color:#fff;'
                f'border-radius:10px;padding:12px 18px;margin-bottom:12px;font-family:JetBrains Mono">'
                f'<div style="font-size:15px;font-weight:700"> EQUITY LIVE — {_n50_alive2} bots active</div>'
                f'<div style="font-size:11px;margin-top:4px;opacity:0.9">Nifty 50 stocks | 5m + 15m + 1H + 4H | Paper: {"ON" if st.session_state.get("sb_paper_mode",True) else "LIVE"}</div>'
                f'</div>', unsafe_allow_html=True)
        else:
            st.markdown(
                '<div style="background:#eff6ff;border:2px solid #1d4ed8;border-radius:10px;'
                'padding:12px 18px;margin-bottom:12px;font-family:JetBrains Mono">'
                '<div style="font-size:15px;font-weight:700;color:#1d4ed8"> Equity Bots Not Running</div>'
                '<div style="font-size:11px;color:#64748b;margin-top:4px">Select stock + TF below and click Start Bot, or use sidebar START.</div>'
                '</div>', unsafe_allow_html=True)

        # ── Bot controls ──────────────────────────────────────────────────────
        _n50_c1, _n50_c2, _n50_c3 = st.columns(3)
        with _n50_c1: _n50_sym = st.selectbox("Stock", NSE_TOP_STOCKS, key="n50_sym")
        with _n50_c2: _n50_seg = st.selectbox("Segment", ["Equity","Futures","Options"], key="n50_seg")
        with _n50_c3: _n50_tf  = st.radio("Timeframe", ["5M","15M","1H","4H"], horizontal=True, key="n50_tf")
        _n50_lots = st.number_input("Lots / Qty", value=1, min_value=1, key="n50_lots")

        _n50b1, _n50b2 = st.columns(2)
        with _n50b1:
            if st.button("Start Bot", type="primary", width='stretch', key="n50_start"):
                if POS_OK:
                    try:
                        _n50bot = POSAutoTrader(instrument=_n50_sym, timeframe=_n50_tf,
                                               qty=_n50_lots,
                                               paper_trade=st.session_state.get('sb_paper_mode',True))
                        _n50bot.start()
                        _n50key = f"{_n50_sym}_{_n50_tf}"
                        _bots_d = dict(st.session_state.get('ind_eq_bots',{}))
                        _bots_d[_n50key] = _n50bot
                        st.session_state.ind_eq_bots = _bots_d
                        st.success(f"Started: {_n50_sym} {_n50_tf} {_n50_seg}"); st.rerun()
                    except Exception as _n50e: st.error(str(_n50e))
                else:
                    st.info("Bot backend not loaded — paper mode simulation only")
        with _n50b2:
            if st.button("Stop Bot", width='stretch', key="n50_stop"):
                _n50key = f"{_n50_sym}_{_n50_tf}"
                _bots_d = dict(st.session_state.get('ind_eq_bots',{}))
                if _n50key in _bots_d:
                    try: _bots_d[_n50key].stop()
                    except: pass
                    del _bots_d[_n50key]
                    st.session_state.ind_eq_bots = _bots_d
                st.success(f"Stopped: {_n50key}"); st.rerun()

        st.markdown("---")

        # ── Nifty 50 price grid ───────────────────────────────────────────────
        st.markdown('<div style="color:#1d4ed8;font-weight:700;font-family:JetBrains Mono;font-size:13px;margin-bottom:6px">Nifty 50 Universe — Live Prices</div>', unsafe_allow_html=True)
        _n50_prices = _fetch_nse_ticker_prices(tuple(NSE_TOP_STOCKS))
        _n50_running = set(st.session_state.get('ind_eq_bots',{}).keys())
        _n50_cols = st.columns(5)
        for _n50i, _n50s in enumerate(NSE_TOP_STOCKS):
            _n50p, _n50c = _n50_prices.get(_n50s, (0.0, 0.0))
            _n50cc = "#16a34a" if _n50c >= 0 else "#dc2626"
            _n50arr = "+" if _n50c >= 0 else ""
            _n50pstr = "Rs{:,.0f}".format(_n50p) if _n50p > 0 else "--"
            _n50_is_running = any(_n50s in k for k in _n50_running)
            _n50_border = "#16a34a" if _n50_is_running else "#dde9ff"
            _n50_badge = ' <span style="background:#16a34a;color:#fff;border-radius:4px;padding:1px 4px;font-size:8px">RUN</span>' if _n50_is_running else ""
            with _n50_cols[_n50i % 5]:
                st.markdown(
                    f'<div style="background:#fff;border:1px solid {_n50_border};border-radius:8px;'
                    f'padding:8px 10px;margin:3px 0;font-family:JetBrains Mono">'
                    f'<div style="color:#1d4ed8;font-weight:700;font-size:11px">{_n50s}{_n50_badge}</div>'
                    f'<div style="color:#1e293b;font-size:12px;font-weight:700">{_n50pstr}</div>'
                    f'<div style="color:{_n50cc};font-size:10px">{_n50arr}{_n50c:.2f}%</div>'
                    f'</div>', unsafe_allow_html=True)

        st.markdown("---")

        # ── Activity Log — yellow/black theme ────────────────────────────────
        st.markdown('<div style="color:#92400e;font-weight:700;font-family:JetBrains Mono;font-size:13px;margin-bottom:6px">Activity Log — All Timeframes (5m | 15m | 1H | 4H)</div>', unsafe_allow_html=True)
        _n50_act_tabs = st.tabs(["5m Activity","15m Activity","1H Activity","4H Activity","All Activity"])
        for _n50_ati, (_n50_att, _n50_attf) in enumerate(zip(_n50_act_tabs[:4], ["5M","15M","1H","4H"])):
            with _n50_att:
                _n50_log_entries = []
                if POS_OK:
                    for _n50_sf in _glob_mod.glob(os.path.join(_APP_DIR, f"pos_state_*_{_n50_attf}.json")):
                        try:
                            _n50_sk = os.path.basename(_n50_sf).replace("pos_state_","").replace(".json","")
                            _n50_ss = POSState.load(_n50_sk)
                            for _n50_le in getattr(_n50_ss,'log',[]):
                                _n50_log_entries.append((_n50_sk, _n50_le))
                        except: pass
                if _n50_log_entries:
                    _n50_log_html = '<div style="background:#f8fafc;border:1px solid #1d4ed8;border-radius:8px;padding:10px;max-height:220px;overflow-y:auto;font-family:JetBrains Mono;font-size:10px">'
                    for _n50_lk, _n50_ll in reversed(_n50_log_entries[-40:]):
                        _n50_lc = "#1d4ed8" if "SIGNAL" in str(_n50_ll).upper() or "ENTRY" in str(_n50_ll).upper() else "#166534" if "EXIT" in str(_n50_ll).upper() else "#1e293b"
                        _n50_log_html += f'<div style="color:{_n50_lc};padding:1px 0;border-bottom:1px solid #e2e8f0"><span style="color:#1d4ed8;font-weight:700;margin-right:6px">[{_n50_lk}]</span>{_n50_ll}</div>'
                    _n50_log_html += '</div>'
                    st.markdown(_n50_log_html, unsafe_allow_html=True)
                else:
                    st.markdown(f'<div style="background:#f8fafc;border:1px solid #dde9ff;border-radius:8px;padding:10px;font-family:JetBrains Mono;font-size:10px;color:#64748b">No {_n50_attf} equity activity yet.</div>', unsafe_allow_html=True)
        with _n50_act_tabs[4]:
            _n50_all_t = _all_trades()
            if _n50_all_t:
                _n50_all_html = '<div style="background:#f8fafc;border:1px solid #1d4ed8;border-radius:8px;padding:10px;max-height:280px;overflow-y:auto;font-family:JetBrains Mono;font-size:10px">'
                for _n50_at2 in reversed(_n50_all_t[-60:]):
                    _n50_pnl2 = _trade_net_pnl(_n50_at2)
                    _n50_pc2  = "#166534" if _n50_pnl2 >= 0 else "#dc2626"
                    _n50_all_html += (
                        f'<div style="color:#e2e8f0;padding:2px 0;border-bottom:1px solid #e2e8f0">'
                        f'<span style="color:#f59e0b">{_n50_at2.get("Symbol","?")}</span>'
                        f' {_n50_at2.get("Direction","?")} '
                        f'<span style="color:#94a3b8">{_n50_at2.get("Strategy","?")}</span>'
                        f' <span style="color:{_n50_pc2};font-weight:700">Rs{_n50_pnl2:+,.0f}</span>'
                        f' <span style="color:#64748b;font-size:9px">{str(_n50_at2.get("Exit Time",""))[:16]}</span>'
                        f'</div>'
                    )
                _n50_all_html += '</div>'
                st.markdown(_n50_all_html, unsafe_allow_html=True)
            else:
                st.markdown('<div style="background:#f8fafc;border:1px solid #dde9ff;border-radius:8px;padding:10px;font-family:JetBrains Mono;font-size:10px;color:#64748b">No equity trade history yet.</div>', unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 2 — ALL STRATEGIES
# ═══════════════════════════════════════════════════════════════════════════════
with _tabs[2]:
    # ──────────────────────────────────────────────────────────────────────────
    # 82 India Strategies: Equity(30) · Options(27) · Futures(25)
    # ──────────────────────────────────────────────────────────────────────────
    _IND_ALL_STRATS = [
        # ===== EQUITY (30) =====
        {"col":"EQ","name":"PPT Fibonacci 61.8%","seg":"EQ · OPT · FUT","tfs":["5m","15m","1h","4h"],
         "logic":"Gareth Soloway's primary setup. Price pulls back to 61.8% Fibonacci retracement of prior impulse. Works on Nifty/BankNifty/top stocks during trend continuation.",
         "entry":"Enter at 61.8% Fib level with reversal candle confirmation (hammer/engulfing). Volume should contract at retracement.",
         "sl_tgt":"SL: below 78.6% Fib | T1: prior high/low | T2: 100% Fib extension"},

        {"col":"EQ","name":"PPT Gap Fill","seg":"EQ · FUT","tfs":["5m","15m"],
         "logic":"Price gaps up/down at open. Gaps in Nifty/top stocks tend to fill within first 30-60 minutes of RTH (9:15-10:15 IST).",
         "entry":"Enter fade of gap direction after first 5-min candle closes with reversal signal. Confirm with VWAP reclaim.",
         "sl_tgt":"SL: beyond gap candle high/low | T1: VWAP | T2: prior day close (gap fill)"},

        {"col":"EQ","name":"PPT Tail Reversal","seg":"EQ · OPT","tfs":["15m","1h"],
         "logic":"Long candle wick (tail) rejecting key level: support, resistance, VWAP, or Fibonacci. Indicates trapped traders and reversal momentum.",
         "entry":"Enter on next candle break of tail candle body. Require volume spike at tail candle.",
         "sl_tgt":"SL: beyond tail tip | T1: 1:1 RR | T2: 1:2 RR"},

        {"col":"EQ","name":"9/20 EMA Pullback","seg":"EQ · FUT","tfs":["5m","15m","1h"],
         "logic":"Price above both 9 EMA and 20 EMA (uptrend) or below both (downtrend). Enter on pullback to 9 or 20 EMA with bounce candle.",
         "entry":"Enter when price touches 9/20 EMA zone and shows reversal candle. EMA separation > 0.1% for valid trend.",
         "sl_tgt":"SL: below 20 EMA (longs) / above 20 EMA (shorts) | T1: prior swing | T2: channel top/bottom"},

        {"col":"EQ","name":"Opening Range Breakout","seg":"EQ · FUT","tfs":["5m","15m"],
         "logic":"Define first 15-min or 30-min range (ORB). Breakout with volume above/below range signals day direction. High-probability on index heavy days.",
         "entry":"Enter on break and close above/below ORB high/low. Volume must be 1.5x average. No entry after 11:00 IST.",
         "sl_tgt":"SL: back inside ORB (< 0.3% retracement) | T1: 1x ORB size | T2: 2x ORB size"},

        {"col":"EQ","name":"VWAP Reclaim","seg":"EQ · FUT","tfs":["5m","15m"],
         "logic":"Price dips below VWAP then reclaims it — signals bulls regaining control. Powerful in first 2 hours of NSE session.",
         "entry":"Enter long when price closes 5-min bar above VWAP after being below. Enter short on loss and fail-to-reclaim.",
         "sl_tgt":"SL: 0.2% below VWAP (longs) | T1: yesterday high / round number | T2: ATH zone"},

        {"col":"EQ","name":"Gap and Go","seg":"EQ","tfs":["5m"],
         "logic":"Stock/index gaps up with strong volume and continues direction. Works best on Nifty stocks with gap > 0.5% on news/result catalyst.",
         "entry":"Enter on first pullback candle after gap-up open holds 9 EMA. Entry must be within first 30 min.",
         "sl_tgt":"SL: below gap candle low | T1: +1.5% from entry | T2: +3%"},

        {"col":"EQ","name":"High of Day Break","seg":"EQ · FUT","tfs":["5m","15m"],
         "logic":"Price consolidates then breaks to new intraday high with volume. Strong trending day signal used for momentum entries.",
         "entry":"Enter on candle close above previous intraday HOD. Volume must be 2x recent bars.",
         "sl_tgt":"SL: below HOD breakout candle | T1: HOD + 0.5% | T2: HOD + 1%"},

        {"col":"EQ","name":"Pre-Market Range Break","seg":"EQ · FUT","tfs":["5m"],
         "logic":"Use 9:00-9:15 IST pre-open price range. Breakout of this range after 9:15 open gives directional bias for entire session.",
         "entry":"Enter on 5-min break above pre-open high or below pre-open low with volume.",
         "sl_tgt":"SL: back inside pre-open range | T1: 2x pre-open range size | T2: prior day HOD/LOD"},

        {"col":"EQ","name":"Scalp Reversal at S/R","seg":"EQ · OPT","tfs":["5m"],
         "logic":"Price approaches major horizontal S/R (yesterday HOD/LOD, round numbers, weekly pivots) and shows 2-bar reversal pattern.",
         "entry":"Enter on second reversal candle completion. Tight 0.2-0.3% stop.",
         "sl_tgt":"SL: 0.25% beyond S/R level | T1: VWAP | T2: mid-range"},

        {"col":"EQ","name":"Volume Climax Reversal","seg":"EQ · OPT","tfs":["5m","15m"],
         "logic":"Exhaustion move: huge volume spike with wide-range candle at support/resistance. Indicates panic selling or euphoric buying near turn.",
         "entry":"Enter counter-trend on next bar after volume climax bar. Confirm with inside bar or doji.",
         "sl_tgt":"SL: beyond climax candle | T1: VWAP | T2: 50% of climax candle"},

        {"col":"EQ","name":"Doji Reversal","seg":"EQ · OPT","tfs":["15m","1h"],
         "logic":"Doji candle at S/R or VWAP after extended move signals indecision and potential reversal. More reliable on higher TF.",
         "entry":"Enter in direction of break of doji high (long) or low (short) on confirming next candle.",
         "sl_tgt":"SL: opposite end of doji body | T1: 1:1.5 RR | T2: 1:2 RR"},

        {"col":"EQ","name":"3-Bar Reversal","seg":"EQ · FUT","tfs":["15m","1h"],
         "logic":"Three consecutive bars in one direction followed by reversal candle. Pattern signals momentum exhaustion. NSE version works well post 10:30 IST.",
         "entry":"Enter on break of reversal candle (4th bar) in opposite direction to the 3-bar move.",
         "sl_tgt":"SL: beyond 3rd bar extreme | T1: start of 3-bar pattern | T2: prior structure"},

        {"col":"EQ","name":"Bollinger Band Squeeze","seg":"EQ · FUT","tfs":["15m","1h"],
         "logic":"BB bandwidth < 10% of 6-month average signals low-volatility squeeze. Breakout from squeeze tends to be explosive (Keltner squeeze indicator).",
         "entry":"Enter on first close outside BB after squeeze. Direction = expansion direction.",
         "sl_tgt":"SL: back inside BB middle | T1: outer BB expanded | T2: 2x ATR from entry"},

        {"col":"EQ","name":"Inside Candle Breakout","seg":"EQ · FUT","tfs":["15m","1h","4h"],
         "logic":"Inside candle (IB) within prior bar range signals coiling energy. Breakout of IB mother bar in trend direction = high-probability entry.",
         "entry":"Enter on break of mother bar high (uptrend) or low (downtrend) after IB formation.",
         "sl_tgt":"SL: opposite end of mother bar | T1: measured move (mother bar height) | T2: 2x measured move"},

        {"col":"EQ","name":"Round Number","seg":"EQ · FUT","tfs":["5m","15m"],
         "logic":"Round numbers (Nifty 22000/22500/23000, BankNifty 48000/50000) act as psychological S/R. Price reacts at these levels predictably.",
         "entry":"Enter reversal at round number with 2-bar confirmation. Enter breakout on close 0.2% beyond round number.",
         "sl_tgt":"SL: 0.3% beyond round number | T1: next round number zone | T2: 2x ATR"},

        {"col":"EQ","name":"Momentum Swing","seg":"EQ · FUT","tfs":["1h","4h"],
         "logic":"Strong trend day: price breaks prior day high/low in first hour. Enter swing in direction. Hold multi-day. Works on Nifty/Bank Nifty trending weeks.",
         "entry":"Enter on pullback to 20 EMA during strong trend. Require RSI > 55 (longs) or RSI < 45 (shorts).",
         "sl_tgt":"SL: below 50 EMA | T1: 1.5x ATR | T2: prior major swing"},

        {"col":"EQ","name":"Support Bounce","seg":"EQ · OPT","tfs":["15m","1h"],
         "logic":"Price approaches daily/weekly support with decreasing volume, then bounces. Support = prior swing low or demand zone from weekly chart.",
         "entry":"Enter long on hammer/engulfing at support with volume. Confirm RSI not in oversold on 1H.",
         "sl_tgt":"SL: 0.3% below support | T1: prior resistance | T2: weekly R1 pivot"},

        {"col":"EQ","name":"MA Crossover 20/50","seg":"EQ · FUT","tfs":["1h","4h"],
         "logic":"20 EMA crosses above/below 50 EMA signaling trend change. Lagging but reliable for swing entries. Used on daily for positional trades.",
         "entry":"Enter on 20/50 crossover bar close in direction. Add on first pullback to crossed MAs.",
         "sl_tgt":"SL: below 50 EMA | T1: 2x ATR | T2: Fibonacci projection"},

        {"col":"EQ","name":"Fabio Daily (Value Area)","seg":"EQ · FUT","tfs":["1h","4h"],
         "logic":"Fabio Ornellas' method: buy deep value in strong stock. Price pulls to 61.8-78.6% of weekly range. Not a reversal — a continuation entry.",
         "entry":"Enter in direction of weekly trend on daily pullback to value zone (61.8-78.6% of prior week range).",
         "sl_tgt":"SL: below 88.6% retracement | T1: prior week high | T2: Fibonacci extension 127%"},

        {"col":"EQ","name":"Breakout Momentum","seg":"EQ · FUT","tfs":["15m","1h"],
         "logic":"Multi-week consolidation breakout on expanding volume. Stock breaks above resistance of 3-week range. Volume must be 3x 20-day average.",
         "entry":"Enter on daily close above range high. Add on first pullback to breakout level (converted to support).",
         "sl_tgt":"SL: back below breakout level | T1: measured move from range | T2: 52-week high"},

        {"col":"EQ","name":"Earnings Momentum","seg":"EQ · OPT","tfs":["5m","15m"],
         "logic":"Stock gaps on quarterly results. If gap holds and direction matches consensus, enter continuation. Avoid fading earnings gaps.",
         "entry":"Enter long after 15-min holding above gap level (no fill). Volume must sustain > 2x average.",
         "sl_tgt":"SL: gap fill (50%) | T1: +3% | T2: +5% from open"},

        {"col":"EQ","name":"Mean Reversion","seg":"EQ · OPT","tfs":["15m","1h"],
         "logic":"Stock extended 2+ standard deviations from 20-day MA. RSI > 75 (overbought) or < 25 (oversold). Fade the extension for quick reversion.",
         "entry":"Enter counter-trend at 2σ BB band with confirming candle. Small size — against trend.",
         "sl_tgt":"SL: 2.5σ BB band | T1: 1σ BB band | T2: 20-day MA"},

        {"col":"EQ","name":"Bull Flag","seg":"EQ · FUT","tfs":["15m","1h"],
         "logic":"Strong pole (>1.5% impulse) followed by orderly 3-5 bar consolidation with declining volume. Classic continuation pattern.",
         "entry":"Enter on break of flag high with volume surge (>1.5x average). Ideal 1-3 PM IST entry.",
         "sl_tgt":"SL: below flag low | T1: pole height projected from flag | T2: 1.618x pole"},

        {"col":"EQ","name":"RSI Divergence","seg":"EQ · OPT","tfs":["1h","4h"],
         "logic":"Price makes new high/low but RSI doesn't confirm. Hidden/regular divergence at key levels signals momentum failure.",
         "entry":"Enter counter-trend on divergence candle reversal. Require structure confirmation (break of minor S/R).",
         "sl_tgt":"SL: beyond divergence swing | T1: 38.2% retracement | T2: 61.8% retracement"},

        {"col":"EQ","name":"MACD Histogram Reversal","seg":"EQ · FUT","tfs":["1h","4h"],
         "logic":"MACD histogram reverses from extreme reading. Zero-line crossover adds confluence. Works best on daily chart for swing trades.",
         "entry":"Enter on first MACD histogram bar shrinking after peak/trough. Confirm with price action reversal candle.",
         "sl_tgt":"SL: beyond swing high/low | T1: MACD zero line cross | T2: prior swing"},

        {"col":"EQ","name":"Supertrend Signal","seg":"EQ · FUT","tfs":["15m","1h"],
         "logic":"Supertrend indicator (10, 3 ATR) flip from red to green (buy) or green to red (sell). Simple trend-following signal widely used in NSE algos.",
         "entry":"Enter on Supertrend color change candle close. Confirm with RSI > 50 (buy) or < 50 (sell).",
         "sl_tgt":"SL: Supertrend line | T1: 1.5x ATR | T2: prior swing"},

        {"col":"EQ","name":"Volume Profile POC","seg":"EQ · FUT","tfs":["1h","4h"],
         "logic":"Point of Control (POC) — highest volume price level of session/week. Price gravitates to POC and tends to auction around it.",
         "entry":"Enter fade at Value Area High/Low (VAH/VAL). Enter breakout when price closes beyond VAH/VAL with volume.",
         "sl_tgt":"SL: 0.3% beyond VAH/VAL | T1: POC | T2: opposite VA boundary"},

        {"col":"EQ","name":"AI Narrative + Technical","seg":"EQ · OPT","tfs":["1h","4h"],
         "logic":"Combine AI sector momentum (e.g., IT, pharma, defence themes) with technical setup. Enter stocks in AI/theme narrative at pullback.",
         "entry":"Enter theme stock at 61.8% Fib pullback during sector strength week. Require Nifty sector index trend up.",
         "sl_tgt":"SL: below 78.6% Fib | T1: prior high | T2: sector ETF target"},

        {"col":"EQ","name":"Relative Strength Leader","seg":"EQ","tfs":["1h","4h"],
         "logic":"Find stock outperforming Nifty on down days and making new highs on up days. RS leader tends to sustain outperformance in current cycle.",
         "entry":"Enter RS leader on Nifty pullback day — if stock holds or rises, it's a strong long.",
         "sl_tgt":"SL: day low | T1: 52-week high | T2: sector rotation target"},

        # ===== OPTIONS (27) =====
        {"col":"OPT","name":"Iron Condor","seg":"OPT","tfs":["4h"],
         "logic":"Sell OTM call + buy further OTM call + sell OTM put + buy further OTM put. Profit if Nifty/BankNifty stays in range. Best before weekly expiry.",
         "entry":"Enter with 5-7 DTE. Short strikes at 1 SD. Collect > 0.5% of index as premium.",
         "sl_tgt":"SL: 2x premium collected | T1: 50% premium decay | T2: 75% premium decay"},

        {"col":"OPT","name":"Iron Butterfly","seg":"OPT","tfs":["4h"],
         "logic":"Sell ATM call and put (same strike) + buy OTM wings. Profit if index pins at ATM strike at expiry. High premium collection, limited range.",
         "entry":"Enter 2-3 DTE near expiry when IV high. Select ATM strike at current index level.",
         "sl_tgt":"SL: index moves > 0.8% from ATM | T1: 40% premium decay | T2: expiry pin"},

        {"col":"OPT","name":"Long Butterfly","seg":"OPT","tfs":["4h"],
         "logic":"Buy lower strike call + sell 2x middle strike calls + buy upper strike call. Low risk, profit at middle strike pin. Used when expecting low movement.",
         "entry":"Enter 5-7 DTE with debit < 20% of spread width. Center strike = ATM.",
         "sl_tgt":"SL: full debit paid | T1: 2x debit | T2: max profit at center strike"},

        {"col":"OPT","name":"Broken Wing Butterfly","seg":"OPT","tfs":["4h"],
         "logic":"Asymmetric butterfly — skip strike on one wing to collect credit. Profit if market goes sideways or in one direction. Zero-cost or credit entry possible.",
         "entry":"Structure: buy 1 call, sell 2 calls (1 strike up), buy 1 call (3 strikes up). Adjust for credit.",
         "sl_tgt":"SL: if net credit, SL = defined loss on downside | T1: middle strike area | T2: full credit retention"},

        {"col":"OPT","name":"Long Straddle","seg":"OPT","tfs":["1h","4h"],
         "logic":"Buy ATM call + ATM put. Profit if Nifty moves > IV-implied move. Use before events: RBI policy, Budget, elections, quarterly results.",
         "entry":"Enter 1-2 days before catalyst event. Buy when IV rank < 50%.",
         "sl_tgt":"SL: 30% of premium paid | T1: breakeven + 0.5% | T2: 2x premium paid"},

        {"col":"OPT","name":"Short Straddle","seg":"OPT","tfs":["1h","4h"],
         "logic":"Sell ATM CE + ATM PE. Maximum theta decay when market is pinned to a level. Paper trade only — unlimited risk both sides. Enter only when VIX < 15.",
         "entry":"VIX < 15 AND spot range < 0.5% over last 5 candles (market pinned). ATM strike only. Best: Tuesday–Wednesday for weekly expiry.",
         "sl_tgt":"SL: individual leg 2× entry premium triggers exit (e.g. CE sold ₹150 → SL if CE hits ₹300) OR spot breaches breakeven (strike ± total premium) | T1: 50% of credit retained | Paper only"},

        {"col":"OPT","name":"Long Strangle","seg":"OPT","tfs":["1h","4h"],
         "logic":"Buy OTM call + OTM put. Cheaper than straddle, needs bigger move. Use when expecting volatility spike but unsure of direction.",
         "entry":"Buy 1-2% OTM strikes. Enter when IV rank < 30%. Hold through event.",
         "sl_tgt":"SL: 40% of premium | T1: 1.5x premium | T2: 3x premium"},

        {"col":"OPT","name":"Short Strangle","seg":"OPT","tfs":["4h"],
         "logic":"Sell OTM call + OTM put. Wider range than short straddle. Collect premium from high IV contraction. Manage at 50% profit.",
         "entry":"Sell 1-2 SD OTM strikes. Enter after IV spike (IV rank > 80%). Weekly expiry.",
         "sl_tgt":"SL: 2x premium per side | T1: 50% decay | T2: expiry worthless"},

        {"col":"OPT","name":"Jade Lizard","seg":"OPT","tfs":["4h"],
         "logic":"Sell OTM put + sell OTM call spread. Collect enough premium so upside credit covers call spread width — no upside risk. Bullish bias.",
         "entry":"Select short put at strong support. Call spread width = 100-200 Nifty points.",
         "sl_tgt":"SL: put breaches — roll or close | T1: 50% premium | T2: expiry pin above short put"},

        {"col":"OPT","name":"Bull Call Spread","seg":"OPT","tfs":["1h","4h"],
         "logic":"Buy lower strike call + sell higher strike call. Reduces cost, defines max profit. Use when moderately bullish on Nifty/BankNifty.",
         "entry":"Buy ATM call, sell 0.5-1% OTM call. Enter on technical breakout confirmation.",
         "sl_tgt":"SL: 50% of debit paid | T1: short strike area | T2: full spread width"},

        {"col":"OPT","name":"Bear Put Spread","seg":"OPT","tfs":["1h","4h"],
         "logic":"Buy higher strike put + sell lower strike put. Defined risk bearish play. Use on Nifty breakdown below support with stop.",
         "entry":"Buy ATM put, sell 0.5% OTM put. Enter on technical breakdown.",
         "sl_tgt":"SL: 50% of debit | T1: short strike area | T2: full spread width"},

        {"col":"OPT","name":"Vertical Credit Spread","seg":"OPT","tfs":["4h"],
         "logic":"Sell OTM option + buy further OTM for protection. Directional bias with defined risk. Suitable for weekly Nifty/BankNifty expiry.",
         "entry":"Credit > 30% of spread width. Short strike at key S/R. Enter Mon/Tue for Thu expiry.",
         "sl_tgt":"SL: short strike breached | T1: 50% premium decay | T2: 80% decay"},

        {"col":"OPT","name":"Calendar Spread","seg":"OPT","tfs":["4h"],
         "logic":"Sell near-month ATM option + buy far-month same strike. Profit from IV differential and theta. Best when near-month IV > far-month.",
         "entry":"Enter when front-month IV > back-month by 5+ points. Use same strike ATM.",
         "sl_tgt":"SL: index moves > 1% from center | T1: near expiry IV collapse | T2: near expiry worthless"},

        {"col":"OPT","name":"Diagonal Spread","seg":"OPT","tfs":["4h"],
         "logic":"Long far-month slightly ITM option + short near-month OTM same type. Poor Man's Covered Call for index. Combines theta and delta.",
         "entry":"Buy 30-45 DTE 0.5% ITM, sell 7-14 DTE OTM. Delta target: net 0.3-0.4.",
         "sl_tgt":"SL: long option loses 30% of value | T1: short option expires worthless | T2: roll short monthly"},

        {"col":"OPT","name":"Ratio Spread","seg":"OPT","tfs":["4h"],
         "logic":"Buy 1 ATM call + sell 2 OTM calls (or put version). Collect credit if market stays below ratio-sold strikes. Undefined upside risk.",
         "entry":"1:2 ratio. Short strikes at strong resistance. Collect small net credit.",
         "sl_tgt":"SL: index closes above both short strikes | T1: credit retention at expiry | T2: close both at 80% decay"},

        {"col":"OPT","name":"Risk Reversal","seg":"OPT","tfs":["1h","4h"],
         "logic":"Sell OTM put + buy OTM call (bullish) or vice versa. Zero-cost or small credit/debit. Directional play with minimal capital outlay.",
         "entry":"Sell put at support, buy call at resistance. Balance strikes for zero-cost.",
         "sl_tgt":"SL: sold option doubles in value | T1: bought option 2x | T2: 3x bought option"},

        {"col":"OPT","name":"Synthetic Long","seg":"OPT","tfs":["1h","4h"],
         "logic":"Buy ATM call + sell ATM put = same P&L as owning futures at lower capital. Bullish directional with full upside, full downside.",
         "entry":"Enter when bullish on trend with IV rank < 50%. Use same expiry as trading plan.",
         "sl_tgt":"SL: put strike (max loss defined by put premium received) | T1: 1% index move | T2: 2% index move"},

        {"col":"OPT","name":"Put Skew Trade","seg":"OPT","tfs":["4h"],
         "logic":"Nifty put skew (higher IV in puts) means OTM puts are expensive. Sell put spreads to harvest skew premium. Mean-reversion on IV skew.",
         "entry":"When put skew (25D put IV - 25D call IV) > 3%. Sell OTM put spread.",
         "sl_tgt":"SL: 2x premium | T1: skew normalises | T2: expiry"},

        {"col":"OPT","name":"Covered Call","seg":"OPT","tfs":["4h"],
         "logic":"Hold stock position (or futures) + sell OTM call above. Earns premium against holding cost. Works in sideways/mild bullish market.",
         "entry":"Sell call at resistance or 1-2% OTM. Monthly expiry for stock positions.",
         "sl_tgt":"SL: buy back call if stock breaks above | T1: call expires worthless | T2: roll up on breakout"},

        {"col":"OPT","name":"Cash Secured Put","seg":"OPT","tfs":["4h"],
         "logic":"Sell OTM put on quality stock/index with cash collateral. Intent: earn premium or acquire stock at discount. Bullish income strategy.",
         "entry":"Sell put at strong support or 52-week low area. IV rank > 50%.",
         "sl_tgt":"SL: put assigned, hold position | T1: expires worthless | T2: roll down if breached"},

        {"col":"OPT","name":"Protective Put","seg":"OPT","tfs":["1h","4h"],
         "logic":"Hold futures long + buy ATM or OTM put as insurance. Limits downside during uncertainty. Cost = put premium.",
         "entry":"Buy put when holding overnight futures during major events (RBI, Fed, quarterly).",
         "sl_tgt":"SL: none (put is the hedge) | T1: futures move > put premium | T2: 2x put premium on futures gain"},

        {"col":"OPT","name":"Collar Strategy","seg":"OPT","tfs":["4h"],
         "logic":"Hold stock + sell OTM call + buy OTM put. Zero-cost collar = no net premium. Caps upside/downside. Useful for large stock positions.",
         "entry":"Put = 1-2% OTM below. Call = 1-2% OTM above. Match premium for zero-cost.",
         "sl_tgt":"SL: put provides floor | T1: call provides cap | T2: close or roll monthly"},

        {"col":"OPT","name":"ATM Call Momentum","seg":"OPT","tfs":["5m","15m"],
         "logic":"Buy ATM call when Nifty breaks intraday high with volume. Quick scalp using options for leverage. Delta ~0.5. Exit same day.",
         "entry":"Enter ATM call on ORB breakout or HOD break. Delta > 0.45 required.",
         "sl_tgt":"SL: 20% of premium | T1: 50% premium gain | T2: 100% premium gain"},

        {"col":"OPT","name":"ATM Put Momentum","seg":"OPT","tfs":["5m","15m"],
         "logic":"Buy ATM put on Nifty breakdown below intraday low or VWAP. Quick directional scalp. Same-day exit mandatory.",
         "entry":"Enter on 5-min close below VWAP or prior LOD with volume surge.",
         "sl_tgt":"SL: 20% of premium | T1: 50% premium gain | T2: 100% premium gain"},

        {"col":"OPT","name":"LEAPS Bull Call","seg":"OPT","tfs":["4h"],
         "logic":"Buy deep ITM call with 3+ month expiry on quality Nifty 50 stock. Low time-decay, high delta. Substitute for stock at 30-40% capital.",
         "entry":"Buy 80 delta call (deep ITM). Strike = stock price - 10%. 90+ DTE.",
         "sl_tgt":"SL: 20% loss on option | T1: stock reaches prior high | T2: 52-week high"},

        {"col":"OPT","name":"Wheel Strategy","seg":"OPT","tfs":["4h"],
         "logic":"Step 1: Sell cash-secured put. If assigned, Step 2: Sell covered calls until called away. Repeat. Generates regular income from quality stocks.",
         "entry":"Start by selling OTM put on strong stock at support. Target > 1% premium per month.",
         "sl_tgt":"SL: stock breaks below major support | T1: put expires worthless (repeat) | T2: called away above entry (profit)"},

        {"col":"OPT","name":"Poor Man's Covered Call","seg":"OPT","tfs":["4h"],
         "logic":"Buy long-dated ITM call (LEAPS) + sell short-dated OTM call monthly. Replaces stock in covered call at 30% capital. Diagonal spread.",
         "entry":"Buy 90+ DTE 70-delta call. Sell 30 DTE 30-delta call against it monthly.",
         "sl_tgt":"SL: long option loses 25% | T1: roll short monthly for income | T2: close full position at 20% gain"},

        # ===== FUTURES (25) =====
        {"col":"FUT","name":"Nifty Fibonacci PPT","seg":"FUT","tfs":["5m","15m","1h","4h"],
         "logic":"Gareth Soloway PPT adapted for Nifty futures. 61.8% retracement of prior impulse on E-mini equivalent (NF/BNF). High-precision reversal entry.",
         "entry":"Enter at 61.8% Fib level with reversal candle on Nifty/BankNifty futures. Volume contraction at retracement.",
         "sl_tgt":"SL: below 78.6% Fib | T1: prior swing | T2: 127% Fib extension"},

        {"col":"FUT","name":"Fibonacci Cluster","seg":"FUT","tfs":["1h","4h"],
         "logic":"Multiple Fib levels from different impulse legs coincide at same price zone. Cluster = high-probability S/R. Nifty futures respect clusters strongly.",
         "entry":"Mark 2-3 Fib sets. Enter at cluster zone (< 0.2% range) with reversal candle.",
         "sl_tgt":"SL: beyond cluster zone | T1: next Fib cluster | T2: measured move"},

        {"col":"FUT","name":"Opening Range Breakout NF","seg":"FUT","tfs":["5m","15m"],
         "logic":"Nifty futures first 15-min range (9:15-9:30 IST) defines day's bias. Breakout of this range with volume = high-conviction entry for rest of day.",
         "entry":"Enter on 5-min close above/below 15-min ORB with volume 2x+ average. No entries after 10:00 IST.",
         "sl_tgt":"SL: back inside ORB | T1: 1x ORB projected | T2: 2x ORB projected"},

        {"col":"FUT","name":"Overnight Range Break","seg":"FUT","tfs":["5m","15m"],
         "logic":"Nifty futures SGX/Gift range (previous close to 9:15 open). Breakout of this pre-market range after RTH open signals directional day.",
         "entry":"Enter on Gift Nifty range breakout sustained for 2 bars after 9:15 IST open.",
         "sl_tgt":"SL: back inside Gift range | T1: 1% from entry | T2: 2% from entry"},

        {"col":"FUT","name":"VWAP Reclaim Futures","seg":"FUT","tfs":["5m","15m"],
         "logic":"Nifty/BankNifty futures reclaim VWAP after dip below (bullish) or break below VWAP after bounce (bearish). Institutional reference level.",
         "entry":"Enter on 5-min close back above VWAP (long) or below VWAP (short). Require volume confirmation.",
         "sl_tgt":"SL: 0.2% beyond VWAP | T1: prior day high/low | T2: 1% from VWAP"},

        {"col":"FUT","name":"Nifty Gap Fill","seg":"FUT","tfs":["5m","15m"],
         "logic":"Nifty opens with gap vs prior close. Gaps < 0.5% tend to fill within first hour. Fade the opening gap direction for a gap-fill scalp.",
         "entry":"Enter fade on first 5-min reversal candle after open if gap < 0.5%. Volume should contract.",
         "sl_tgt":"SL: beyond gap candle | T1: VWAP | T2: prior close (gap fill)"},

        {"col":"FUT","name":"Gift Nifty High/Low Break","seg":"FUT","tfs":["5m"],
         "logic":"Gift Nifty (SGX equivalent) trades 6:30-9:15 IST. Its high and low become first key levels for Nifty futures at open. Breakout = momentum.",
         "entry":"Enter on break of Gift Nifty high/low within first 15 min of NSE open. Quick scalp.",
         "sl_tgt":"SL: back inside Gift range | T1: 50 Nifty points | T2: 100 Nifty points"},

        {"col":"FUT","name":"Nifty Momentum Scalp","seg":"FUT","tfs":["5m"],
         "logic":"Strong directional push (>0.3% in 5 min) with expanding volume. Enter pullback continuation. Works in morning session 9:15-11:00 IST.",
         "entry":"Enter on first 2-bar pullback after momentum thrust. Price must hold 9 EMA.",
         "sl_tgt":"SL: below 9 EMA (longs) | T1: 0.5% from entry | T2: 1% from entry"},

        {"col":"FUT","name":"RTH Open Drive","seg":"FUT","tfs":["5m","15m"],
         "logic":"NSE session opens (9:15-9:30 IST) with strong directional push that doesn't pause. Open drive = trend day likely. Hold position all session.",
         "entry":"Enter long/short if first 3 bars all same direction with volume. No contra-candle.",
         "sl_tgt":"SL: back below open candle | T1: +1% | T2: +1.5% (close at 3:00 IST)"},

        {"col":"FUT","name":"BankNifty Tech Momentum","seg":"FUT","tfs":["5m","15m"],
         "logic":"BankNifty is high-beta version of Nifty. Strong banking sector news (PSU, HDFC, ICICI) drives BNF momentum. Use for amplified directional trades.",
         "entry":"Enter BankNifty futures on confirmed ORB or VWAP reclaim with banking sector catalyst.",
         "sl_tgt":"SL: 0.3% | T1: 0.7% | T2: 1.5% (BNF moves bigger than Nifty)"},

        {"col":"FUT","name":"Sensex Divergence","seg":"FUT","tfs":["15m","1h"],
         "logic":"When Sensex (BSE) diverges from Nifty (NSE) — one makes new high while other doesn't — it signals distribution/rotation. Trade the weaker index short.",
         "entry":"Enter short on the lagging index future when divergence confirmed on 15-min close.",
         "sl_tgt":"SL: beyond divergence high | T1: VWAP of lagging index | T2: divergence resolved"},

        {"col":"FUT","name":"Reversal at VWAP","seg":"FUT","tfs":["5m","15m"],
         "logic":"Price extends far from VWAP (> 0.5%) then shows reversal candle returning to VWAP. Mean-reversion scalp toward VWAP. High win rate.",
         "entry":"Enter fade at 0.6%+ from VWAP on reversal candle. Target VWAP.",
         "sl_tgt":"SL: 0.3% beyond extreme | T1: VWAP | T2: VWAP ± 0.2%"},

        {"col":"FUT","name":"Cumulative Delta Divergence","seg":"FUT","tfs":["5m","15m"],
         "logic":"Price rises but cumulative delta (buy-sell volume) falls — bearish divergence. Indicates absorption at highs. Used on Nifty/BankNifty futures.",
         "entry":"Enter short when price makes new intraday high but cum-delta makes lower high on 5-min.",
         "sl_tgt":"SL: beyond price high | T1: VWAP | T2: POC"},

        {"col":"FUT","name":"Lunch Hour Fade","seg":"FUT","tfs":["15m"],
         "logic":"12:00-1:30 PM IST is low-volume lunch session in India. Trending moves during this period often reverse. Fade extremes at lunch highs/lows.",
         "entry":"Enter counter-trend at 12:30-1:00 PM IST if price is at intraday extreme with low volume.",
         "sl_tgt":"SL: beyond lunch extreme | T1: VWAP | T2: open price"},

        {"col":"FUT","name":"Power Hour Breakout","seg":"FUT","tfs":["5m","15m"],
         "logic":"2:00-3:30 PM IST is high-volume power hour in NSE. Breakouts of day's range in this period have strong follow-through. Enter late-day breakouts.",
         "entry":"Enter on 15-min close above/below day's HOD/LOD after 2:00 PM IST with volume surge.",
         "sl_tgt":"SL: back inside day range | T1: HOD + 0.5% | T2: close position at 3:25 IST"},

        {"col":"FUT","name":"Market Profile TPO","seg":"FUT","tfs":["1h","4h"],
         "logic":"TPO (Time Price Opportunity) chart shows value area (70% of volume). Trade from VAL (Value Area Low) to VAH (Value Area High) in balanced market.",
         "entry":"Enter long at VAL with acceptance (2 TPO periods). Enter short at VAH with rejection.",
         "sl_tgt":"SL: beyond VA boundary | T1: POC | T2: opposite VA boundary"},

        {"col":"FUT","name":"Nifty 50 EMA Support","seg":"FUT","tfs":["1h","4h"],
         "logic":"50 EMA on hourly/4H chart is major institutional support/resistance for Nifty futures. Price bounces from 50 EMA in trending markets.",
         "entry":"Enter long on 50 EMA touch with hammer/bullish engulfing on 1H. Trend must be up (above 200 EMA).",
         "sl_tgt":"SL: 0.3% below 50 EMA | T1: prior swing high | T2: 127% Fib extension"},

        {"col":"FUT","name":"BankNifty Relative Strength","seg":"FUT","tfs":["15m","1h"],
         "logic":"When BankNifty outperforms Nifty (BNF/NF ratio rising), go long BNF futures. When underperforming, go short BNF relative to Nifty.",
         "entry":"Enter BNF long when BNF/NF ratio breaks above 5-day high. Pair trade: long BNF short Nifty.",
         "sl_tgt":"SL: ratio reverses | T1: 0.5% BNF outperformance | T2: 1% BNF outperformance"},

        {"col":"FUT","name":"Nifty Inside Day","seg":"FUT","tfs":["4h"],
         "logic":"Day's range inside prior day's range = coiled energy. Next day's break of inside day high/low = powerful directional move. Position trade.",
         "entry":"Enter on break of inside day high (bullish) or low (bearish) on daily chart. Set alert at levels.",
         "sl_tgt":"SL: back inside inside day | T1: measured move = inside day range | T2: 2x measured move"},

        {"col":"FUT","name":"Volume Node Rejection","seg":"FUT","tfs":["1h","4h"],
         "logic":"Volume Profile shows high-volume nodes (HVN) as S/R and low-volume nodes (LVN) as price moves through quickly. Trade rejections at HVN.",
         "entry":"Enter when price reaches HVN and shows reversal candle. Volume should decline at HVN.",
         "sl_tgt":"SL: beyond HVN | T1: next LVN | T2: opposite HVN"},

        {"col":"FUT","name":"Nifty Trend Follow","seg":"FUT","tfs":["1h","4h"],
         "logic":"Classical trend following: 20 EMA > 50 EMA > 200 EMA (uptrend). Enter pullbacks to 20 EMA in bull trend. Ride medium-term directional moves.",
         "entry":"Enter on 20 EMA bounce with RSI > 50 in uptrend. Exit on 20 EMA break.",
         "sl_tgt":"SL: below 50 EMA | T1: channel top | T2: prior major swing"},

        {"col":"FUT","name":"MNF/MBNF Swing","seg":"FUT","tfs":["4h"],
         "logic":"Mini Nifty (MNF) and Mini BankNifty (MBNF) contracts for smaller capital. Swing trades 2-5 days. Same setups as regular contracts at 1/10th size.",
         "entry":"Enter mini contracts at same technical levels. Better for position sizing discipline.",
         "sl_tgt":"SL: same % as regular | T1: 1.5x ATR swing | T2: prior weekly swing"},

        {"col":"FUT","name":"FII DII Flow Follow","seg":"FUT","tfs":["4h"],
         "logic":"Track daily FII (Foreign) and DII (Domestic) institutional buying/selling from NSE data. Multi-day consistent FII buying = bullish signal for Nifty.",
         "entry":"Enter after 3+ consecutive days of FII net buying > 1000 crore. Confirm with Nifty above 20 EMA.",
         "sl_tgt":"SL: FII turns net seller | T1: 1% Nifty gain | T2: 2% Nifty gain"},

        {"col":"FUT","name":"India VIX Spike Reversal","seg":"FUT","tfs":["1h","4h"],
         "logic":"India VIX > 20 signals fear. VIX spike above 25 then reversal = buy Nifty. VIX mean-reverts quickly. High win rate after VIX spikes.",
         "entry":"Enter long Nifty futures when VIX falls from spike high by >10%. Require reversal candle on Nifty.",
         "sl_tgt":"SL: VIX makes new high | T1: VIX returns to 18 | T2: VIX returns to 14"},

        {"col":"FUT","name":"Budget/Event Straddle Unwind","seg":"FUT","tfs":["15m"],
         "logic":"Before major events (Union Budget, RBI Policy, FOMC), IV spikes. After event, IV crush. Short straddle/strangle day before, unwind day after.",
         "entry":"Enter short straddle/strangle 1 day before event. Exit within 1 hour after event announcement.",
         "sl_tgt":"SL: Nifty moves > 1.5% intraday | T1: IV crush 30% | T2: close 2 hours post event"},
    ]

    # ── Renderer ──────────────────────────────────────────────────────────────
    def _strat_card(s):
        _col_colors = {"EQ": ("#1d4ed8","#eff6ff"), "OPT": ("#d97706","#fffbeb"), "FUT": ("#16a34a","#f0fdf4")}
        _cc, _cbg = _col_colors.get(s["col"], ("#64748b","#f8fafc"))
        return (
            f'<div style="background:{_cbg};border:1px solid {_cc}33;border-left:3px solid {_cc};'
            f'border-radius:6px;padding:10px 14px;margin:6px 0;font-family:JetBrains Mono">'
            f'<div style="display:flex;justify-content:space-between;align-items:center">'
            f'<span style="color:{_cc};font-weight:700;font-size:12px">{s["name"]}</span>'
            f'<span style="background:{_cc};color:#fff;font-size:8px;padding:2px 6px;border-radius:4px">{s["seg"]}</span>'
            f'</div>'
            f'<div style="color:#475569;font-size:10px;margin-top:4px">TF: {" · ".join(s["tfs"])}</div>'
            f'<div style="color:#1e293b;font-size:10px;margin-top:6px"><b>Logic:</b> {s["logic"]}</div>'
            f'<div style="color:#1e293b;font-size:10px;margin-top:4px"><b>Entry:</b> {s["entry"]}</div>'
            f'<div style="color:#64748b;font-size:10px;margin-top:4px"><b>SL/Target:</b> {s["sl_tgt"]}</div>'
            f'</div>'
        )

    # ── Header banner ─────────────────────────────────────────────────────────
    _eq_n  = sum(1 for s in _IND_ALL_STRATS if s["col"]=="EQ")
    _opt_n = sum(1 for s in _IND_ALL_STRATS if s["col"]=="OPT")
    _fut_n = sum(1 for s in _IND_ALL_STRATS if s["col"]=="FUT")
    st.markdown(
        f'<div style="background:#dde9ff;border-radius:10px;padding:12px 18px;'
        f'margin-bottom:14px;font-family:JetBrains Mono">'
        f'<div style="color:#1d4ed8;font-size:18px;font-weight:700">India Strategies — {len(_IND_ALL_STRATS)} Strategies</div>'
        f'<div style="display:flex;gap:16px;margin-top:6px">'
        f'<span style="background:#1d4ed8;color:#fff;padding:3px 10px;border-radius:20px;font-size:11px">Equity {_eq_n}</span>'
        f'<span style="background:#d97706;color:#fff;padding:3px 10px;border-radius:20px;font-size:11px">Options {_opt_n}</span>'
        f'<span style="background:#16a34a;color:#fff;padding:3px 10px;border-radius:20px;font-size:11px">Futures {_fut_n}</span>'
        f'<span style="color:#5a72a0;font-size:11px;margin-top:3px">3 columns · 4 timeframes · Nifty · BankNifty · Sensex</span>'
        f'</div></div>',
        unsafe_allow_html=True)

    # ── TF sub-tabs ───────────────────────────────────────────────────────────
    _strat_tabs = st.tabs(["Overview", "5-Min Scalp", "15-Min Intraday", "1-Hour Swing", "4-Hour Position", "Backtest", "Bulk Backtest"])
    _tf_filter  = {"5-Min Scalp":"5m","15-Min Intraday":"15m","1-Hour Swing":"1h","4-Hour Position":"4h"}

    # ── Overview tab ─────────────────────────────────────────────────────────
    with _strat_tabs[0]:
        import plotly.express as _px_ov
        import plotly.graph_objects as _go_ov
        from plotly.subplots import make_subplots as _msp_ov

        _ov_rows = st.session_state.get('_bbk_trade_rows', [])

        if _ov_rows:
            import pandas as _pd_ov
            _ov_df = _pd_ov.DataFrame(_ov_rows)

            # normalise column names
            _seg_col  = 'Segment' if 'Segment' in _ov_df.columns else None
            _tf_col   = 'Timeframe' if 'Timeframe' in _ov_df.columns else None
            _strat_col= 'Strategy' if 'Strategy' in _ov_df.columns else None
            _won_col  = 'Won' if 'Won' in _ov_df.columns else None
            _pnl_col  = 'Net PnL ₹' if 'Net PnL ₹' in _ov_df.columns else None

            if _seg_col and _strat_col:
                st.markdown('<div style="color:#1d4ed8;font-size:15px;font-weight:700;font-family:JetBrains Mono;margin-bottom:6px">Strategy Performance Overview</div>', unsafe_allow_html=True)
                st.caption("Data from last Bulk Backtest run. Re-run Bulk Backtest to refresh.")

                # ── helper: build per-segment charts ─────────────────────────
                def _ov_seg_charts(seg_label, seg_key, color):
                    _seg_df = _ov_df[_ov_df[_seg_col] == seg_key].copy() if seg_key else _ov_df.copy()
                    if _seg_df.empty:
                        st.info(f"No {seg_label} trades in last backtest.")
                        return
                    if _won_col and _won_col in _seg_df.columns:
                        _seg_df['_won_int'] = _seg_df[_won_col].apply(lambda x: 1 if str(x).strip().lower() in ('true','1','yes','win') else 0)
                    else:
                        _seg_df['_won_int'] = 0
                    if _pnl_col and _pnl_col in _seg_df.columns:
                        _seg_df['_pnl_num'] = _pd_ov.to_numeric(_seg_df[_pnl_col], errors='coerce').fillna(0)
                    else:
                        _seg_df['_pnl_num'] = 0
                    _agg = _seg_df.groupby(_strat_col).agg(
                        Trades = (_strat_col, 'count'),
                        Wins   = ('_won_int',  'sum'),
                        PnL    = ('_pnl_num',  'sum'),
                    ).reset_index()
                    _agg['Win Rate %'] = (_agg['Wins'].astype(float) / _agg['Trades'].astype(float) * 100).round(1)
                    _agg['PnL'] = _agg['PnL'].round(0)
                    _agg = _agg.sort_values('Trades', ascending=False).head(20)

                    _col_bar, _col_pie = st.columns([3, 2])
                    with _col_bar:
                        _fig_bar = _go_ov.Figure()
                        _fig_bar.add_bar(x=_agg[_strat_col], y=_agg['Trades'],   name='Trades',    marker_color='#3b82f6', opacity=0.8)
                        _fig_bar.add_bar(x=_agg[_strat_col], y=_agg['Win Rate %'], name='Win Rate %', marker_color='#22c55e', opacity=0.8, yaxis='y2')
                        _fig_bar.update_layout(
                            title=dict(text=f"{seg_label} — Trades & Win Rate", font_size=12),
                            barmode='group', height=320,
                            legend=dict(orientation='h', y=1.1, x=0),
                            xaxis=dict(tickangle=-40, tickfont_size=9),
                            yaxis=dict(title='Trades', side='left'),
                            yaxis2=dict(title='Win Rate %', side='right', overlaying='y', range=[0,100]),
                            margin=dict(l=40, r=40, t=50, b=80),
                        )
                        st.plotly_chart(_fig_bar, width='stretch')

                        # PnL bar
                        _agg_pnl = _agg.sort_values('PnL', ascending=False)
                        _pnl_colors = ['#22c55e' if v >= 0 else '#ef4444' for v in _agg_pnl['PnL']]
                        _fig_pnl = _go_ov.Figure(_go_ov.Bar(
                            x=_agg_pnl[_strat_col], y=_agg_pnl['PnL'],
                            marker_color=_pnl_colors, name='Net PnL ₹'
                        ))
                        _fig_pnl.update_layout(
                            title=dict(text=f"{seg_label} — Net PnL ₹ by Strategy", font_size=12),
                            height=280, xaxis=dict(tickangle=-40, tickfont_size=9),
                            yaxis_title='Net PnL ₹', margin=dict(l=40, r=20, t=40, b=80),
                        )
                        st.plotly_chart(_fig_pnl, width='stretch')

                    with _col_pie:
                        _fig_pie = _px_ov.pie(
                            _agg, values='Trades', names=_strat_col,
                            title=f"{seg_label} — Trade Distribution",
                            color_discrete_sequence=_px_ov.colors.qualitative.Pastel,
                        )
                        _fig_pie.update_traces(textposition='inside', textinfo='percent+label', textfont_size=8)
                        _fig_pie.update_layout(height=320, showlegend=False, margin=dict(l=10, r=10, t=40, b=10))
                        st.plotly_chart(_fig_pie, width='stretch')

                        _fig_pie2 = _px_ov.pie(
                            _agg[_agg['PnL'] > 0] if (_agg['PnL'] > 0).any() else _agg,
                            values='Wins', names=_strat_col,
                            title=f"{seg_label} — Winning Trades Distribution",
                            color_discrete_sequence=_px_ov.colors.qualitative.Safe,
                        )
                        _fig_pie2.update_traces(textposition='inside', textinfo='percent+label', textfont_size=8)
                        _fig_pie2.update_layout(height=280, showlegend=False, margin=dict(l=10, r=10, t=40, b=10))
                        st.plotly_chart(_fig_pie2, width='stretch')

                # ── Segment sections ─────────────────────────────────────────
                _segs = [s for s in _ov_df[_seg_col].unique() if s]
                _seg_map = {}
                for _s in _segs:
                    _sl = str(_s).upper()
                    if 'EQ' in _sl or 'EQUITY' in _sl:
                        _seg_map['Equity'] = _s
                    elif 'OPT' in _sl or 'OPTION' in _sl:
                        _seg_map['Options'] = _s
                    elif 'FUT' in _sl or 'FUTURE' in _sl:
                        _seg_map['Futures'] = _s

                for _sname, _skey in _seg_map.items():
                    st.markdown(f'<div style="background:#f0f4ff;border-left:4px solid #3b82f6;padding:6px 12px;border-radius:0 6px 6px 0;font-weight:700;font-family:JetBrains Mono;font-size:12px;margin:12px 0 4px 0">{_sname}</div>', unsafe_allow_html=True)
                    _ov_seg_charts(_sname, _skey, '#3b82f6')

                # ── Per-Timeframe charts ──────────────────────────────────────
                if _tf_col:
                    st.markdown('<div style="background:#f0faf4;border-left:4px solid #22c55e;padding:6px 12px;border-radius:0 6px 6px 0;font-weight:700;font-family:JetBrains Mono;font-size:12px;margin:16px 0 4px 0">Performance by Timeframe</div>', unsafe_allow_html=True)
                    _tf_vals = sorted(_ov_df[_tf_col].dropna().unique(), key=lambda x: ['5m','15m','1h','4h'].index(x) if x in ['5m','15m','1h','4h'] else 99)
                    _tf_tabs_ov = st.tabs([f"{tf}" for tf in _tf_vals])
                    for _tfi, _tft in enumerate(_tf_tabs_ov):
                        with _tft:
                            _tf_name = _tf_vals[_tfi]
                            _tf_sub  = _ov_df[_ov_df[_tf_col] == _tf_name]
                            if _tf_sub.empty:
                                st.info(f"No trades for {_tf_name}")
                                continue
                            _tf_sub = _tf_sub.copy()
                            if _won_col and _won_col in _tf_sub.columns:
                                _tf_sub['_won_int'] = _tf_sub[_won_col].apply(lambda x: 1 if str(x).strip().lower() in ('true','1','yes','win') else 0)
                            else:
                                _tf_sub['_won_int'] = 0
                            if _pnl_col and _pnl_col in _tf_sub.columns:
                                _tf_sub['_pnl_num'] = _pd_ov.to_numeric(_tf_sub[_pnl_col], errors='coerce').fillna(0)
                            else:
                                _tf_sub['_pnl_num'] = 0
                            _tf_agg = _tf_sub.groupby(_strat_col).agg(
                                Trades = (_strat_col, 'count'),
                                Wins   = ('_won_int',  'sum'),
                                PnL    = ('_pnl_num',  'sum'),
                            ).reset_index()
                            _tf_agg['Win Rate %'] = (_tf_agg['Wins'].astype(float) / _tf_agg['Trades'].astype(float) * 100).round(1)
                            _tf_agg['PnL'] = _tf_agg['PnL'].round(0)
                            _tf_agg = _tf_agg.sort_values('PnL', ascending=False).head(20)

                            _tc1, _tc2 = st.columns([3, 2])
                            with _tc1:
                                _tfig = _go_ov.Figure()
                                _tfig.add_bar(x=_tf_agg[_strat_col], y=_tf_agg['Trades'],     name='Trades',    marker_color='#818cf8', opacity=0.85)
                                _tfig.add_bar(x=_tf_agg[_strat_col], y=_tf_agg['Win Rate %'], name='Win Rate %', marker_color='#f59e0b', opacity=0.85, yaxis='y2')
                                _tfig.update_layout(
                                    title=dict(text=f"{_tf_name} — Trades & Win Rate", font_size=12),
                                    barmode='group', height=300,
                                    legend=dict(orientation='h', y=1.1, x=0),
                                    xaxis=dict(tickangle=-40, tickfont_size=9),
                                    yaxis=dict(title='Trades', side='left'),
                                    yaxis2=dict(title='Win Rate %', side='right', overlaying='y', range=[0,100]),
                                    margin=dict(l=40, r=40, t=50, b=80),
                                )
                                st.plotly_chart(_tfig, width='stretch')

                                _pnl_c = ['#22c55e' if v >= 0 else '#ef4444' for v in _tf_agg['PnL']]
                                _tfig2 = _go_ov.Figure(_go_ov.Bar(x=_tf_agg[_strat_col], y=_tf_agg['PnL'], marker_color=_pnl_c, name='Net PnL ₹'))
                                _tfig2.update_layout(
                                    title=dict(text=f"{_tf_name} — Net PnL ₹", font_size=12),
                                    height=260, xaxis=dict(tickangle=-40, tickfont_size=9),
                                    yaxis_title='Net PnL ₹', margin=dict(l=40, r=20, t=40, b=80),
                                )
                                st.plotly_chart(_tfig2, width='stretch')

                            with _tc2:
                                _tpie = _px_ov.pie(
                                    _tf_agg, values='Trades', names=_strat_col,
                                    title=f"{_tf_name} — Trade Mix",
                                    color_discrete_sequence=_px_ov.colors.qualitative.Vivid,
                                )
                                _tpie.update_traces(textposition='inside', textinfo='percent+label', textfont_size=8)
                                _tpie.update_layout(height=300, showlegend=False, margin=dict(l=10, r=10, t=40, b=10))
                                st.plotly_chart(_tpie, width='stretch')

                                if _seg_col:
                                    _seg_agg = _tf_sub.groupby(_seg_col).agg(Trades=(_strat_col,'count')).reset_index()
                                    _spie = _px_ov.pie(
                                        _seg_agg, values='Trades', names=_seg_col,
                                        title=f"{_tf_name} — By Segment",
                                        color_discrete_sequence=['#3b82f6','#f59e0b','#22c55e'],
                                    )
                                    _spie.update_traces(textposition='inside', textinfo='percent+label', textfont_size=10)
                                    _spie.update_layout(height=260, showlegend=False, margin=dict(l=10, r=10, t=40, b=10))
                                    st.plotly_chart(_spie, width='stretch')
            else:
                st.info("No trade data found. Run the Bulk Backtest first to see charts here.")
        else:
            # ── Static overview from _IND_ALL_STRATS when no backtest run ────
            import plotly.express as _px_ov2
            import plotly.graph_objects as _go_ov2
            st.info("Run the Bulk Backtest tab to populate performance charts. Showing strategy distribution from the strategy library.")

            _ov_tfs = ['5m','15m','1h','4h']
            _ov_segs = [('Equity','EQ','#3b82f6'), ('Options','OPT','#f59e0b'), ('Futures','FUT','#22c55e')]

            _sc1, _sc2, _sc3 = st.columns(3)
            for _sci, (_sname, _skey, _scol) in enumerate(_ov_segs):
                _col_ref = [_sc1, _sc2, _sc3][_sci]
                with _col_ref:
                    _s_strats = [s for s in _IND_ALL_STRATS if s['col'] == _skey]
                    _tf_counts = {tf: sum(1 for s in _s_strats if tf in s.get('tfs',[])) for tf in _ov_tfs}
                    _fig_s = _go_ov2.Figure(_go_ov2.Bar(
                        x=list(_tf_counts.keys()), y=list(_tf_counts.values()),
                        marker_color=_scol, name=_sname
                    ))
                    _fig_s.update_layout(
                        title=dict(text=f"{_sname} strategies by TF", font_size=11),
                        height=220, margin=dict(l=20,r=10,t=40,b=20),
                        yaxis_title='Count',
                    )
                    st.plotly_chart(_fig_s, width='stretch')

            _all_names = [s['name'] for s in _IND_ALL_STRATS]
            _all_segs  = [s['col']  for s in _IND_ALL_STRATS]
            _seg_cnts  = {k: _all_segs.count(k) for k in ['EQ','OPT','FUT']}
            _fig_tot = _px_ov2.pie(
                values=list(_seg_cnts.values()),
                names=['Equity','Options','Futures'],
                title='Strategy Library — Segment Split',
                color_discrete_sequence=['#3b82f6','#f59e0b','#22c55e'],
            )
            _fig_tot.update_traces(textposition='inside', textinfo='percent+label', textfont_size=10)
            _fig_tot.update_layout(height=280, margin=dict(l=10,r=10,t=40,b=10))
            st.plotly_chart(_fig_tot, width='stretch')

    for _sti, _stt in enumerate(_strat_tabs[1:5]):
        with _stt:
            _tf_key = list(_tf_filter.values())[_sti]
            _tf_strats = [s for s in _IND_ALL_STRATS if _tf_key in s["tfs"]]
            _eq_s  = [s for s in _tf_strats if s["col"]=="EQ"]
            _opt_s = [s for s in _tf_strats if s["col"]=="OPT"]
            _fut_s = [s for s in _tf_strats if s["col"]=="FUT"]
            st.markdown(
                f'<div style="background:#f8fafc;border-radius:8px;padding:8px 14px;'
                f'margin-bottom:10px;font-family:JetBrains Mono;font-size:11px">'
                f'<span style="color:#1d4ed8;font-weight:700">{len(_tf_strats)} strategies on {_tf_key}</span>'
                f'&nbsp;·&nbsp;EQ:{len(_eq_s)} OPT:{len(_opt_s)} FUT:{len(_fut_s)}'
                f'</div>', unsafe_allow_html=True)
            _c1, _c2, _c3 = st.columns(3)
            with _c1:
                st.markdown('<div style="color:#1d4ed8;font-weight:700;font-size:11px;font-family:JetBrains Mono;margin-bottom:4px">EQUITY</div>', unsafe_allow_html=True)
                for _s in _eq_s:
                    st.markdown(_strat_card(_s), unsafe_allow_html=True)
            with _c2:
                st.markdown('<div style="color:#d97706;font-weight:700;font-size:11px;font-family:JetBrains Mono;margin-bottom:4px">OPTIONS</div>', unsafe_allow_html=True)
                for _s in _opt_s:
                    st.markdown(_strat_card(_s), unsafe_allow_html=True)
            with _c3:
                st.markdown('<div style="color:#16a34a;font-weight:700;font-size:11px;font-family:JetBrains Mono;margin-bottom:4px">FUTURES</div>', unsafe_allow_html=True)
                for _s in _fut_s:
                    st.markdown(_strat_card(_s), unsafe_allow_html=True)

    # ── Backtest tab ──────────────────────────────────────────────────────────
    with _strat_tabs[5]:
        st.markdown('<div style="color:#1d4ed8;font-size:16px;font-weight:700;font-family:JetBrains Mono">Strategy Backtester</div>', unsafe_allow_html=True)
        _bt_col1, _bt_col2 = st.columns([2,1])
        with _bt_col1:
            _bt_strat = st.selectbox("Select Strategy", [s["name"] for s in _IND_ALL_STRATS], key="ind_bt_strat")
            _bt_inst  = st.selectbox("Index", ["NIFTY","BANKNIFTY","SENSEX"], key="ind_bt_inst")
        with _bt_col2:
            _bt_capital = st.number_input("Capital (₹)", value=100000, step=10000, key="ind_bt_cap")
            _bt_lots    = st.number_input("Lots", value=1, min_value=1, key="ind_bt_lots")
        _bt_tf = st.radio("Timeframe", ["5m","15m","1h","4h"], horizontal=True, key="ind_bt_tf")

        # ── Segment + Expiry selector ─────────────────────────────────────────
        _bt_segment = st.radio("Segment", ["Equity", "Options", "Futures"],
                                horizontal=True, key="ind_bt_seg")

        _bt_selected_exps = []
        _bt_iv_pct = 20.0
        if _bt_segment in ("Options", "Futures") and OFS_OK:
            _bt_exp_fn  = _ofs_past_opt if _bt_segment == "Options" else _ofs_past_fut
            _bt_past_e  = _bt_exp_fn(_bt_inst, n=6, offset=0)
            _bt_e_labels = [e.strftime('%d %b %y  (%a)') for e in _bt_past_e]
            _seg_c1, _seg_c2 = st.columns([3, 1])
            with _seg_c1:
                _bt_e_sel = st.multiselect(
                    f"Expiries — {_bt_inst} ({'monthly' if _bt_segment == 'Options' and _bt_inst == 'BANKNIFTY' else 'weekly' if _bt_segment == 'Options' else 'monthly'})",
                    options=_bt_e_labels,
                    default=_bt_e_labels[:3],
                    key="ind_bt_exp_sel")
                _bt_selected_exps = [_bt_past_e[_bt_e_labels.index(l)] for l in _bt_e_sel if l in _bt_e_labels]
            with _seg_c2:
                if _bt_segment == "Options":
                    _bt_iv_pct = st.number_input("IV %", min_value=10, max_value=80, value=20, step=5,
                                                  key="ind_bt_iv")
                    _bt_rr_seg = st.selectbox("R:R", [2.0, 3.0],
                                              format_func=lambda x: f"1:{int(x)}", key="ind_bt_rr")
                else:
                    _bt_rr_seg = st.selectbox("R:R", [2.0, 3.0],
                                              format_func=lambda x: f"1:{int(x)}", key="ind_bt_rr_fut")
        else:
            _bt_rr_seg = 2.0

        # ── Strategy → backtester routing ────────────────────────────────────
        _BT_ROUTE = {
            # EQ — Fibonacci/reversal
            "PPT Fibonacci 61.8%":"FIB","PPT Tail Reversal":"FIB","Scalp Reversal at S/R":"FIB",
            "Volume Climax Reversal":"FIB","Doji Reversal":"FIB","Support Bounce":"FIB",
            "RSI Divergence":"FIB","MACD Histogram Reversal":"FIB","Volume Profile POC":"FIB",
            "Mean Reversion":"FIB","AI Narrative + Technical":"FIB",
            # EQ — Gap/intraday hunter
            "PPT Gap Fill":"HUNTER","Gap and Go":"HUNTER","Earnings Momentum":"HUNTER",
            "VWAP Reclaim":"HUNTER","Opening Range Breakout":"HUNTER","Pre-Market Range Break":"HUNTER",
            "High of Day Break":"HUNTER",
            # EQ — EMA-based
            "9/20 EMA Pullback":"5EMA","MA Crossover 20/50":"5EMA","Momentum Swing":"5EMA",
            "Bull Flag":"5EMA","Relative Strength Leader":"5EMA",
            # EQ — Inside/breakout
            "Inside Candle Breakout":"INSIDE","Breakout Momentum":"INSIDE",
            # EQ — Momentum/traffic
            "3-Bar Reversal":"TRAFFIC","Supertrend Signal":"TRAFFIC","Bollinger Band Squeeze":"TRAFFIC",
            # EQ — Round numbers
            "Round Number":"ROUND",
            # EQ — Fabio
            "Fabio Daily (Value Area)":"FABIO",
            # FUT — Fibonacci
            "Nifty Fibonacci PPT":"FIB","Fibonacci Cluster":"FIB","Reversal at VWAP":"FIB",
            "Cumulative Delta Divergence":"FIB","Lunch Hour Fade":"FIB","Market Profile TPO":"FIB",
            "Volume Node Rejection":"FIB","FII DII Flow Follow":"FIB","India VIX Spike Reversal":"FIB",
            # FUT — Hunter
            "Opening Range Breakout NF":"HUNTER","Overnight Range Break":"HUNTER",
            "VWAP Reclaim Futures":"HUNTER","Nifty Gap Fill":"HUNTER","Gift Nifty High/Low Break":"HUNTER",
            "RTH Open Drive":"HUNTER","Power Hour Breakout":"HUNTER","Sensex Divergence":"HUNTER",
            "Budget/Event Straddle Unwind":"HUNTER",
            # FUT — 5-min
            "Nifty Momentum Scalp":"5MIN",
            # FUT — EMA
            "BankNifty Tech Momentum":"5EMA","Nifty 50 EMA Support":"5EMA",
            "BankNifty Relative Strength":"5EMA","MNF/MBNF Swing":"5EMA",
            # FUT — Inside/Traffic
            "Nifty Inside Day":"INSIDE","Nifty Trend Follow":"TRAFFIC",
        }
        _OPT_STRATS = {
            "Iron Condor","Iron Butterfly","Long Butterfly","Broken Wing Butterfly",
            "Long Straddle","Short Straddle","Long Strangle","Short Strangle","Jade Lizard",
            "Bull Call Spread","Bear Put Spread","Vertical Credit Spread","Calendar Spread",
            "Diagonal Spread","Ratio Spread","Risk Reversal","Synthetic Long","Put Skew Trade",
            "Covered Call","Cash Secured Put","Protective Put","Collar Strategy",
            "ATM Call Momentum","ATM Put Momentum","LEAPS Bull Call","Wheel Strategy",
            "Poor Man's Covered Call",
        }
        _BT_LABELS = {
            "FIB":    "Fibonacci 61.8% + OI",
            "HUNTER": "Intraday Hunter (Gap/Divergence/Trap)",
            "5EMA":   "5/20 EMA Strategy",
            "INSIDE": "Inside Candle Breakout",
            "TRAFFIC":"Momentum Traffic Light",
            "ROUND":  "Round Number Reaction",
            "5MIN":   "5-Min Scalp (EMA)",
            "FABIO":  "Fabio Value Area",
        }

        def _display_bt_result(_res, _label):
            if isinstance(_res, dict):
                _candidates = [(k, v) for k, v in _res.items() if v and v.total_trades > 0]
                if not _candidates:
                    st.warning("No trades generated. Try a different instrument or check data availability.")
                    return
                _best_name, _best = max(_candidates, key=lambda x: x[1].total_trades)
                st.caption(f"Engine: {_label} — showing sub-strategy: {_best_name}")
                _display_single(_best)
            elif _res and _res.total_trades > 0:
                st.caption(f"Engine: {_label}")
                _display_single(_res)
            else:
                st.warning("No trades generated. Try a different instrument or check data availability.")

        def _display_single(_r):
            _c1, _c2, _c3, _c4 = st.columns(4)
            _c1.metric("Trades", _r.total_trades)
            _c2.metric("Win Rate", f"{_r.win_rate}%")
            _c3.metric("Total P&L (pts)", _r.total_pnl_pts)
            _c4.metric("Profit Factor", _r.profit_factor)
            _c1b, _c2b, _c3b = st.columns(3)
            _c1b.metric("Avg Win", f"{_r.avg_win_pts} pts")
            _c2b.metric("Avg Loss", f"{_r.avg_loss_pts} pts")
            _c3b.metric("Max Drawdown", f"{_r.max_drawdown_pts} pts")
            st.success(f"Backtest complete — {_r.total_trades} trades | Win Rate: {_r.win_rate}%")

        if st.button("Run Backtest", key="ind_bt_run", type="primary"):
            _bt_info = next((s for s in _IND_ALL_STRATS if s["name"]==_bt_strat), None)
            if _bt_info:
                st.info(f"Running backtest: {_bt_strat} on {_bt_inst} ({_bt_tf}) — {_bt_segment}")

                # ── Options / Futures: run StrategyBacktester per expiry ──────
                if _bt_segment in ("Options", "Futures") and OFS_OK:
                    if not _bt_selected_exps:
                        st.warning("Select at least one expiry date to run Options/Futures backtest.")
                    else:
                        _bt_ofs_mode = 'options' if _bt_segment == "Options" else 'futures'
                        _bt_ofs_tf   = _bt_tf if _bt_tf in ('5m','15m','1h') else '1h'
                        _bt_ofs_rows = []
                        _prog_bt = st.progress(0.0)
                        for _ei, _exp in enumerate(_bt_selected_exps):
                            _prog_bt.progress((_ei + 1) / len(_bt_selected_exps))
                            _window = 8 if _bt_ofs_mode == 'options' else 31
                            _pstart = _exp - timedelta(days=_window)
                            _btr_ofs = _OFSBacktester(_bt_inst)
                            for _ofs_s in _OFS_STRATEGIES:
                                try:
                                    _tr_ofs = _btr_ofs.run(
                                        _ofs_s, _bt_ofs_tf, _pstart, _exp,
                                        mode=_bt_ofs_mode, expiry_date=_exp,
                                        rr_mult=_bt_rr_seg,
                                        vix_override=float(_bt_iv_pct) / 100.0)
                                    for _t in _tr_ofs:
                                        _bt_ofs_rows.append({
                                            'Expiry':    str(_exp),
                                            'Strategy':  _ofs_s,
                                            'Dir':       _t.direction,
                                            'Entry':     _t.entry,
                                            'SL':        _t.sl,
                                            'Target':    _t.target,
                                            'Exit':      _t.exit_price,
                                            'Reason':    _t.exit_reason,
                                            'P&L pts':   round(_t.pnl_pts, 1),
                                            'P&L INR':   round(_t.pnl_inr, 0),
                                        })
                                except Exception:
                                    pass
                        _prog_bt.empty()
                        if _bt_ofs_rows:
                            import pandas as _pd_bt
                            _df_ofs = _pd_bt.DataFrame(_bt_ofs_rows)
                            _tot = _df_ofs['P&L INR'].sum()
                            _wins = (_df_ofs['P&L INR'] > 0).sum()
                            _n    = len(_df_ofs)
                            _c1b, _c2b, _c3b, _c4b = st.columns(4)
                            _c1b.metric("Trades", _n)
                            _c2b.metric("Win Rate", f"{round(_wins/_n*100,1)}%" if _n else "0%")
                            _c3b.metric("Net P&L", f"Rs{_tot:+,.0f}")
                            _c4b.metric("Segment", _bt_segment)
                            st.dataframe(_df_ofs, hide_index=True, use_container_width=True)
                        else:
                            st.warning("No trades generated for selected expiries/timeframe.")

                else:  # Equity segment — existing backtester path
                    if _bt_strat in _OPT_STRATS:
                        st.warning("Options strategies need live premium data for accurate backtesting. Showing Fibonacci proxy on underlying index.")
                    if BT_OK:
                        try:
                            _sym_map = {"NIFTY":"^NSEI","BANKNIFTY":"^NSEBANK","SENSEX":"^BSESN"}
                            _idx_map = {"NIFTY":"Nifty","BANKNIFTY":"BankNifty","SENSEX":"Sensex"}
                            _lmap    = {"NIFTY":75,"BANKNIFTY":35,"SENSEX":20}
                            _total_lot = _bt_lots * _lmap.get(_bt_inst, 50)
                            _sym = _sym_map.get(_bt_inst, "^NSEI")
                            _idx = _idx_map.get(_bt_inst, "Nifty")
                            _route = _BT_ROUTE.get(_bt_strat, "FIB")
                            _label = _BT_LABELS.get(_route, "Fibonacci 61.8% + OI")
                            if _route == "HUNTER":
                                _res = IntraHunterBacktester(days=30, lot_size=_total_lot).run()
                            elif _route == "5EMA":
                                _res = SubasishBacktester(symbol=_sym, days=365, lot_size=_total_lot, strategy="5EMA").run()
                            elif _route == "INSIDE":
                                _res = SubasishBacktester(symbol=_sym, days=365, lot_size=_total_lot, strategy="INSIDE").run()
                            elif _route == "TRAFFIC":
                                _res = SubasishBacktester(symbol=_sym, days=365, lot_size=_total_lot, strategy="TRAFFIC").run()
                            elif _route == "ROUND":
                                _res = SubasishBacktester(symbol=_sym, days=365, lot_size=_total_lot, strategy="ROUND").run()
                            elif _route == "5MIN":
                                _res = IntraDay5MinBacktester(index=_idx, days=55).run()
                            elif _route == "FABIO":
                                from backtester import FabioNSEIndexBacktester
                                _res = FabioNSEIndexBacktester(index=_idx, days=55).run()
                            else:  # FIB (default)
                                _res = FibOIBacktester(instrument=_bt_inst, days=30, lot_size=_total_lot).run()
                            _display_bt_result(_res, _label)
                        except Exception as _bte:
                            st.error(f"Backtest error: {_bte}")
                            st.markdown(
                                f'<div class="mcard">'
                                f'<div class="mlbl">{_bt_strat}</div>'
                                f'<div style="font-size:11px;color:#475569;font-family:JetBrains Mono">'
                                f'<b>Logic:</b> {_bt_info["logic"]}<br>'
                                f'<b>Entry:</b> {_bt_info["entry"]}<br>'
                                f'<b>SL/Target:</b> {_bt_info["sl_tgt"]}'
                                f'</div></div>',
                                unsafe_allow_html=True)
                    else:
                        st.markdown(
                            f'<div class="mcard">'
                            f'<div class="mlbl">{_bt_strat}</div>'
                            f'<div style="font-size:11px;color:#475569;font-family:JetBrains Mono">'
                            f'<b>Logic:</b> {_bt_info["logic"]}<br>'
                            f'<b>Entry:</b> {_bt_info["entry"]}<br>'
                            f'<b>SL/Target:</b> {_bt_info["sl_tgt"]}'
                            f'</div></div>',
                            unsafe_allow_html=True)


    # ── Bulk Backtest tab ─────────────────────────────────────────────────────
    with _strat_tabs[6]:
        import io as _bbk_io
        from datetime import datetime as _bbk_dt

        st.markdown(
            '<div style="color:#1d4ed8;font-size:16px;font-weight:700;font-family:JetBrains Mono;margin-bottom:4px">'
            'Bulk Backtest — All Strategies × All Symbols × All Timeframes → Excel</div>',
            unsafe_allow_html=True)
        st.caption("Runs all implemented strategies (4 per timeframe) via MultiTFBacktester for any mix of indices and Nifty 50 stocks.")

        _bc1, _bc2, _bc3, _bc4 = st.columns([1,1,1,2])
        with _bc1:
            _bbk_tfs  = st.multiselect("Timeframes", ["5M","15M","1H","4H"],
                                        default=["5M","15M","1H","4H"], key="bbk_tfs")
        with _bc2:
            _bbk_days = st.selectbox("Period", [30,60,90,120,150,180],
                                     format_func=lambda x: f"{x} days", key="bbk_days")
        with _bc3:
            _bbk_idx  = st.multiselect("Indices", ["NIFTY","BANKNIFTY","SENSEX"],
                                        default=["NIFTY","BANKNIFTY","SENSEX"], key="bbk_idx")
        with _bc4:
            _bbk_stocks = st.multiselect("Nifty 50 Stocks", NSE_TOP_STOCKS,
                                          default=NSE_TOP_STOCKS, key="bbk_stocks",
                                          placeholder="Add Nifty 50 stocks…")

        _bbk_tfs_prev   = _bbk_tfs or ["5M"]
        _bbk_strat_n    = sum(len(TF_STRATEGIES.get(tf, [])) for tf in _bbk_tfs_prev)
        _bbk_sym_n      = len(_bbk_idx) + len(_bbk_stocks if _bbk_stocks else [])
        _bbk_total_runs = _bbk_strat_n * _bbk_sym_n
        _bbk_warn  = "  ⚠️ 5m/15m data capped at 60 days" if any(t in _bbk_tfs_prev for t in ["5M","15M"]) and _bbk_days > 60 else ""
        st.markdown(
            f'<div style="background:#f0f6ff;border:1px solid #bfdbfe;border-radius:8px;padding:8px 16px;'
            f'margin-bottom:10px;font-family:JetBrains Mono;font-size:12px">'
            f'<b>{_bbk_strat_n}</b> strategy×TF combos · '
            f'<b>{_bbk_sym_n}</b> symbols ({len(_bbk_idx)} indices + {len(_bbk_stocks or [])} stocks) · '
            f'<b>{_bbk_total_runs}</b> total backtests | Period: <b>{_bbk_days} days</b>{_bbk_warn}</div>',
            unsafe_allow_html=True)

        if st.button("▶ Run Bulk Backtest & Export Excel", key="bbk_run", type="primary"):
            if not MULTI_TF_OK:
                st.error("india_all_strategies_engine.py not found.")
            elif not _bbk_idx and not _bbk_stocks:
                st.warning("Select at least one index or stock.")
            elif not _bbk_tfs:
                st.warning("Select at least one timeframe.")
            else:
                import pandas as _bbk_pd
                import io as _bbk_io
                from datetime import datetime as _bbk_dt

                # Build symbol universe: selected indices + selected stocks
                _IDX_YF  = {"NIFTY":"^NSEI","BANKNIFTY":"^NSEBANK","SENSEX":"^BSESN"}
                _IDX_LOT = {"NIFTY":75,"BANKNIFTY":35,"SENSEX":20}
                _bbk_universe = {}   # symbol → {yf, lot, seg}
                for _xi in (_bbk_idx or []):
                    _bbk_universe[_xi] = {"yf": _IDX_YF.get(_xi,"^NSEI"), "lot": _IDX_LOT.get(_xi,75), "seg":"INDEX"}
                for _sk in (_bbk_stocks or []):
                    _bbk_universe[_sk] = {"yf": NSE_YF_MAP.get(_sk, _sk+".NS"), "lot": 1, "seg":"EQUITY"}

                _all_summary_rows, _all_trade_rows = [], []
                _total_runs = sum(len(TF_STRATEGIES.get(tf,[])) for tf in _bbk_tfs) * len(_bbk_universe)
                _prog   = st.progress(0, text="Initialising…")
                _status = st.empty()
                _done   = 0

                for _bbk_tf in _bbk_tfs:
                    _eff      = min(_bbk_days, 58) if _bbk_tf in ["5M","15M"] else _bbk_days
                    _bt_eng   = MultiTFBacktester(tf=_bbk_tf, days=_eff)
                    _strats   = TF_STRATEGIES.get(_bbk_tf, [])

                    for _sym, _smeta in _bbk_universe.items():
                        _yft = _smeta["yf"]; _lot = _smeta["lot"]; _seg = _smeta["seg"]
                        for _strat in _strats:
                            _done += 1
                            _prog.progress(_done / max(_total_runs, 1),
                                           text=f"[{_bbk_tf}] {_sym} × {_strat} ({_done}/{_total_runs})")
                            try:
                                _res = _bt_eng._bt_symbol(_yft, _sym, _strat, _lot)
                            except Exception:
                                _res = None
                            if _res and _res.total_trades > 0:
                                _all_summary_rows.append({
                                    'Timeframe': _bbk_tf, 'Symbol': _sym, 'Segment': _seg,
                                    'Strategy': _strat,
                                    'Trades':   _res.total_trades,
                                    'Wins':     _res.winning,
                                    'Win%':     round(_res.win_rate, 1),
                                    'PF':       round(_res.profit_factor, 2),
                                    'PnL Pts':  round(_res.total_pnl_pts, 2),
                                    'PnL ₹':    round(_res.total_pnl_rs, 2),
                                    'Max DD%':  round(_res.max_drawdown, 2),
                                })
                                for _t in _res.trades:
                                    _all_trade_rows.append({
                                        'Date':       _t.date,
                                        'Index':      _sym,
                                        'Segment':    _seg,
                                        'Timeframe':  _bbk_tf,
                                        'Strategy':   _strat,
                                        'Direction':  _t.direction,
                                        'Entry':      _t.entry,
                                        'Target':     _t.target,
                                        'Stop':       _t.stop,
                                        'Exit':       _t.exit_price,
                                        'Exit Reason':_t.exit_reason,
                                        'PnL Pts':    _t.pnl_pts,
                                        'PnL%':       _t.pnl_pct,
                                        'RR':         _t.rr,
                                        'Won':        _t.won,
                                    })

                _prog.empty(); _status.empty()

                # ── Display summary ───────────────────────────────────────────────
                _df_sum = _bbk_pd.DataFrame(_all_summary_rows) if _all_summary_rows else _bbk_pd.DataFrame()
                _df_trd = _bbk_pd.DataFrame(_all_trade_rows)   if _all_trade_rows   else _bbk_pd.DataFrame()

                if not _df_sum.empty:
                    st.markdown(f"**{len(_all_summary_rows)} strategy×symbol combos with trades | {len(_all_trade_rows)} total trades**")
                    _res_tabs = st.tabs(_bbk_tfs + ["All"])
                    for _ti, _rtf in enumerate(_bbk_tfs):
                        with _res_tabs[_ti]:
                            _dft = _df_sum[_df_sum['Timeframe']==_rtf]
                            if not _dft.empty:
                                st.dataframe(_dft.sort_values('PnL Pts', ascending=False).reset_index(drop=True),
                                             height=420, width='stretch')
                            else:
                                st.info("No trades for this timeframe.")
                    with _res_tabs[-1]:
                        st.dataframe(_df_sum.sort_values('PnL Pts', ascending=False).reset_index(drop=True),
                                     height=520, width='stretch')
                else:
                    st.warning("No trades found. Try a longer period or different symbols/timeframes.")

                # ── Excel export ──────────────────────────────────────────────────
                if not _df_sum.empty or not _df_trd.empty:
                    _xl = _bbk_io.BytesIO()
                    try:
                        with _bbk_pd.ExcelWriter(_xl, engine='openpyxl') as _xw:
                            if not _df_sum.empty:
                                _df_sum.sort_values(['Timeframe','PnL Pts'], ascending=[True,False]).to_excel(_xw, sheet_name='Summary', index=False)
                            if not _df_trd.empty:
                                _df_trd.to_excel(_xw, sheet_name='All Trades', index=False)
                            for _btf3 in _bbk_tfs:
                                _dft3 = _df_trd[_df_trd['Timeframe']==_btf3] if not _df_trd.empty else _bbk_pd.DataFrame()
                                if not _dft3.empty:
                                    _dft3.to_excel(_xw, sheet_name=f'Trades_{_btf3}', index=False)
                            for _xi2 in list(_bbk_idx or []) + list(_bbk_stocks or []):
                                _dfi2 = _df_trd[_df_trd['Index']==_xi2] if not _df_trd.empty else _bbk_pd.DataFrame()
                                if not _dfi2.empty:
                                    _dfi2.to_excel(_xw, sheet_name=f'{_xi2[:28]}', index=False)
                            _map_rows = []
                            for _tf2 in _bbk_tfs:
                                for _s2 in TF_STRATEGIES.get(_tf2, []):
                                    _desc = STRATEGY_DESCRIPTIONS.get(_s2, ("",""))
                                    _map_rows.append({'TF':_tf2,'Code':_s2,'Name':_desc[0],'Description':_desc[1]})
                            _bbk_pd.DataFrame(_map_rows).to_excel(_xw, sheet_name='Strategy_Map', index=False)
                        _tfs_str = "+".join(_bbk_tfs)
                        _fname = f"india_bulk_bt_{_tfs_str}_{_bbk_days}d_{_bbk_dt.now().strftime('%Y%m%d_%H%M')}.xlsx"
                        st.download_button(label=f"⬇ Download Excel — {_fname}",
                                           data=_xl.getvalue(), file_name=_fname,
                                           mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                           type='primary', key="bbk_dl")
                    except ImportError:
                        st.error("openpyxl not installed. Run: pip install openpyxl")

                # Persist for chart section
                st.session_state['_bbk_trade_rows'] = _all_trade_rows
                st.session_state['_bbk_indices']    = list(_bbk_idx or []) + list(_bbk_stocks or [])
                st.session_state['_bbk_tfs_run']    = _bbk_tfs


        # ── Trade Chart Section (persists after backtest runs) ─────────────────
        _bbk_tr_stored = st.session_state.get('_bbk_trade_rows', [])
        if _bbk_tr_stored:
            st.markdown("---")
            st.markdown(
                '<div style="color:#1d4ed8;font-size:14px;font-weight:700;'
                'font-family:JetBrains Mono;margin-bottom:4px">'
                '📊 Trade Chart — Day-wise Candlestick with Strategy Markers</div>',
                unsafe_allow_html=True)
            st.caption("Select a date and index to see all trades taken that day with strategy labels overlaid on the price chart.")

            # Collect unique dates and indices from stored trade rows
            _chart_dates   = sorted(set(str(t.get('Date',''))[:10] for t in _bbk_tr_stored if t.get('Date')), reverse=True)
            _chart_indices = sorted(set(t.get('Index','') for t in _bbk_tr_stored if t.get('Index')))
            _chart_tfs     = sorted(set(t.get('Timeframe','') for t in _bbk_tr_stored if t.get('Timeframe')))

            _cht_c1, _cht_c2, _cht_c3 = st.columns(3)
            with _cht_c1:
                _cht_date  = st.selectbox("Select Date", _chart_dates, key="cht_date")
            with _cht_c2:
                _cht_idx   = st.selectbox("Select Index", _chart_indices, key="cht_idx")
            with _cht_c3:
                _cht_tf    = st.selectbox("Timeframe", _chart_tfs if _chart_tfs else ["5m","15m","1h"], key="cht_tf")

            if _cht_date and _cht_idx:
                # Filter trades for selected date + index
                _day_trades = [t for t in _bbk_tr_stored
                               if str(t.get('Date',''))[:10] == _cht_date
                               and t.get('Index','') == _cht_idx
                               and t.get('Timeframe','') == _cht_tf]

                try:
                    import yfinance as _cht_yf
                    import plotly.graph_objects as _pgo
                    from plotly.subplots import make_subplots as _msp
                    import pandas as _cht_pd

                    _cht_sym_map = {"NIFTY":"^NSEI","BANKNIFTY":"^NSEBANK","SENSEX":"^BSESN"}
                    _cht_sym     = _cht_sym_map.get(_cht_idx, NSE_YF_MAP.get(_cht_idx, _cht_idx + ".NS"))
                    _cht_tf_map  = {"5M":"5m","15M":"15m","1H":"1h","4H":"1h","5m":"5m","15m":"15m","1h":"1h","4h":"1h"}
                    _cht_yf_tf   = _cht_tf_map.get(_cht_tf, "5m")

                    # Fetch ±1 day around selected date for context
                    try:
                        _cht_start = str(_cht_date)
                        _cht_end   = (_cht_pd.Timestamp(_cht_start) + _cht_pd.Timedelta(days=2)).strftime('%Y-%m-%d')
                        _cht_df    = _cht_yf.download(_cht_sym, start=_cht_start, end=_cht_end,
                                                       interval=_cht_yf_tf, progress=False, auto_adjust=True)
                    except Exception:
                        _cht_df = None

                    if _cht_df is not None and len(_cht_df) > 2:
                        # Filter to just the selected day (by date part of index)
                        _cht_df.index = _cht_pd.DatetimeIndex(_cht_df.index)
                        _cht_day_df   = _cht_df[_cht_df.index.date == _cht_pd.Timestamp(_cht_date).date()]
                        if _cht_day_df.empty:
                            _cht_day_df = _cht_df  # fallback: show all data

                        # Build subplots: price + volume
                        _fig = _msp(rows=2, cols=1, shared_xaxes=True,
                                    row_heights=[0.75, 0.25], vertical_spacing=0.03)

                        # Candlestick
                        _fig.add_trace(_pgo.Candlestick(
                            x=_cht_day_df.index,
                            open=_cht_day_df['Open'].squeeze(),
                            high=_cht_day_df['High'].squeeze(),
                            low=_cht_day_df['Low'].squeeze(),
                            close=_cht_day_df['Close'].squeeze(),
                            name='Price', increasing_line_color='#16a34a',
                            decreasing_line_color='#dc2626', showlegend=False,
                        ), row=1, col=1)

                        # Volume bars
                        if 'Volume' in _cht_day_df.columns:
                            _vol_colors = ['#16a34a' if c >= o else '#dc2626'
                                           for c, o in zip(_cht_day_df['Close'].squeeze(),
                                                           _cht_day_df['Open'].squeeze())]
                            _fig.add_trace(_pgo.Bar(
                                x=_cht_day_df.index,
                                y=_cht_day_df['Volume'].squeeze(),
                                name='Volume', marker_color=_vol_colors,
                                opacity=0.6, showlegend=False,
                            ), row=2, col=1)

                        # VWAP line
                        try:
                            _vwap_num = (_cht_day_df['Close'].squeeze() *
                                         _cht_day_df['Volume'].squeeze()).cumsum()
                            _vwap_den = _cht_day_df['Volume'].squeeze().cumsum()
                            _vwap = _vwap_num / _vwap_den
                            _fig.add_trace(_pgo.Scatter(
                                x=_cht_day_df.index, y=_vwap,
                                name='VWAP', line=dict(color='#7c3aed', width=1.5, dash='dot'),
                                showlegend=True,
                            ), row=1, col=1)
                        except Exception:
                            pass

                        # Trade markers
                        _seg_color = {'EQ':'#1d4ed8','FUT':'#059669','OPT':'#d97706'}
                        _dir_sym   = {'LONG':'triangle-up','SHORT':'triangle-down'}
                        _dir_col   = {'LONG':'#16a34a','SHORT':'#dc2626'}

                        for _tr in _day_trades:
                            _tr_ep  = float(_tr.get('Entry', 0) or 0)
                            _tr_xp  = float(_tr.get('Exit Price', 0) or 0)
                            _tr_dir = str(_tr.get('Direction','LONG'))
                            _tr_seg = str(_tr.get('Segment','EQ'))
                            _tr_strat = str(_tr.get('Strategies', _tr.get('Strategy','?')) or '?')[:30]
                            _tr_won = str(_tr.get('Won','NO'))
                            _tr_pnl = _tr.get('PnL pts', 0) or 0
                            _tr_chg = _tr.get('Total Charges', 0) or 0
                            _tr_net = _tr.get('Net PnL ₹', 0) or 0
                            _marker_sym = _dir_sym.get(_tr_dir, 'circle')
                            _entry_col  = _dir_col.get(_tr_dir, '#1d4ed8')
                            _exit_col   = '#16a34a' if _tr_won == 'YES' else '#dc2626'

                            # Entry marker
                            if _tr_ep > 0:
                                _fig.add_trace(_pgo.Scatter(
                                    x=[_cht_pd.Timestamp(_cht_date + ' 09:15:00')],
                                    y=[_tr_ep],
                                    mode='markers+text',
                                    marker=dict(symbol=_marker_sym, size=14,
                                                color=_entry_col,
                                                line=dict(color='white', width=1.5)),
                                    text=[f"▶ {_tr_seg} {_tr_dir[:1]}"],
                                    textposition='top center',
                                    textfont=dict(size=8, color=_entry_col),
                                    name=f"Entry: {_tr_strat[:20]}",
                                    hovertemplate=(
                                        f"<b>ENTRY</b><br>"
                                        f"Strategy: {_tr_strat}<br>"
                                        f"Seg: {_tr_seg} | Dir: {_tr_dir}<br>"
                                        f"Entry: {_tr_ep:.2f}<br>"
                                        f"Target: {_tr.get('Target','?')} | SL: {_tr.get('Stop','?')}<br>"
                                        f"<extra></extra>"
                                    ),
                                    showlegend=True,
                                ), row=1, col=1)

                            # Exit marker
                            if _tr_xp > 0:
                                _fig.add_trace(_pgo.Scatter(
                                    x=[_cht_pd.Timestamp(_cht_date + ' 15:15:00')],
                                    y=[_tr_xp],
                                    mode='markers+text',
                                    marker=dict(symbol='x', size=12,
                                                color=_exit_col,
                                                line=dict(color='white', width=1)),
                                    text=[f"{'✓' if _tr_won=='YES' else '✗'} {_tr_pnl:+.0f}pt"],
                                    textposition='bottom center',
                                    textfont=dict(size=8, color=_exit_col),
                                    name=f"Exit ({_tr.get('Exit Reason','?')})",
                                    hovertemplate=(
                                        f"<b>EXIT</b> — {_tr.get('Exit Reason','?')}<br>"
                                        f"Exit Price: {_tr_xp:.2f}<br>"
                                        f"PnL: {_tr_pnl:+.1f} pts<br>"
                                        f"Charges: ₹{_tr_chg:.2f}<br>"
                                        f"Net PnL: ₹{_tr_net:+.0f}<br>"
                                        f"<extra></extra>"
                                    ),
                                    showlegend=True,
                                ), row=1, col=1)

                            # Horizontal lines for target and stop
                            for _lvl, _lvl_col, _lvl_name in [
                                (_tr.get('Target'), '#16a34a', 'T'),
                                (_tr.get('Stop'),   '#dc2626', 'SL'),
                            ]:
                                try:
                                    _lvl_f = float(_lvl)
                                    if _lvl_f > 0:
                                        _fig.add_hline(y=_lvl_f, line_dash='dash',
                                                       line_color=_lvl_col, line_width=0.8,
                                                       opacity=0.5, row=1, col=1)
                                except Exception:
                                    pass

                        _fig.update_layout(
                            title=dict(
                                text=f"{_cht_idx} · {_cht_tf} · {_cht_date}  —  {len(_day_trades)} trade(s)",
                                font=dict(size=13, family='JetBrains Mono'), x=0.01,
                            ),
                            height=600, xaxis_rangeslider_visible=False,
                            plot_bgcolor='#f8fafc', paper_bgcolor='white',
                            legend=dict(orientation='h', y=-0.12, font=dict(size=9)),
                            margin=dict(l=40, r=20, t=50, b=40),
                            xaxis2=dict(title='Time (IST)'),
                            yaxis=dict(title=f'{_cht_idx} Price (pts)', gridcolor='#e2e8f0'),
                            yaxis2=dict(title='Volume', gridcolor='#e2e8f0'),
                        )
                        st.plotly_chart(_fig, width='stretch')

                        # Trade detail cards below chart
                        if _day_trades:
                            st.markdown(
                                f'<div style="font-family:JetBrains Mono;font-size:11px;'
                                f'font-weight:700;color:#1d4ed8;margin:8px 0">'
                                f'{len(_day_trades)} trade(s) on {_cht_date} — {_cht_idx} [{_cht_tf}]</div>',
                                unsafe_allow_html=True)
                            _card_html = '<div style="display:flex;flex-wrap:wrap;gap:8px;margin-bottom:12px">'
                            for _tr2 in _day_trades:
                                _tr2_won = str(_tr2.get('Won','NO'))
                                _card_bg = '#f0fdf4' if _tr2_won=='YES' else '#fff1f2'
                                _card_bc = '#16a34a' if _tr2_won=='YES' else '#dc2626'
                                _tr2_net = _tr2.get('Net PnL ₹', 0) or 0
                                _tr2_chg = _tr2.get('Total Charges', 0) or 0
                                _tr2_inv = _tr2.get('Investment ₹', 0) or 0
                                _tr2_ang = _tr2.get('Angel Brokerage', 0) or 0
                                _tr2_stt = _tr2.get('STT', 0) or 0
                                _tr2_strat = str(_tr2.get('Strategies', _tr2.get('Strategy','?')) or '?')
                                _card_html += (
                                    f'<div style="background:{_card_bg};border:1.5px solid {_card_bc};'
                                    f'border-radius:8px;padding:10px 14px;min-width:260px;'
                                    f'font-family:JetBrains Mono;font-size:11px">'
                                    f'<div style="font-weight:700;color:{_card_bc};margin-bottom:4px">'
                                    f'{_tr2.get("Segment","?")} · {_tr2.get("Direction","?")} · {_tr2.get("Exit Reason","?")}</div>'
                                    f'<div style="color:#1e293b;margin-bottom:2px"><b>Strategy:</b> {_tr2_strat[:35]}</div>'
                                    f'<div style="color:#64748b;margin-bottom:2px">'
                                    f'Entry: <b>{_tr2.get("Entry","?")}</b> → '
                                    f'Exit: <b>{_tr2.get("Exit Price","?")}</b> · '
                                    f'Lot: {_tr2.get("Lot","?")}</div>'
                                    f'<div style="color:#64748b;margin-bottom:4px">'
                                    f'PnL: <b style="color:{_card_bc}">{_tr2.get("PnL pts",0):+.1f} pts</b> · '
                                    f'Invested: ₹{_tr2_inv:,.0f}</div>'
                                    f'<div style="background:#f1f5f9;border-radius:4px;padding:4px 8px;font-size:10px">'
                                    f'Angel: ₹{_tr2_ang:.0f} | STT: ₹{_tr2_stt:.2f} | '
                                    f'Chg: ₹{_tr2_chg:.2f} | '
                                    f'<b>Net: ₹{_tr2_net:+,.0f}</b></div>'
                                    f'</div>'
                                )
                            _card_html += '</div>'
                            st.markdown(_card_html, unsafe_allow_html=True)
                    else:
                        st.warning(f"No intraday data available for {_cht_idx} on {_cht_date}. Market may have been closed or data unavailable.")
                except ImportError:
                    st.error("plotly not installed. Run: pip install plotly")
                except Exception as _cht_err:
                    st.error(f"Chart error: {_cht_err}")


def _plot_bt_trade_chart(t):
    """Plotly candlestick for a backtested TradeResult dict/namespace. Uses t.instrument + t.tf."""
    try:
        import yfinance as _yf_bch
        import plotly.graph_objects as _go_bch
        _sym_map = {"NIFTY": "^NSEI", "BANKNIFTY": "^NSEBANK", "SENSEX": "^BSESN"}
        _sym  = _sym_map.get(getattr(t, 'instrument', 'NIFTY'), "^NSEI")
        _tf_v = getattr(t, 'tf', '15m')
        _days = {"5m": 4, "15m": 8, "1h": 20}.get(_tf_v, 5)
        try:
            from datetime import datetime as _dtb, timedelta as _tdb
            _td    = _dtb.strptime(str(getattr(t, 'date', '')), "%Y-%m-%d").date()
        except Exception:
            from datetime import date as _dtb2, timedelta as _tdb
            _td = _dtb2.today()
        _start = _td - _tdb(days=_days)
        _end   = _td + _tdb(days=3)
        _df_b  = _yf_bch.Ticker(_sym).history(start=str(_start), end=str(_end),
                                               interval=_tf_v, auto_adjust=True)
        if _df_b is None or len(_df_b) < 5:
            return None
        _df_b.index = pd.to_datetime(_df_b.index)
        try:
            _df_b.index = _df_b.index.tz_convert(None)
        except Exception:
            try:
                _df_b.index = _df_b.index.tz_localize(None)
            except Exception:
                pass
        _entry_v  = getattr(t, 'entry', 0)
        _sl_v     = getattr(t, 'sl', 0)
        _tgt_v    = getattr(t, 'target', 0)
        _pnl_inr  = getattr(t, 'pnl_inr', 0)
        _reason   = getattr(t, 'exit_reason', '')
        _dir_v    = getattr(t, 'direction', '')
        _strat_v  = str(getattr(t, 'strategy', '')).replace('_SUPERTREND', '+ST')
        _iv_v     = getattr(t, 'iv_pct', 0)
        _pnl_c    = "#16a34a" if _pnl_inr >= 0 else "#dc2626"
        fig = _go_bch.Figure()
        fig.add_trace(_go_bch.Candlestick(
            x=_df_b.index, open=_df_b['Open'], high=_df_b['High'],
            low=_df_b['Low'], close=_df_b['Close'],
            increasing_line_color='#16a34a', decreasing_line_color='#dc2626',
            name=_tf_v))
        fig.add_hline(y=_entry_v, line_color='#3b82f6', line_dash='dash',
                      annotation_text=f"Entry {_entry_v:.0f}", annotation_position="right")
        fig.add_hline(y=_sl_v,   line_color='#ef4444', line_dash='dot',
                      annotation_text=f"SL {_sl_v:.0f}", annotation_position="right")
        fig.add_hline(y=_tgt_v,  line_color='#22c55e', line_dash='dot',
                      annotation_text=f"TGT {_tgt_v:.0f}", annotation_position="right")
        _iv_str = f" IV={_iv_v:.0f}%" if _iv_v else ""
        fig.update_layout(
            title=(f"{getattr(t,'instrument','?')} {_tf_v} | {_strat_v} {_dir_v}"
                   f"{_iv_str} | {_reason} | ₹{_pnl_inr:+,.0f}"),
            height=300, margin=dict(l=10, r=90, t=40, b=10),
            xaxis_rangeslider_visible=False,
            plot_bgcolor='#0f172a', paper_bgcolor='#0f172a',
            font_color='#e2e8f0', title_font_color=_pnl_c)
        return fig
    except Exception:
        return None


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 3 — OPT/FUT STRATEGIES
# ═══════════════════════════════════════════════════════════════════════════════
with _tabs[3]:
    st.markdown(
        '<div style="background:linear-gradient(135deg,#0f172a,#1e3a5f);border-radius:12px;'
        'padding:16px 20px;margin-bottom:14px">'
        '<div style="color:#60a5fa;font-size:20px;font-weight:700;font-family:JetBrains Mono">'
        'Options & Futures Strategies</div>'
        '<div style="color:#94a3b8;font-size:11px;margin-top:4px;font-family:JetBrains Mono">'
        '3 Signal Strategies · 10 Option Strategies · 3 Indices · 6 Option + 6 Futures Expiries per Index</div>'
        '</div>',
        unsafe_allow_html=True)

    # Invalidate stale session state if expiry dict has old 3-key format
    if "ofs_all_exp" in st.session_state:
        _sample_inner = next(iter(st.session_state["ofs_all_exp"].values()), {})
        if isinstance(_sample_inner, dict) and 'opt_mid' not in _sample_inner:
            for _sk in ["ofs_all_exp", "ofs_all_signals", "ofs_scan_time"]:
                st.session_state.pop(_sk, None)

    if not OFS_OK:
        st.error(f"options_futures_strategies.py failed to import: {_ofs_err_msg}")
        st.info("Run: pip install scipy yfinance")
    else:
        _ofs_sub = st.tabs(["📡 Live Signals", "📋 All Option Strategies", "📊 Signal Backtest", "📈 Option Strategy Backtest"])
        _OFS_INDICES = ["NIFTY", "BANKNIFTY", "SENSEX"]
        _OFS_COLORS  = {"NIFTY": "#1d4ed8", "BANKNIFTY": "#7c3aed", "SENSEX": "#b45309"}

        # ── helper: render one signal card ──────────────────────────────────
        def _ofs_sig_card(sig):
            _dir   = sig.get('direction', 'FLAT')
            _fresh = sig.get('fresh_signal')
            _dc    = "#16a34a" if _dir == 'LONG' else ("#dc2626" if _dir == 'SHORT' else "#64748b")
            _bg    = "#f0fdf4" if _dir == 'LONG' else ("#fff1f2" if _dir == 'SHORT' else "#f8fafc")
            _badge = ("LONG" if _dir == 'LONG' else ("SHORT" if _dir == 'SHORT' else "FLAT"))
            _arrow = ("&#8593;" if _dir == 'LONG' else ("&#8595;" if _dir == 'SHORT' else "&#8596;"))
            _fresh_row = (f'<div style="color:#f59e0b;font-size:9px;font-weight:700">&#9889; FRESH {_fresh}</div>'
                          if _fresh else '')
            return (
                f'<div style="background:{_bg};border:1.5px solid {_dc}55;border-radius:9px;'
                f'padding:10px;margin-bottom:6px;font-family:JetBrains Mono">'
                f'<div style="display:flex;justify-content:space-between;align-items:center">'
                f'<span style="color:#374151;font-size:10px;font-weight:700">{sig.get("strategy","")[:20]}</span>'
                f'<span style="color:{_dc};font-size:13px;font-weight:800">{_arrow} {_badge}</span>'
                f'</div>'
                f'{_fresh_row}'
                f'<div style="font-size:9px;color:#374151;margin-top:5px;line-height:1.6">'
                f'Entry <b>&#8377;{sig.get("entry",0):,.0f}</b> &nbsp;'
                f'SL <span style="color:#dc2626">&#8377;{sig.get("sl",0):,.0f}</span> &nbsp;'
                f'Tgt <span style="color:#16a34a">&#8377;{sig.get("target",0):,.0f}</span> &nbsp;'
                f'R:R <b>{sig.get("rr",0):.1f}</b>'
                f'</div>'
                f'<div style="color:#94a3b8;font-size:8px;margin-top:3px">{sig.get("reason","")[:70]}</div>'
                f'</div>'
            )

        # ── Live Signals ─────────────────────────────────────────────────────
        with _ofs_sub[0]:
            _ofs_tf_choices = st.multiselect(
                "Timeframes", ["5m", "15m", "1h"], default=["5m", "15m", "1h"], key="ofs_tfs")
            _ofs_refresh = st.button("&#128260; Refresh All Signals", key="ofs_refresh_btn")

            if _ofs_refresh or "ofs_all_signals" not in st.session_state:
                if not _ofs_tf_choices:
                    st.warning("Select at least one timeframe")
                else:
                    with st.spinner("Scanning NIFTY, BANKNIFTY, SENSEX..."):
                        _all_sigs = {}
                        _all_exp  = {}
                        for _idx in _OFS_INDICES:
                            try:
                                _all_sigs[_idx] = _ofs_scan_all(_idx, _ofs_tf_choices)
                            except Exception as _se:
                                _all_sigs[_idx] = {}
                            try:
                                _all_exp[_idx] = _ofs_all_expiries(_idx)
                            except Exception:
                                _all_exp[_idx] = {}
                        st.session_state["ofs_all_signals"] = _all_sigs
                        st.session_state["ofs_all_exp"]     = _all_exp
                        st.session_state["ofs_scan_time"]   = datetime.now(_OFS_IST).strftime('%H:%M:%S IST')

            if "ofs_all_signals" in st.session_state:
                _scan_ts   = st.session_state.get("ofs_scan_time", "—")
                _all_sigs  = st.session_state["ofs_all_signals"]
                _all_exp   = st.session_state.get("ofs_all_exp", {})
                st.caption(f"Last scan: {_scan_ts}")

                # 3-column layout: one per index
                _col_n, _col_b, _col_s = st.columns(3)
                for _icol, _idx in zip([_col_n, _col_b, _col_s], _OFS_INDICES):
                    _ic = _OFS_COLORS[_idx]
                    _exp  = _all_exp.get(_idx, {})
                    _on   = _exp.get('opt_near', ('—', '—'))
                    _om   = _exp.get('opt_mid',  ('—', '—'))
                    _of   = _exp.get('opt_far',  ('—', '—'))
                    _fn   = _exp.get('fut_near', ('—', '—'))
                    _ff   = _exp.get('fut_far',  ('—', '—'))

                    with _icol:
                        # Index header
                        st.markdown(
                            f'<div style="background:{_ic};border-radius:10px;padding:10px 14px;margin-bottom:8px">'
                            f'<div style="color:white;font-size:15px;font-weight:800;font-family:JetBrains Mono">{_idx}</div>'
                            f'</div>',
                            unsafe_allow_html=True)

                        # Expiry badges: 3 opt + 2 fut
                        st.markdown(
                            f'<div style="font-family:JetBrains Mono;font-size:10px;margin-bottom:8px">'
                            f'<div style="background:#dbeafe;border-radius:6px;padding:4px 8px;margin-bottom:3px">'
                            f'<b>Opt 1:</b> {_on[0]} &nbsp;<span style="color:#64748b">(DTE {_on[1]})</span></div>'
                            f'<div style="background:#ede9fe;border-radius:6px;padding:4px 8px;margin-bottom:3px">'
                            f'<b>Opt 2:</b> {_om[0]} &nbsp;<span style="color:#64748b">(DTE {_om[1]})</span></div>'
                            f'<div style="background:#e0e7ff;border-radius:6px;padding:4px 8px;margin-bottom:3px">'
                            f'<b>Opt 3:</b> {_of[0]} &nbsp;<span style="color:#64748b">(DTE {_of[1]})</span></div>'
                            f'<div style="background:#fef3c7;border-radius:6px;padding:4px 8px;margin-bottom:3px">'
                            f'<b>Fut 1:</b> {_fn[0]} &nbsp;<span style="color:#64748b">(DTE {_fn[1]})</span></div>'
                            f'<div style="background:#fed7aa;border-radius:6px;padding:4px 8px">'
                            f'<b>Fut 2:</b> {_ff[0]} &nbsp;<span style="color:#64748b">(DTE {_ff[1]})</span></div>'
                            f'</div>',
                            unsafe_allow_html=True)

                        # Signals per timeframe
                        _idx_sigs = _all_sigs.get(_idx, {})
                        for _tf in (_ofs_tf_choices or ["5m", "15m", "1h"]):
                            if _tf not in _idx_sigs:
                                continue
                            tf_signals = [s for s in _idx_sigs[_tf]
                                          if isinstance(s, dict) and 'strategy' in s and 'direction' in s]
                            if not tf_signals:
                                continue
                            st.markdown(
                                f'<div style="color:{_ic};font-size:10px;font-weight:700;'
                                f'font-family:JetBrains Mono;margin:8px 0 4px">{_tf.upper()}</div>',
                                unsafe_allow_html=True)
                            for _sig in tf_signals:
                                st.markdown(_ofs_sig_card(_sig), unsafe_allow_html=True)

        # ── All Option Strategies ────────────────────────────────────────────
        with _ofs_sub[1]:
            _as_row = st.columns([2, 1])
            with _as_row[0]:
                _as_idx = st.radio("Index", _OFS_INDICES, horizontal=True, key="as_idx_radio")
            with _as_row[1]:
                _as_load = st.button("&#128260; Load Strategies", key="as_load_btn")

            if _as_load or f"as_data_{_as_idx}" not in st.session_state:
                with st.spinner(f"Computing 30 strategies for {_as_idx}..."):
                    try:
                        st.session_state[f"as_data_{_as_idx}"] = _ofs_all_strategies(_as_idx)
                    except Exception as _ase:
                        st.error(f"Error: {_ase}")
                        st.session_state[f"as_data_{_as_idx}"] = []

            _as_data = st.session_state.get(f"as_data_{_as_idx}", [])
            if _as_data:
                # Group into dict: strategy_name -> {expiry_str: setup}
                _strategy_order = [
                    'Long Call', 'Long Put', 'Long Straddle', 'Long Strangle',
                    'Covered Call', 'Protective Put', 'Collar',
                    'Synthetic Long', 'Call Backspread', 'Put Backspread'
                ]
                _strat_map = {s: {} for s in _strategy_order}
                _expiries_seen = []
                for _setup in _as_data:
                    _sn  = _setup['strategy_name']
                    _exp = _setup['expiry']
                    if _sn in _strat_map:
                        _strat_map[_sn][_exp] = _setup
                    if _exp not in _expiries_seen:
                        _expiries_seen.append(_exp)

                # Header row: strategy | Expiry1 | Expiry2 | Expiry3
                _gh = ['Strategy', 'Market View'] + [f"Expiry {i+1}" for i in range(len(_expiries_seen))]
                _hcols = st.columns([2, 1.5] + [2] * len(_expiries_seen))
                for _hc, _ht in zip(_hcols, _gh):
                    _hc.markdown(f'<div style="font-family:JetBrains Mono;font-size:10px;'
                                 f'font-weight:700;color:#475569;padding:4px 0">{_ht}</div>',
                                 unsafe_allow_html=True)

                # Expiry labels
                _hcols2 = st.columns([2, 1.5] + [2] * len(_expiries_seen))
                _hcols2[0].markdown("")
                _hcols2[1].markdown("")
                for _ei, _ev in enumerate(_expiries_seen):
                    _dte_v = next((s['dte'] for s in _as_data if s['expiry'] == _ev), 0)
                    _hcols2[2 + _ei].markdown(
                        f'<div style="font-family:JetBrains Mono;font-size:9px;'
                        f'color:#64748b;padding:2px 0">{_ev}<br>DTE {_dte_v}</div>',
                        unsafe_allow_html=True)

                st.markdown('<hr style="margin:4px 0;border-color:#e2e8f0">', unsafe_allow_html=True)

                _view_colors = {
                    'Bullish': '#16a34a', 'Bearish': '#dc2626',
                    'High Volatility': '#7c3aed', 'Extreme Move': '#6d28d9',
                    'Neutral/Mild Bull': '#0284c7', 'Bullish + Hedged': '#059669',
                    'Neutral/Capped': '#0891b2', 'Futures-Like Risk': '#1d4ed8',
                    'Explosive Up': '#d97706', 'Explosive Down': '#b91c1c',
                }

                for _sname in _strategy_order:
                    _row_data = _strat_map.get(_sname, {})
                    if not _row_data:
                        continue
                    _first   = next(iter(_row_data.values()))
                    _view    = _first.get('market_view', '')
                    _vc      = _view_colors.get(_view, '#374151')
                    _rcols   = st.columns([2, 1.5] + [2] * len(_expiries_seen))
                    _rcols[0].markdown(
                        f'<div style="font-family:JetBrains Mono;font-size:10px;'
                        f'font-weight:700;color:#1e293b;padding:6px 0">{_sname}</div>',
                        unsafe_allow_html=True)
                    _rcols[1].markdown(
                        f'<div style="font-family:JetBrains Mono;font-size:9px;'
                        f'color:{_vc};font-weight:600;padding:6px 0">{_view}</div>',
                        unsafe_allow_html=True)

                    for _ei, _ev in enumerate(_expiries_seen):
                        _cell = _row_data.get(_ev)
                        if _cell is None:
                            _rcols[2 + _ei].markdown("—")
                            continue
                        _sign = _cell.get('net_sign', 'DEBIT')
                        _bg   = "#f0fdf4" if _sign == 'DEBIT' else "#fff1f2"
                        _bc   = "#16a34a" if _sign == 'DEBIT' else "#dc2626"
                        _net  = _cell.get('net_premium', 0)
                        _blo  = _cell.get('breakeven_lo', 0)
                        _bhi  = _cell.get('breakeven_hi', 0)
                        _legs = _cell.get('legs', [])
                        _leg_str = ' | '.join(
                            f'{"BUY" if l["type"]=="BUY" else ("SELL" if l["type"]=="SELL" else l["type"])}'
                            f' {l["option_type"]}'
                            + (f'×{l.get("qty",1)}' if l.get('qty',1)>1 else '')
                            + (f' {l["strike"]}' if l.get('strike') else '')
                            for l in _legs if l['option_type'] != 'FUT'
                        )
                        _rcols[2 + _ei].markdown(
                            f'<div style="background:{_bg};border:1px solid {_bc}44;border-radius:6px;'
                            f'padding:5px 7px;font-family:JetBrains Mono;font-size:8px">'
                            f'<div style="color:{_bc};font-weight:700">{_sign} &#8377;{_net:.0f}</div>'
                            f'<div style="color:#374151;margin-top:2px">{_leg_str[:40]}</div>'
                            f'<div style="color:#64748b;margin-top:2px">'
                            f'BEP: {_blo:.0f}{"/" + str(round(_bhi)) if _blo != _bhi else ""}</div>'
                            f'</div>',
                            unsafe_allow_html=True)

                    st.markdown('<hr style="margin:2px 0;border-color:#f1f5f9">', unsafe_allow_html=True)
            elif not _as_load:
                st.info(f"Click 'Load Strategies' to compute all 10 strategies × 3 expiries for {_as_idx}")

        # ── Backtest ─────────────────────────────────────────────────────────
        with _ofs_sub[2]:

            # ── Excel generator helper ───────────────────────────────────────
            def _build_excel(flat_trades: list) -> bytes:
                import io as _io
                _buf = _io.BytesIO()
                if not flat_trades:
                    pd.DataFrame().to_excel(_buf, index=False)
                    _buf.seek(0)
                    return _buf.read()

                _df_all = pd.DataFrame(flat_trades)
                _df_all['cumul_pnl'] = _df_all.groupby(
                    ['idx','strategy','tf'])['pnl_inr'].cumsum()

                try:
                    from openpyxl import Workbook
                    from openpyxl.chart import LineChart, Reference
                    from openpyxl.chart.series import SeriesLabel
                    from openpyxl.styles import Font, PatternFill, Alignment
                    from openpyxl.utils import get_column_letter

                    wb = Workbook()
                    ws_sum = wb.active
                    ws_sum.title = "P&L Summary"

                    # Summary sheet
                    _hdr_fill = PatternFill("solid", fgColor="0F172A")
                    _hdr_font = Font(color="FFFFFF", bold=True)
                    _sum_cols = ['Index','Strategy','TF','Mode','R:R','Trades','Wins','Losses',
                                 'Win%','Profit Factor','Max DD (Rs)','Net P&L (Rs)']
                    for ci, col in enumerate(_sum_cols, 1):
                        cell = ws_sum.cell(1, ci, col)
                        cell.fill = _hdr_fill
                        cell.font = _hdr_font
                        cell.alignment = Alignment(horizontal='center')

                    _xl_grp_cols = ['idx','strategy','tf','mode','rr_ratio'] if 'rr_ratio' in _df_all.columns else ['idx','strategy','tf','mode']
                    _grp = _df_all.groupby(_xl_grp_cols)
                    _row = 2
                    for _gk, _gdf in _grp:
                        if len(_xl_grp_cols) == 5:
                            idx, strat, tf, mode, rr = _gk
                        else:
                            idx, strat, tf, mode = _gk; rr = 2.0
                        pnl_list = _gdf['pnl_inr'].tolist()
                        wins  = sum(1 for p in pnl_list if p > 0)
                        losses= len(pnl_list) - wins
                        gross_w = sum(p for p in pnl_list if p > 0)
                        gross_l = abs(sum(p for p in pnl_list if p < 0))
                        pf = round(gross_w / gross_l, 2) if gross_l > 0 else 99.0
                        cumul = 0.0; peak = 0.0; max_dd = 0.0
                        for p in pnl_list:
                            cumul += p; peak = max(peak, cumul)
                            max_dd = max(max_dd, peak - cumul)
                        wr = round(wins / len(pnl_list) * 100, 1) if pnl_list else 0
                        _vals = [idx, strat, tf, mode, f"1:{rr:.0f}", len(pnl_list), wins, losses,
                                 wr, pf, round(max_dd, 2), round(sum(pnl_list), 2)]
                        for ci, v in enumerate(_vals, 1):
                            ws_sum.cell(_row, ci, v)
                        _row += 1

                    for col in ws_sum.columns:
                        ws_sum.column_dimensions[get_column_letter(col[0].column)].width = 14

                    # Per-strategy sheets
                    for strat in _OFS_STRATEGIES:
                        _sdf = _df_all[_df_all['strategy'] == strat].copy()
                        if _sdf.empty:
                            continue
                        _sname = strat.replace('_SUPERTREND', '+ST').replace('_', ' ')[:31]
                        ws = wb.create_sheet(title=_sname)

                        _trade_cols = ['idx','tf','mode','expiry_used','date','entry_time','exit_time',
                                       'direction','exit_reason','rr_ratio',
                                       'entry','exit_price','sl','target',
                                       'theta_entry','theta_exit',
                                       'hit_sl','hit_target','hold_bars','pnl_pts','pnl_inr','cumul_pnl']
                        _trade_cols = [c for c in _trade_cols if c in _sdf.columns]
                        # Write header
                        for ci, col in enumerate(_trade_cols, 1):
                            cell = ws.cell(1, ci, col)
                            cell.fill = _hdr_fill; cell.font = _hdr_font
                        # Write rows with color coding
                        _green_fill = PatternFill("solid", fgColor="F0FDF4")
                        _red_fill   = PatternFill("solid", fgColor="FFF1F2")
                        for ri, (_, row) in enumerate(_sdf[_trade_cols].iterrows(), 2):
                            _fill = _green_fill if row.get('pnl_inr', 0) > 0 else _red_fill
                            for ci, col in enumerate(_trade_cols, 1):
                                cell = ws.cell(ri, ci, row[col] if col in row else "")
                                cell.fill = _fill

                        for col in ws.columns:
                            ws.column_dimensions[get_column_letter(col[0].column)].width = 12

                        # Equity curve chart
                        if 'cumul_pnl' in _sdf.columns and len(_sdf) > 1:
                            _cumul_col = _trade_cols.index('cumul_pnl') + 1
                            chart = LineChart()
                            chart.title = f"{strat} — Cumulative P&L"
                            chart.style = 10
                            chart.y_axis.title = "P&L (Rs)"
                            chart.x_axis.title = "Trade #"
                            chart.width = 20; chart.height = 12
                            _data_ref = Reference(ws, min_col=_cumul_col, min_row=1,
                                                  max_row=len(_sdf) + 1)
                            chart.add_data(_data_ref, titles_from_data=True)
                            chart.series[0].graphicalProperties.line.solidFill = "1D4ED8"
                            ws.add_chart(chart, f"{get_column_letter(len(_trade_cols)+2)}2")

                    wb.save(_buf)
                    _buf.seek(0)
                    return _buf.read()

                except ImportError:
                    # Fallback: plain Excel via pandas
                    with pd.ExcelWriter(_buf, engine='openpyxl') as _wr:
                        _grp_cols = ['idx','strategy','tf','mode']
                        _sum_rows = []
                        for (idx, strat, tf, mode), _gdf in _df_all.groupby(_grp_cols):
                            _pl = _gdf['pnl_inr'].tolist()
                            _sum_rows.append({'Index':idx,'Strategy':strat,'TF':tf,'Mode':mode,
                                'Trades':len(_pl),'Wins':sum(1 for p in _pl if p>0),
                                'Win%':round(sum(1 for p in _pl if p>0)/len(_pl)*100,1) if _pl else 0,
                                'Net P&L':round(sum(_pl),2)})
                        pd.DataFrame(_sum_rows).to_excel(_wr, sheet_name='P&L Summary', index=False)
                        for strat in _OFS_STRATEGIES:
                            _sdf = _df_all[_df_all['strategy']==strat]
                            if not _sdf.empty:
                                _sdf.to_excel(_wr, sheet_name=strat[:31], index=False)
                    _buf.seek(0)
                    return _buf.read()

            # ── Session state: expiry offsets ────────────────────────────────
            for _bidx2 in _OFS_INDICES:
                for _knd in ['opt', 'fut']:
                    _osk = f"bt_{_knd}_offset_{_bidx2}"
                    if _osk not in st.session_state:
                        st.session_state[_osk] = 0

            # ── Options expiry selector ──────────────────────────────────────
            st.markdown(
                '<div style="background:#dbeafe;border-radius:8px;padding:8px 12px;'
                'margin-bottom:8px;font-family:JetBrains Mono;font-size:11px;font-weight:700;color:#1d4ed8">'
                'OPTIONS EXPIRY — select past weekly contracts to backtest (theta-decay ON)</div>',
                unsafe_allow_html=True)

            _all_opt_sel = {}
            for _bidx in _OFS_INDICES:
                _off = st.session_state[f"bt_opt_offset_{_bidx}"]
                _past_opt = _ofs_past_opt(_bidx, n=6, offset=_off)
                _opt_labels = [e.strftime('%d %b %y') for e in _past_opt]

                _oc1, _oc2 = st.columns([5, 1])
                with _oc1:
                    _sel_labels = st.multiselect(
                        f"{_bidx} Options (weekly)",
                        options=_opt_labels, default=_opt_labels,
                        key=f"ms_opt_{_bidx}_{_off}")
                    _all_opt_sel[_bidx] = [_past_opt[_opt_labels.index(l)] for l in _sel_labels]
                with _oc2:
                    st.markdown("<div style='margin-top:26px'></div>", unsafe_allow_html=True)
                    _pb_c1, _pb_c2 = st.columns(2)
                    def _prev_opt(_i=_bidx):
                        st.session_state[f"bt_opt_offset_{_i}"] += 6
                    def _next_opt(_i=_bidx):
                        st.session_state[f"bt_opt_offset_{_i}"] = max(0, st.session_state[f"bt_opt_offset_{_i}"] - 6)
                    _pb_c1.button("◀ Prev", key=f"prev_opt_{_bidx}", on_click=_prev_opt)
                    _pb_c2.button("Next ▶", key=f"next_opt_{_bidx}", on_click=_next_opt,
                                  disabled=_off == 0)

            # ── Futures expiry selector ──────────────────────────────────────
            st.markdown(
                '<div style="background:#fef3c7;border-radius:8px;padding:8px 12px;'
                'margin:8px 0;font-family:JetBrains Mono;font-size:11px;font-weight:700;color:#b45309">'
                'FUTURES EXPIRY — select past monthly contracts to backtest</div>',
                unsafe_allow_html=True)

            _all_fut_sel = {}
            for _bidx in _OFS_INDICES:
                _off = st.session_state[f"bt_fut_offset_{_bidx}"]
                _past_fut = _ofs_past_fut(_bidx, n=6, offset=_off)
                _fut_labels = [e.strftime('%d %b %y') for e in _past_fut]

                _fc1, _fc2 = st.columns([5, 1])
                with _fc1:
                    _fsel_labels = st.multiselect(
                        f"{_bidx} Futures (monthly)",
                        options=_fut_labels, default=_fut_labels,
                        key=f"ms_fut_{_bidx}_{_off}")
                    _all_fut_sel[_bidx] = [_past_fut[_fut_labels.index(l)] for l in _fsel_labels]
                with _fc2:
                    st.markdown("<div style='margin-top:26px'></div>", unsafe_allow_html=True)
                    _fb_c1, _fb_c2 = st.columns(2)
                    def _prev_fut(_i=_bidx):
                        st.session_state[f"bt_fut_offset_{_i}"] += 6
                    def _next_fut(_i=_bidx):
                        st.session_state[f"bt_fut_offset_{_i}"] = max(0, st.session_state[f"bt_fut_offset_{_i}"] - 6)
                    _fb_c1.button("◀ Prev", key=f"prev_fut_{_bidx}", on_click=_prev_fut)
                    _fb_c2.button("Next ▶", key=f"next_fut_{_bidx}", on_click=_next_fut,
                                  disabled=_off == 0)

            st.caption("Note: 5m/15m data limited to last ~58 days by yfinance. Older expiries automatically use 1h only.")

            # ── R:R control (IV runs automatically across all values) ────────
            _bt_rr_opts  = {"1:2 (conservative)": 2.0, "1:3 (aggressive)": 3.0}
            _bt_rr_sel   = st.multiselect(
                "R:R Ratio",
                options=list(_bt_rr_opts.keys()),
                default=list(_bt_rr_opts.keys()),
                key="bt_rr_sel",
                help="Select one or both risk-reward ratios to test")
            _bt_rr_mults = [_bt_rr_opts[k] for k in _bt_rr_sel] or [2.0]
            st.markdown(
                '<div style="background:#fefce8;border:1px solid #fde047;border-radius:8px;'
                'padding:8px 14px;margin-bottom:6px;font-size:12px">'
                '<b>IV auto-tested:</b> 15% · 20% · 25% · 30% · 35% · 40% · 45% · 50% · 55% · 60%'
                ' &nbsp;—&nbsp; all 10 IVs are run automatically for options; futures ignore IV.</div>',
                unsafe_allow_html=True)

            _ALL_BT_IVS = [15, 20, 25, 30, 35, 40, 45, 50, 55, 60]

            st.info(
                f"**Strategies:** {', '.join(_OFS_STRATEGIES)} | "
                f"**R:R:** {', '.join(_bt_rr_sel) if _bt_rr_sel else '1:2'} | "
                "**IV:** all 10 levels (15–60%) for options | "
                "SL = indicator stop level | Target = Entry ± R:R × SL dist | "
                "CE_REGIME exits early when market turns sideways (exit_reason=SIDEWAYS)")

            # ── Run button ───────────────────────────────────────────────────
            _bt_rr_label = " + ".join(_bt_rr_sel) if _bt_rr_sel else "1:2"
            if st.button(
                f"▶ Run All Strategies — {_bt_rr_label} R:R × All TFs × All Indices × All IVs",
                key="ofs_run_bt", type="primary"):
                _bt_flat   = []
                _prog_lbl  = st.empty()
                # total = options (10 IVs × expiries × 3 TFs) + futures (1 × expiries × 3 TFs)
                _n_opt_exp = sum(len(_all_opt_sel[_i]) for _i in _OFS_INDICES)
                _n_fut_exp = sum(len(_all_fut_sel[_i]) for _i in _OFS_INDICES)
                _total_runs = (len(_ALL_BT_IVS) * _n_opt_exp * 3 + _n_fut_exp * 3) * len(_bt_rr_mults)
                _prog = st.progress(0.0)
                _prog_n = 0

                for _rr_m in _bt_rr_mults:
                    # ── Options: loop all IVs ─────────────────────────────
                    for _iv in _ALL_BT_IVS:
                        for _bidx in _OFS_INDICES:
                            for _exp in _all_opt_sel[_bidx]:
                                _pstart = _exp - timedelta(days=8)
                                for _tf in ['5m', '15m', '1h']:
                                    _cutoff = date.today() - timedelta(days=57)
                                    if _tf in ('5m', '15m') and _exp < _cutoff:
                                        _prog_n += 1
                                        _prog.progress(min(_prog_n / max(_total_runs, 1), 1.0))
                                        continue
                                    _prog_lbl.caption(
                                        f"Running IV={_iv}% · {_bidx} · options · {_tf} · {_exp} …")
                                    try:
                                        _btr = _OFSBacktester(_bidx)
                                        for _strat in _OFS_STRATEGIES:
                                            _tr = _btr.run(_strat, _tf, _pstart, _exp,
                                                           mode='options', expiry_date=_exp,
                                                           rr_mult=_rr_m,
                                                           vix_override=float(_iv) / 100.0)
                                            for _t in _tr:
                                                _row = vars(_t).copy()
                                                _row.update({'idx': _bidx, 'mode': 'options',
                                                             'expiry_used': str(_exp),
                                                             'instrument': _bidx})
                                                _t.instrument = _bidx
                                                _bt_flat.append(_row)
                                    except Exception:
                                        pass
                                    _prog_n += 1
                                    _prog.progress(min(_prog_n / max(_total_runs, 1), 1.0))

                    # ── Futures: single IV pass (IV unused in P&L) ────────
                    for _bidx in _OFS_INDICES:
                        for _exp in _all_fut_sel[_bidx]:
                            _pstart = _exp - timedelta(days=31)
                            for _tf in ['5m', '15m', '1h']:
                                _cutoff = date.today() - timedelta(days=57)
                                if _tf in ('5m', '15m') and _exp < _cutoff:
                                    _prog_n += 1
                                    _prog.progress(min(_prog_n / max(_total_runs, 1), 1.0))
                                    continue
                                _prog_lbl.caption(
                                    f"Running {_bidx} · futures · {_tf} · {_exp} …")
                                try:
                                    _btr = _OFSBacktester(_bidx)
                                    for _strat in _OFS_STRATEGIES:
                                        _tr = _btr.run(_strat, _tf, _pstart, _exp,
                                                       mode='futures', expiry_date=_exp,
                                                       rr_mult=_rr_m,
                                                       vix_override=0.20)
                                        for _t in _tr:
                                            _row = vars(_t).copy()
                                            _row.update({'idx': _bidx, 'mode': 'futures',
                                                         'expiry_used': str(_exp),
                                                         'instrument': _bidx})
                                            _t.instrument = _bidx
                                            _bt_flat.append(_row)
                                except Exception:
                                    pass
                                _prog_n += 1
                                _prog.progress(min(_prog_n / max(_total_runs, 1), 1.0))

                _prog.empty()
                _prog_lbl.empty()
                st.session_state["ofs_bt_flat"] = _bt_flat
                n_opt = sum(len(v) for v in _all_opt_sel.values())
                n_fut = sum(len(v) for v in _all_fut_sel.values())
                st.session_state["ofs_bt_label"] = (
                    f"Options: {n_opt} expiries × 10 IVs | Futures: {n_fut} expiries | "
                    f"R:R: {_bt_rr_label} | All TFs | {len(_bt_flat)} total trades")

            # ── Results display ──────────────────────────────────────────────
            if "ofs_bt_flat" in st.session_state and st.session_state["ofs_bt_flat"]:
                _bt_flat = st.session_state["ofs_bt_flat"]
                st.caption(st.session_state.get("ofs_bt_label", ""))

                _df_bt = pd.DataFrame(_bt_flat)

                # Excel download
                try:
                    _xl_bytes = _build_excel(_bt_flat)
                    st.download_button(
                        label="⬇ Download Excel Report",
                        data=_xl_bytes,
                        file_name=f"backtest_{date.today()}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="bt_excel_dl")
                except Exception as _xle:
                    st.warning(f"Excel generation failed: {_xle}")

                # ── Inner sub-tabs ───────────────────────────────────────
                import plotly.graph_objects as _go_bt
                _bt_inner = st.tabs(["📋 Trade Report", "📊 Analytics"])

                # ════════════════════════════════════════════════════════
                # TAB A — Trade Report
                # ════════════════════════════════════════════════════════
                with _bt_inner[0]:

                    # P&L Summary (grouped by index × strategy × direction × mode)
                    st.markdown("#### P&L Summary by Strategy & Direction")
                    _sum_rows = []
                    for (_idx, _strat, _dir, _mode), _gdf in _df_bt.groupby(
                            ['idx', 'strategy', 'direction', 'mode']):
                        _pl   = _gdf['pnl_inr'].tolist()
                        _wins = sum(1 for p in _pl if p > 0)
                        _gl   = abs(sum(p for p in _pl if p < 0))
                        _gw   = sum(p for p in _pl if p > 0)
                        _ivs  = _gdf['iv_pct'].dropna() if 'iv_pct' in _gdf.columns else pd.Series([0])
                        _iv_rng = (f"{_ivs.min():.0f}%–{_ivs.max():.0f}%"
                                   if _mode == 'options' else "N/A")
                        _sum_rows.append({
                            'Index': _idx, 'Strategy': _strat, 'Direction': _dir,
                            'Mode': _mode, 'IV Range': _iv_rng,
                            'Trades': len(_pl), 'Wins': _wins,
                            'Losses': len(_pl) - _wins,
                            'Win%': f"{_wins/len(_pl)*100:.1f}%" if _pl else "0%",
                            'Profit Factor': f"{_gw/_gl:.2f}" if _gl > 0 else "∞",
                            'Net P&L (₹)': round(sum(_pl), 0)})
                    if _sum_rows:
                        _sum_df = pd.DataFrame(_sum_rows)
                        def _pnl_color_bt(row):
                            c = "#f0fdf4" if row.get('Net P&L (₹)', 0) >= 0 else "#fff1f2"
                            return [f"background-color:{c}"] * len(row)
                        st.dataframe(_sum_df.style.apply(_pnl_color_bt, axis=1),
                                     hide_index=True, use_container_width=True)

                    # Per-strategy expanders with trade table + chart gallery
                    for _strat in _OFS_STRATEGIES:
                        _sdf = _df_bt[_df_bt['strategy'] == _strat]
                        if _sdf.empty:
                            continue
                        _sname     = _strat.replace('_SUPERTREND', '+Supertrend').replace('_', ' ')
                        _total_pnl = _sdf['pnl_inr'].sum()
                        with st.expander(
                                f"{_sname} — {len(_sdf)} trades | Net ₹{_total_pnl:+,.0f}",
                                expanded=True):

                            # Equity curve per TF
                            _fig_eq = _go_bt.Figure()
                            for _tf_eq in ['5m', '15m', '1h']:
                                _tdf_eq = _sdf[_sdf['tf'] == _tf_eq]
                                if not _tdf_eq.empty:
                                    _eq_s = _tdf_eq['pnl_inr'].cumsum().reset_index(drop=True)
                                    _fig_eq.add_trace(_go_bt.Scatter(
                                        y=_eq_s, mode='lines', name=_tf_eq,
                                        line={'width': 2}))
                            _fig_eq.update_layout(
                                title=f"{_sname} — Cumulative P&L by Timeframe",
                                yaxis_title="P&L (₹)", xaxis_title="Trade #",
                                height=260, margin=dict(l=10, r=10, t=35, b=10),
                                legend=dict(orientation='h', y=-0.15),
                                paper_bgcolor='#f8fafc', plot_bgcolor='#f8fafc')
                            st.plotly_chart(_fig_eq, use_container_width=True,
                                            key=f"bt_eq_{_strat}")

                            # Trade table
                            _show_cols = ['idx', 'instrument', 'iv_pct', 'tf', 'mode',
                                          'expiry_used', 'date', 'entry_time', 'exit_time',
                                          'direction', 'exit_reason', 'rr_ratio',
                                          'entry', 'exit_price', 'sl', 'target',
                                          'theta_entry', 'theta_exit',
                                          'pnl_pts', 'pnl_inr', 'hit_sl', 'hit_target']
                            _show_cols = [c for c in _show_cols if c in _sdf.columns]
                            _disp_s = _sdf[_show_cols].copy()
                            _disp_s['pnl_inr'] = _disp_s['pnl_inr'].round(0)
                            _rename_map = {'idx': 'Index', 'iv_pct': 'IV%',
                                           'instrument': 'Instrument',
                                           'entry_time': 'Time In',
                                           'exit_time': 'Time Out',
                                           'exit_reason': 'Reason',
                                           'rr_ratio': 'R:R',
                                           'exit_price': 'Exit',
                                           'expiry_used': 'Expiry',
                                           'pnl_pts': 'P&L pts',
                                           'pnl_inr': 'P&L ₹',
                                           'hit_sl': 'Hit SL',
                                           'hit_target': 'Hit Tgt',
                                           'theta_entry': 'Θ Entry',
                                           'theta_exit': 'Θ Exit'}
                            _disp_s = _disp_s.rename(columns=_rename_map)

                            def _bt_row_color(row):
                                c = "#f0fdf4" if row.get('P&L ₹', 0) >= 0 else "#fff1f2"
                                return [f"background-color:{c}"] * len(row)

                            st.dataframe(_disp_s.style.apply(_bt_row_color, axis=1),
                                         hide_index=True, use_container_width=True)

                            # Chart gallery (collapsed by default — opens on demand)
                            _strat_trades_for_chart = _sdf.to_dict('records')
                            with st.expander(
                                    f"📈 Trade Charts ({len(_strat_trades_for_chart)} trades, "
                                    "max 24 shown)", expanded=False):
                                _cg1, _cg2 = st.columns(2)
                                _ci = 0
                                for _crow in _strat_trades_for_chart[:24]:
                                    # build a mini namespace so _plot_bt_trade_chart can access fields
                                    class _TR:
                                        pass
                                    _tobj = _TR()
                                    for _k, _v in _crow.items():
                                        setattr(_tobj, _k, _v)
                                    _tobj.instrument = _crow.get('instrument') or _crow.get('idx', 'NIFTY')
                                    _fig_ch = _plot_bt_trade_chart(_tobj)
                                    if _fig_ch:
                                        (_cg1 if _ci % 2 == 0 else _cg2).plotly_chart(
                                            _fig_ch, use_container_width=True,
                                            key=f"btch_{_strat}_{_ci}")
                                        _ci += 1

                # ════════════════════════════════════════════════════════
                # TAB B — Analytics
                # ════════════════════════════════════════════════════════
                with _bt_inner[1]:

                    _STARTING_CAPITAL = 1_00_00_000  # ₹1 Cr

                    # Section A — Bar chart: net P&L by strategy × direction
                    st.markdown("#### Net P&L by Strategy & Direction")
                    _an_fig = _go_bt.Figure()
                    _an_strats = _OFS_STRATEGIES
                    _an_strat_labels = [s.replace('_SUPERTREND', '+ST').replace('_', ' ')
                                        for s in _an_strats]
                    for _an_dir, _an_col in [("LONG", "#16a34a"), ("SHORT", "#dc2626")]:
                        _an_vals = [
                            round(_df_bt[(_df_bt['strategy'] == _s) &
                                         (_df_bt['direction'] == _an_dir)]['pnl_inr'].sum(), 0)
                            for _s in _an_strats]
                        _an_fig.add_trace(_go_bt.Bar(
                            name=_an_dir, x=_an_strat_labels, y=_an_vals,
                            marker_color=_an_col,
                            text=[f"₹{v:+,.0f}" for v in _an_vals],
                            textposition='outside'))
                    _an_fig.update_layout(
                        barmode='group',
                        title="Net P&L by Strategy & Direction",
                        yaxis_title="Net P&L (₹)",
                        height=380,
                        margin=dict(l=10, r=10, t=50, b=10),
                        legend=dict(orientation='h', y=1.08),
                        plot_bgcolor='#0f172a', paper_bgcolor='#0f172a',
                        font_color='#e2e8f0', title_font_color='#e2e8f0',
                        yaxis=dict(gridcolor='#334155'))
                    st.plotly_chart(_an_fig, use_container_width=True, key="bt_an_bar")

                    # Section B — Win/loss stats table
                    st.markdown("#### Win / Loss Statistics by Strategy & Direction")
                    _wl_rows = []
                    for _s in _an_strats:
                        for _d in ("LONG", "SHORT"):
                            _sub = _df_bt[(_df_bt['strategy'] == _s) &
                                          (_df_bt['direction'] == _d)]['pnl_inr'].tolist()
                            if not _sub:
                                continue
                            _ww = [p for p in _sub if p > 0]
                            _ll = [p for p in _sub if p < 0]
                            _avg_w = round(sum(_ww) / len(_ww), 0) if _ww else 0
                            _avg_l = round(sum(_ll) / len(_ll), 0) if _ll else 0
                            _pf    = (round((len(_ww) * _avg_w) /
                                     max(abs(len(_ll) * _avg_l), 1), 2)) if _ll else 99.0
                            _wl_rows.append({
                                'Strategy': _s.replace('_SUPERTREND', '+ST').replace('_', ' '),
                                'Direction': _d,
                                'Trades': len(_sub),
                                'Win%': f"{len(_ww)/len(_sub)*100:.1f}%",
                                'Avg Win ₹': f"₹{_avg_w:+,.0f}",
                                'Avg Loss ₹': f"₹{_avg_l:+,.0f}",
                                'Profit Factor': _pf})
                    if _wl_rows:
                        st.dataframe(pd.DataFrame(_wl_rows), hide_index=True,
                                     use_container_width=True)

                    # Section C — Cumulative P&L table (time-sorted, ₹1 Cr start)
                    st.markdown("#### Cumulative P&L (Starting Capital: ₹1,00,00,000)")
                    _sort_key = ['date', 'entry_time'] if 'entry_time' in _df_bt.columns else ['date']
                    _sorted_trades = _df_bt.sort_values(_sort_key).reset_index(drop=True)
                    _cumul   = _STARTING_CAPITAL
                    _cum_rows = []
                    for _, _tr_r in _sorted_trades.iterrows():
                        _cumul += _tr_r.get('pnl_inr', 0)
                        _cum_rows.append({
                            'Date': _tr_r.get('date', ''),
                            'Time': _tr_r.get('entry_time', ''),
                            'Instrument': _tr_r.get('instrument') or _tr_r.get('idx', ''),
                            'Strategy': str(_tr_r.get('strategy', '')).replace('_SUPERTREND', '+ST'),
                            'Direction': _tr_r.get('direction', ''),
                            'IV%': _tr_r.get('iv_pct', 0),
                            'Mode': _tr_r.get('mode', ''),
                            'P&L ₹': round(_tr_r.get('pnl_inr', 0), 0),
                            'Cumulative ₹': round(_cumul, 0)})

                    if _cum_rows:
                        _cum_df = pd.DataFrame(_cum_rows)

                        def _cum_row_color(row):
                            c = "#f0fdf4" if row.get('P&L ₹', 0) >= 0 else "#fff1f2"
                            return [f"background-color:{c}"] * len(row)

                        st.dataframe(_cum_df.style.apply(_cum_row_color, axis=1),
                                     hide_index=True, use_container_width=True)

                    # Section D — Net portfolio value
                    _net_val   = _STARTING_CAPITAL + _df_bt['pnl_inr'].sum()
                    _net_delta = _net_val - _STARTING_CAPITAL
                    st.metric(
                        label="Portfolio Net Value  (started ₹1,00,00,000)",
                        value=f"₹{_net_val:,.0f}",
                        delta=f"₹{_net_delta:+,.0f}")

            elif "ofs_bt_flat" in st.session_state:
                st.info("No trades found — try selecting more expiry dates or check data availability.")

        # ── Option Strategy Backtest ─────────────────────────────────────────
        with _ofs_sub[3]:

            # ── Excel generator for option strategy backtest ─────────────────
            def _build_excel_ostrat(all_rows: list) -> bytes:
                import io as _io
                _buf = _io.BytesIO()
                if not all_rows:
                    pd.DataFrame().to_excel(_buf, index=False)
                    _buf.seek(0)
                    return _buf.read()

                _df = pd.DataFrame(all_rows)
                _strat_order = [
                    'Long Call', 'Long Put', 'Long Straddle', 'Long Strangle',
                    'Covered Call', 'Protective Put', 'Collar',
                    'Synthetic Long', 'Call Backspread', 'Put Backspread'
                ]

                try:
                    from openpyxl import Workbook
                    from openpyxl.chart import LineChart, Reference
                    from openpyxl.styles import Font, PatternFill, Alignment
                    from openpyxl.utils import get_column_letter

                    wb = Workbook()
                    ws_sum = wb.active
                    ws_sum.title = "P&L Summary"

                    _hdr_fill = PatternFill("solid", fgColor="0F172A")
                    _hdr_font = Font(color="FFFFFF", bold=True)
                    _sum_cols = ['Index', 'Strategy', 'Expiries Run', 'Total P&L (₹)',
                                 'Best Expiry P&L', 'Worst Expiry P&L', 'Win Expiries', 'Lose Expiries']
                    for ci, col in enumerate(_sum_cols, 1):
                        cell = ws_sum.cell(1, ci, col)
                        cell.fill = _hdr_fill; cell.font = _hdr_font
                        cell.alignment = Alignment(horizontal='center')

                    _row = 2
                    for (_idx, _sn), _gdf in _df.groupby(['idx', 'strategy_name']):
                        _pl = _gdf['net_pnl_inr'].tolist()
                        _ws_sum_vals = [
                            _idx, _sn, len(_pl), round(sum(_pl), 0),
                            round(max(_pl), 0), round(min(_pl), 0),
                            sum(1 for p in _pl if p >= 0), sum(1 for p in _pl if p < 0)
                        ]
                        for ci, v in enumerate(_ws_sum_vals, 1):
                            ws_sum.cell(_row, ci, v)
                        _row += 1
                    for col in ws_sum.columns:
                        ws_sum.column_dimensions[get_column_letter(col[0].column)].width = 16

                    # Per-index detail sheets
                    for _idx in _OFS_INDICES:
                        _idf = _df[_df['idx'] == _idx]
                        if _idf.empty:
                            continue
                        ws = wb.create_sheet(title=_idx[:31])
                        _cols = ['expiry', 'strategy_name', 'entry_spot', 'exit_spot',
                                 'dte', 'atm', 'net_pnl_pts', 'net_pnl_inr']
                        for ci, col in enumerate(_cols, 1):
                            cell = ws.cell(1, ci, col)
                            cell.fill = _hdr_fill; cell.font = _hdr_font
                        _green_fill = PatternFill("solid", fgColor="F0FDF4")
                        _red_fill   = PatternFill("solid", fgColor="FFF1F2")
                        for ri, (_, row) in enumerate(_idf[_cols].iterrows(), 2):
                            _fill = _green_fill if row.get('net_pnl_inr', 0) >= 0 else _red_fill
                            for ci, col in enumerate(_cols, 1):
                                ws.cell(ri, ci, row[col] if col in row else "").fill = _fill
                        for col in ws.columns:
                            ws.column_dimensions[get_column_letter(col[0].column)].width = 14

                        # Equity chart: cumulative P&L per strategy across expiries
                        if len(_idf) > 1:
                            _pnl_col = _cols.index('net_pnl_inr') + 1
                            chart = LineChart()
                            chart.title = f"{_idx} — Strategy P&L per Expiry"
                            chart.style = 10
                            chart.y_axis.title = "P&L (₹)"
                            chart.x_axis.title = "Expiry #"
                            chart.width = 22; chart.height = 12
                            _data_ref = Reference(ws, min_col=_pnl_col, min_row=1,
                                                  max_row=len(_idf) + 1)
                            chart.add_data(_data_ref, titles_from_data=True)
                            ws.add_chart(chart, f"{get_column_letter(len(_cols)+2)}2")

                    wb.save(_buf)
                    _buf.seek(0)
                    return _buf.read()

                except ImportError:
                    with pd.ExcelWriter(_buf, engine='openpyxl') as _wr:
                        _df.to_excel(_wr, sheet_name='All Results', index=False)
                    _buf.seek(0)
                    return _buf.read()

            # ── Session state: expiry offsets (separate keys from Signal BT) ─
            for _bidx2 in _OFS_INDICES:
                for _knd in ['opt', 'fut']:
                    _osk = f"osbt_{_knd}_offset_{_bidx2}"
                    if _osk not in st.session_state:
                        st.session_state[_osk] = 0

            st.markdown(
                '<div style="background:#ede9fe;border-radius:8px;padding:8px 12px;'
                'margin-bottom:10px;font-family:JetBrains Mono;font-size:11px;font-weight:700;color:#5b21b6">'
                'OPTION STRATEGY BACKTEST — Enter at period start (BS premiums) · Exit at intrinsic on expiry</div>',
                unsafe_allow_html=True)

            # ── Options expiry selector ──────────────────────────────────────
            st.markdown(
                '<div style="background:#dbeafe;border-radius:8px;padding:6px 12px;'
                'margin-bottom:6px;font-family:JetBrains Mono;font-size:10px;font-weight:700;color:#1d4ed8">'
                'OPTIONS EXPIRY (weekly) — select past expiries</div>',
                unsafe_allow_html=True)

            _osbt_opt_sel = {}
            for _bidx in _OFS_INDICES:
                _off = st.session_state[f"osbt_opt_offset_{_bidx}"]
                _past_opt = _ofs_past_opt(_bidx, n=6, offset=_off)
                _opt_labels = [e.strftime('%d %b %y') for e in _past_opt]

                _oc1, _oc2 = st.columns([5, 1])
                with _oc1:
                    _sel_labels = st.multiselect(
                        f"{_bidx} Options (weekly)",
                        options=_opt_labels, default=_opt_labels,
                        key=f"ms_osbt_opt_{_bidx}_{_off}")
                    _osbt_opt_sel[_bidx] = [_past_opt[_opt_labels.index(l)] for l in _sel_labels]
                with _oc2:
                    st.markdown("<div style='margin-top:26px'></div>", unsafe_allow_html=True)
                    _pb_c1, _pb_c2 = st.columns(2)
                    def _osbt_prev_opt(_i=_bidx):
                        st.session_state[f"osbt_opt_offset_{_i}"] += 6
                    def _osbt_next_opt(_i=_bidx):
                        st.session_state[f"osbt_opt_offset_{_i}"] = max(0, st.session_state[f"osbt_opt_offset_{_i}"] - 6)
                    _pb_c1.button("◀ Prev", key=f"osbt_prev_opt_{_bidx}", on_click=_osbt_prev_opt)
                    _pb_c2.button("Next ▶", key=f"osbt_next_opt_{_bidx}", on_click=_osbt_next_opt,
                                  disabled=_off == 0)

            # ── Futures expiry selector ──────────────────────────────────────
            st.markdown(
                '<div style="background:#fef3c7;border-radius:8px;padding:6px 12px;'
                'margin:8px 0 6px;font-family:JetBrains Mono;font-size:10px;font-weight:700;color:#b45309">'
                'FUTURES EXPIRY (monthly) — select past expiries</div>',
                unsafe_allow_html=True)

            _osbt_fut_sel = {}
            for _bidx in _OFS_INDICES:
                _off = st.session_state[f"osbt_fut_offset_{_bidx}"]
                _past_fut = _ofs_past_fut(_bidx, n=6, offset=_off)
                _fut_labels = [e.strftime('%d %b %y') for e in _past_fut]

                _fc1, _fc2 = st.columns([5, 1])
                with _fc1:
                    _fsel_labels = st.multiselect(
                        f"{_bidx} Futures (monthly)",
                        options=_fut_labels, default=_fut_labels,
                        key=f"ms_osbt_fut_{_bidx}_{_off}")
                    _osbt_fut_sel[_bidx] = [_past_fut[_fut_labels.index(l)] for l in _fsel_labels]
                with _fc2:
                    st.markdown("<div style='margin-top:26px'></div>", unsafe_allow_html=True)
                    _fb_c1, _fb_c2 = st.columns(2)
                    def _osbt_prev_fut(_i=_bidx):
                        st.session_state[f"osbt_fut_offset_{_i}"] += 6
                    def _osbt_next_fut(_i=_bidx):
                        st.session_state[f"osbt_fut_offset_{_i}"] = max(0, st.session_state[f"osbt_fut_offset_{_i}"] - 6)
                    _fb_c1.button("◀ Prev", key=f"osbt_prev_fut_{_bidx}", on_click=_osbt_prev_fut)
                    _fb_c2.button("Next ▶", key=f"osbt_next_fut_{_bidx}", on_click=_osbt_next_fut,
                                  disabled=_off == 0)

            st.caption("Options: ~8-day window per expiry. Futures: ~31-day window. Entry via Black-Scholes premium, exit via intrinsic value.")

            # ── Run button ───────────────────────────────────────────────────
            if st.button("▶ Run All Option Strategies — All Indices", key="osbt_run_btn", type="primary"):
                _osbt_rows = []
                _total_osbt = sum(
                    len(_osbt_opt_sel[_i]) + len(_osbt_fut_sel[_i]) for _i in _OFS_INDICES)
                _osbt_prog = st.progress(0.0)
                _osbt_n = 0

                for _bidx in _OFS_INDICES:
                    for _exp in _osbt_opt_sel[_bidx]:
                        try:
                            _strat_results = _ofs_bt_opt_strat(_bidx, _exp, period_days=8)
                            for _sr in _strat_results:
                                _sr['idx'] = _bidx
                                _sr['expiry_type'] = 'options'
                                _osbt_rows.append(_sr)
                        except Exception:
                            pass
                        _osbt_n += 1
                        _osbt_prog.progress(min(_osbt_n / max(_total_osbt, 1), 1.0))

                    for _exp in _osbt_fut_sel[_bidx]:
                        try:
                            _strat_results = _ofs_bt_opt_strat(_bidx, _exp, period_days=31)
                            for _sr in _strat_results:
                                _sr['idx'] = _bidx
                                _sr['expiry_type'] = 'futures'
                                _osbt_rows.append(_sr)
                        except Exception:
                            pass
                        _osbt_n += 1
                        _osbt_prog.progress(min(_osbt_n / max(_total_osbt, 1), 1.0))

                _osbt_prog.empty()
                st.session_state["osbt_rows"] = _osbt_rows
                _n_exp = sum(len(v) for v in _osbt_opt_sel.values()) + sum(len(v) for v in _osbt_fut_sel.values())
                st.session_state["osbt_label"] = (
                    f"{_n_exp} expiries × 10 strategies × 3 indices | {len(_osbt_rows)} results computed")

            # ── Results display ──────────────────────────────────────────────
            if "osbt_rows" in st.session_state and st.session_state["osbt_rows"]:
                _osbt_rows = st.session_state["osbt_rows"]
                st.caption(st.session_state.get("osbt_label", ""))

                _osbt_df = pd.DataFrame(_osbt_rows)

                # Excel download
                try:
                    _xl_bytes2 = _build_excel_ostrat(_osbt_rows)
                    st.download_button(
                        label="⬇ Download Excel Report",
                        data=_xl_bytes2,
                        file_name=f"option_strategy_bt_{date.today()}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="osbt_excel_dl")
                except Exception as _xle2:
                    st.warning(f"Excel failed: {_xle2}")

                # P&L matrix: rows = expiry, columns = strategies (per index tab)
                _strat_order_disp = [
                    'Long Call', 'Long Put', 'Long Straddle', 'Long Strangle',
                    'Covered Call', 'Protective Put', 'Collar',
                    'Synthetic Long', 'Call Backspread', 'Put Backspread'
                ]

                st.markdown("#### P&L per Expiry × Strategy")
                _tabs_idx = st.tabs([f"**{_i}**" for _i in _OFS_INDICES])
                for _ti, _bidx in zip(_tabs_idx, _OFS_INDICES):
                    with _ti:
                        _idf = _osbt_df[_osbt_df['idx'] == _bidx]
                        if _idf.empty:
                            st.info(f"No results for {_bidx}")
                            continue

                        # Build pivot: rows = expiry, cols = strategy
                        _pivot_rows = []
                        for _exp_str in sorted(_idf['expiry'].unique()):
                            _edf = _idf[_idf['expiry'] == _exp_str]
                            _entry_d = _edf['entry_date'].iloc[0] if 'entry_date' in _edf.columns and len(_edf) > 0 else '—'
                            _exit_d  = _edf['exit_date'].iloc[0]  if 'exit_date'  in _edf.columns and len(_edf) > 0 else _exp_str
                            _row = {'Expiry (Exit)': _exp_str, 'Entry Date': _entry_d, 'Exit Date': _exit_d}
                            _tot = 0.0
                            for _sn in _strat_order_disp:
                                _v = _edf[_edf['strategy_name'] == _sn]['net_pnl_inr'].sum()
                                _row[_sn] = round(_v, 0)
                                _tot += _v
                            _row['TOTAL'] = round(_tot, 0)
                            _pivot_rows.append(_row)

                        # Summary row
                        _sum_row = {'Expiry (Exit)': 'TOTAL', 'Entry Date': '', 'Exit Date': ''}
                        for _sn in _strat_order_disp:
                            _sum_row[_sn] = round(sum(r.get(_sn, 0) for r in _pivot_rows if r.get('Expiry (Exit)') != 'TOTAL'), 0)
                        _sum_row['TOTAL'] = round(sum(r['TOTAL'] for r in _pivot_rows if r.get('Expiry (Exit)') != 'TOTAL'), 0)
                        _pivot_rows.append(_sum_row)

                        _pvdf = pd.DataFrame(_pivot_rows)

                        def _pnl_bg(val):
                            if isinstance(val, (int, float)):
                                return "background-color:#f0fdf4;color:#166534" if val >= 0 else "background-color:#fff1f2;color:#991b1b"
                            return ""

                        _color_cols = [c for c in _strat_order_disp + ['TOTAL'] if c in _pvdf.columns]
                        st.dataframe(
                            _pvdf.style.map(_pnl_bg, subset=_color_cols),
                            hide_index=True, use_container_width=True)

                        # Per-strategy cumulative P&L chart across expiries
                        st.markdown("##### Cumulative P&L by Strategy")
                        _fig2 = go.Figure()
                        for _sn in _strat_order_disp:
                            _sdf2 = _idf[_idf['strategy_name'] == _sn].sort_values('expiry')
                            if _sdf2.empty:
                                continue
                            _cumul = _sdf2['net_pnl_inr'].cumsum().reset_index(drop=True)
                            _fig2.add_trace(go.Scatter(
                                y=_cumul, x=list(range(1, len(_cumul) + 1)),
                                mode='lines+markers', name=_sn, line={'width': 2}))
                        _fig2.update_layout(
                            title=f"{_bidx} — Cumulative P&L per Strategy across Expiries",
                            yaxis_title="Cumulative P&L (₹)", xaxis_title="Expiry #",
                            height=320, margin=dict(l=10, r=10, t=35, b=10),
                            legend=dict(orientation='h', y=-0.25, font=dict(size=9)),
                            paper_bgcolor='#f8fafc', plot_bgcolor='#f8fafc')
                        st.plotly_chart(_fig2, use_container_width=True)

                        # Summary: best / worst strategy
                        _strat_total = {
                            _sn: round(_idf[_idf['strategy_name'] == _sn]['net_pnl_inr'].sum(), 0)
                            for _sn in _strat_order_disp
                        }
                        _sorted_st = sorted(_strat_total.items(), key=lambda x: x[1], reverse=True)
                        _st_c1, _st_c2 = st.columns(2)
                        with _st_c1:
                            st.markdown(
                                f'<div style="background:#f0fdf4;border:1.5px solid #16a34a44;border-radius:8px;'
                                f'padding:10px;font-family:JetBrains Mono;font-size:11px">'
                                f'<b style="color:#16a34a">Best Strategy</b><br>'
                                f'{_sorted_st[0][0]}<br>'
                                f'<span style="font-size:14px;font-weight:700">₹{_sorted_st[0][1]:+,.0f}</span>'
                                f'</div>', unsafe_allow_html=True)
                        with _st_c2:
                            st.markdown(
                                f'<div style="background:#fff1f2;border:1.5px solid #dc262644;border-radius:8px;'
                                f'padding:10px;font-family:JetBrains Mono;font-size:11px">'
                                f'<b style="color:#dc2626">Worst Strategy</b><br>'
                                f'{_sorted_st[-1][0]}<br>'
                                f'<span style="font-size:14px;font-weight:700">₹{_sorted_st[-1][1]:+,.0f}</span>'
                                f'</div>', unsafe_allow_html=True)

            elif "osbt_rows" in st.session_state:
                st.info("No results — check data availability or select more expiries.")


def _plot_of_trade_chart(t):
    """Plotly candlestick for a single OFTrade. Fetches bars around trade date, marks entry/SL/target."""
    try:
        import yfinance as _yf_ch
        import plotly.graph_objects as _go_ch
        _sym_map = {"NIFTY": "^NSEI", "BANKNIFTY": "^NSEBANK", "SENSEX": "^BSESN"}
        _sym  = _sym_map.get(t.instrument, "^NSEI")
        _days = {"5m": 5, "15m": 10, "1h": 30}.get(t.tf, 5)
        try:
            from datetime import datetime as _dtch, timedelta as _tdch, date as _dch
            _td    = _dtch.strptime(t.date, "%Y-%m-%d").date()
        except Exception:
            from datetime import date as _dch, timedelta as _tdch
            _td = _dch.today()
        _start = _td - _tdch(days=_days)
        _end   = _td + _tdch(days=2)
        _df    = _yf_ch.Ticker(_sym).history(start=str(_start), end=str(_end),
                                              interval=t.tf, auto_adjust=True)
        if _df is None or len(_df) < 5:
            return None
        _df.index = pd.to_datetime(_df.index)
        try:
            _df.index = _df.index.tz_convert(None)
        except Exception:
            try:
                _df.index = _df.index.tz_localize(None)
            except Exception:
                pass

        fig = _go_ch.Figure()
        fig.add_trace(_go_ch.Candlestick(
            x=_df.index, open=_df['Open'], high=_df['High'],
            low=_df['Low'], close=_df['Close'],
            increasing_line_color='#16a34a', decreasing_line_color='#dc2626',
            name=t.tf))
        fig.add_hline(y=t.entry, line_color='#3b82f6', line_dash='dash',
                      annotation_text=f"Entry {t.entry:.0f}", annotation_position="right")
        fig.add_hline(y=t.sl,    line_color='#ef4444', line_dash='dot',
                      annotation_text=f"SL {t.sl:.0f}", annotation_position="right")
        fig.add_hline(y=t.target,line_color='#22c55e', line_dash='dot',
                      annotation_text=f"TGT {t.target:.0f}", annotation_position="right")
        _pnl_c  = "#16a34a" if getattr(t, 'pnl_inr', 0) >= 0 else "#dc2626"
        _reason = getattr(t, 'exit_reason', '')
        if _reason == 'SL' and getattr(t, 'pnl_inr', 0) > 0:
            _reason = 'TARGET'
        elif _reason == 'TARGET' and getattr(t, 'pnl_inr', 0) < 0:
            _reason = 'SL'
        fig.update_layout(
            title=(f"{t.instrument} {t.direction} | {t.tf} | {t.date} | "
                   f"{_reason} | P&L: Rs{getattr(t,'pnl_inr',0):+,.0f}"),
            height=320, margin=dict(l=10, r=90, t=40, b=10),
            xaxis_rangeslider_visible=False,
            plot_bgcolor='#0f172a', paper_bgcolor='#0f172a',
            font_color='#e2e8f0', title_font_color=_pnl_c,
        )
        return fig
    except Exception:
        return None


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 4 — OPT/FUT AUTO TRADE
# ═══════════════════════════════════════════════════════════════════════════════
with _tabs[4]:
    st.markdown(
        '<div style="background:linear-gradient(135deg,#0f172a,#1e3a5f);border-radius:12px;'
        'padding:16px 20px;margin-bottom:14px">'
        '<div style="color:#60a5fa;font-size:20px;font-weight:700;font-family:JetBrains Mono">'
        'Opt/Fut Auto Trade</div>'
        '<div style="color:#94a3b8;font-size:11px;margin-top:4px;font-family:JetBrains Mono">'
        'Chandelier Exit · TUX+Supertrend — Options & Futures only · Paper/Live</div>'
        '</div>',
        unsafe_allow_html=True)

    if not OFS_OK:
        st.error(f"Strategy engine unavailable: {_ofs_err_msg}")
    else:
        _oat_sub = st.tabs(["🎯 Controls & Live Trades", "📡 Live Signals", "📋 All Option Strategies", "📊 P&L Dashboard"])

        # ── Controls row ─────────────────────────────────────────────────────
        # Mode / Qty / Paper are configured in the sidebar (shared session_state keys).
        # Only TFs and Strategies are unique to this tab.
        _oat_mode   = st.session_state.get('oat_mode', 'both')
        _oat_qty    = int(st.session_state.get('oat_qty', 1))
        _oat_paper  = bool(st.session_state.get('oat_paper', True))
        _oat_ctrl_c2, _oat_ctrl_c3 = st.columns([2, 3])
        with _oat_ctrl_c2:
            _oat_tfs    = st.multiselect("Timeframes", ["5m", "15m", "1h"],
                                          default=st.session_state.get('oat_tfs', ["5m", "15m", "1h"]),
                                          key="oat_tfs")
            st.info("Always enters **both 1:2 and 1:3** R:R trades simultaneously per signal")
        with _oat_ctrl_c3:
            _oat_strats = st.multiselect("Strategies", _OFS_STRATEGIES,
                                          default=st.session_state.get('oat_strats', _OFS_STRATEGIES),
                                          key="oat_strats")
        st.markdown(
            f'<div style="font-family:JetBrains Mono;font-size:11px;color:#64748b;margin:4px 0">'
            f'Mode: <b>{_oat_mode}</b> | Lots: <b>{_oat_qty}</b> | '
            f'{"📋 Paper" if _oat_paper else "⚡ LIVE"} — configure Mode/Lots/Paper in sidebar</div>',
            unsafe_allow_html=True)

        _oat_rr_mult = 2.0  # always both — passed as default; _scan_and_enter overrides with [2.0, 3.0]
        _oat_key     = "_oat_trader_ALL"   # single instance for all 3 indices
        _oat_running = (st.session_state.get(_oat_key) is not None and
                        st.session_state[_oat_key]._running)
        _oat_trader_obj = st.session_state.get(_oat_key)

        # Heartbeat indicator
        _hb_col1, _hb_col2 = st.columns([3, 2])
        with _hb_col1:
            if _oat_running:
                _last = getattr(_oat_trader_obj, '_last_cycle_at', None)
                _last_txt = _last.strftime('%H:%M:%S') if _last else 'pending first cycle'
                st.markdown(f'<div style="font-family:JetBrains Mono;font-size:11px;color:#10b981">'
                            f'● BOT RUNNING — last cycle: {_last_txt}</div>', unsafe_allow_html=True)
            else:
                st.markdown('<div style="font-family:JetBrains Mono;font-size:11px;color:#f59e0b">'
                            '○ BOT STOPPED — click START below</div>', unsafe_allow_html=True)

        st.caption("Use the sidebar ▶ START ALL / ⏹ STOP ALL to control this engine.")

        _sc = "#16a34a" if _oat_running else "#64748b"
        _sl_indices = "NIFTY · BANKNIFTY · SENSEX" if _oat_running else "STOPPED"
        _sl_detail  = (f" | {_oat_trader_obj.mode.upper()} | R:R 1:2 + 1:3 | "
                       f"{'PAPER' if _oat_trader_obj.paper else 'LIVE'} | TFs={','.join(_oat_trader_obj.timeframes)}"
                       if _oat_running and _oat_trader_obj else "")
        st.markdown(
            f'<div style="background:#f8fafc;border:1.5px solid {_sc};border-radius:8px;'
            f'padding:8px 14px;font-family:JetBrains Mono;font-size:12px;font-weight:700;'
            f'color:{_sc};margin:8px 0">{_sl_indices}{_sl_detail}</div>', unsafe_allow_html=True)

        # ── Tab 1: Controls & Live Trades ────────────────────────────────────
        with _oat_sub[0]:
            if _oat_trader_obj:
                _positions = _oat_trader_obj.positions
                _closed    = _oat_trader_obj.closed
                _total_pnl = _oat_trader_obj.total_pnl

                # Open positions
                st.markdown("##### Open Positions")
                if _positions:
                    _op_rows = []
                    for p in _positions:
                        _spot_now = _oat_trader_obj._get_spot(p.instrument)
                        _live_pnl = (_spot_now - p.entry) * (1 if p.direction == 'LONG' else -1) * p.lot_size * p.qty if _spot_now > 0 else 0
                        _op_invest = round(p.entry * getattr(p,'lot_size',75) * getattr(p,'qty',1), 0)
                        _op_rows.append({
                            'Instr':     p.instrument,
                            'Mode':      p.mode.upper(),
                            'Expiry':    p.expiry_str,
                            'TF':        p.tf,
                            'Strategy':  p.strategy,
                            'Dir':       p.direction,
                            'Entry':     round(p.entry, 2),
                            'Spot':      round(_spot_now, 2) if _spot_now > 0 else '—',
                            'SL':        round(p.sl, 2),
                            'Risk pts':  round((p.sl - p.entry) if p.direction == 'LONG' else (p.entry - p.sl), 1),
                            'Target':    round(p.target, 2),
                            'Invest Rs': f'{_op_invest:,.0f}',
                            'Live P&L':  f'{_live_pnl:+,.0f}',
                            'R:R':       f"1:{getattr(p,'rr_ratio',2.0):.0f}",
                            'Lots':      p.qty,
                            'Time':      (p.time or '')[:5],
                        })
                    st.dataframe(pd.DataFrame(_op_rows), hide_index=True, use_container_width=True)
                else:
                    st.info("No open positions — auto-trader will enter on next fresh signal")

                # Closed trades
                st.markdown("##### Closed Trades Today")
                if _closed:
                    def _sl_risk(t):
                        # always negative — how many pts you lose if SL hit
                        return round((t.sl - t.entry) if t.direction == 'LONG' else (t.entry - t.sl), 1)

                    def _invest(t):
                        ls = getattr(t, 'lot_size', 75)
                        qt = getattr(t, 'qty', 1)
                        return round(t.entry * ls * qt, 0)

                    _cl_rows = [{
                        'Date':      t.date,
                        'Entry T':   (t.time or '')[:5],
                        'Exit T':    (t.exit_time or '')[:5],
                        'Instr':     t.instrument,
                        'Mode':      t.mode.upper(),
                        'Expiry':    t.expiry_str,
                        'TF':        t.tf,
                        'Strategy':  t.strategy,
                        'Dir':       t.direction,
                        'IV%':       round(getattr(t, 'vix_at_entry', 0.0), 1),
                        'Entry':     round(t.entry, 2),
                        'SL':        round(t.sl, 2),
                        'Risk pts':  _sl_risk(t),       # negative = pts at risk
                        'Target':    round(t.target, 2),
                        'Exit':      round(t.exit_price, 2),
                        'Invest Rs': int(_invest(t)),   # capital deployed
                        'R:R':       f"1:{getattr(t,'rr_ratio',2.0):.0f}",
                        'Reason':    ('TARGET' if t.exit_reason == 'SL' and t.pnl_inr > 0
                                      else ('SL' if t.exit_reason == 'TARGET' and t.pnl_inr < 0
                                      else t.exit_reason)),
                        'P&L pts':   round(t.pnl_pts, 1),
                        'P&L Rs':    int(round(t.pnl_inr, 0)),
                    } for t in _closed]

                    _cl_df = pd.DataFrame(_cl_rows)
                    _cl_df['_pnl'] = [t.pnl_inr for t in _closed]

                    def _cl_color(row):
                        pnl = row.get('_pnl', 0)
                        bg  = '#f0fdf4' if pnl > 0 else '#fff1f2'
                        return [f'background-color:{bg}'] * len(row)

                    def _cl_fmt(col, val):
                        if col == 'Risk pts':
                            return f'{val:+.1f}'
                        if col in ('P&L pts',):
                            return f'{val:+.1f}'
                        if col in ('P&L Rs', 'Invest Rs'):
                            return f'{val:+,.0f}' if col == 'P&L Rs' else f'{val:,.0f}'
                        return val

                    _cl_display = _cl_df.drop(columns=['_pnl'])
                    # apply sign formatting to numeric columns
                    for _cc in ['Risk pts', 'P&L pts', 'P&L Rs', 'Invest Rs']:
                        if _cc in _cl_display.columns:
                            if _cc in ('Risk pts', 'P&L pts'):
                                _cl_display[_cc] = _cl_display[_cc].apply(lambda v: f'{v:+.1f}')
                            elif _cc == 'P&L Rs':
                                _cl_display[_cc] = _cl_display[_cc].apply(lambda v: f'{v:+,.0f}')
                            elif _cc == 'Invest Rs':
                                _cl_display[_cc] = _cl_display[_cc].apply(lambda v: f'{v:,.0f}')

                    st.dataframe(
                        _cl_display.style.apply(_cl_color, axis=1),
                        hide_index=True, use_container_width=True)

                    # ── Trade Charts ────────────────────────────────────────
                    st.markdown("##### Trade Charts")
                    _ch_col1, _ch_col2 = st.columns([1, 1])
                    with _ch_col1:
                        _ch_dir = st.selectbox(
                            "Filter by direction", ["All", "LONG", "SHORT"],
                            key="oat_chart_dir")
                    with _ch_col2:
                        _ch_max = st.slider("Max charts", 2, 20, 6, key="oat_chart_max")
                    _ch_trades = [t for t in reversed(_closed)
                                  if _ch_dir == "All" or t.direction == _ch_dir][:_ch_max]
                    _gc1, _gc2 = st.columns(2)
                    for _ci, _ct in enumerate(_ch_trades):
                        _fig_ch = _plot_of_trade_chart(_ct)
                        if _fig_ch:
                            (_gc1 if _ci % 2 == 0 else _gc2).plotly_chart(
                                _fig_ch, use_container_width=True,
                                key=f"oat_ch_{_ct.id}_{_ci}")
                else:
                    st.info("No closed trades yet today")

                if st.toggle("Show Log", key="oat_show_log"):
                    st.code("\n".join(reversed(_oat_trader_obj.log[-40:])), language=None)
            else:
                st.info("Click ▶ START to begin auto-trading NIFTY, BANKNIFTY and SENSEX simultaneously.")

        # ── Tab 2: Live Signals (mirrors Tab 3 live signals for all 3 indices) ──
        with _oat_sub[1]:
            _oat_tf_choices = st.multiselect(
                "Timeframes to scan", ["5m", "15m", "1h"],
                default=["5m", "15m", "1h"], key="oat_sig_tfs")
            _oat_sig_refresh = st.button("&#128260; Refresh Signals", key="oat_sig_refresh")

            if _oat_sig_refresh or "oat_live_sigs" not in st.session_state:
                with st.spinner("Scanning NIFTY, BANKNIFTY, SENSEX..."):
                    _oat_all_sigs = {}
                    _oat_all_exp  = {}
                    for _idx in _OFS_INDICES:
                        try:
                            _oat_all_sigs[_idx] = _ofs_scan_all(_idx, _oat_tf_choices or ["5m"])
                        except Exception:
                            _oat_all_sigs[_idx] = {}
                        try:
                            _oat_all_exp[_idx] = _ofs_all_expiries(_idx)
                        except Exception:
                            _oat_all_exp[_idx] = {}
                    st.session_state["oat_live_sigs"] = _oat_all_sigs
                    st.session_state["oat_live_exp"]  = _oat_all_exp
                    st.session_state["oat_sig_time"]  = datetime.now(_OFS_IST).strftime('%H:%M:%S IST')

            if "oat_live_sigs" in st.session_state:
                st.caption(f"Last scan: {st.session_state.get('oat_sig_time', '—')}")
                _oat_all_sigs = st.session_state["oat_live_sigs"]
                _oat_all_exp  = st.session_state.get("oat_live_exp", {})

                _col_n2, _col_b2, _col_s2 = st.columns(3)
                for _icol2, _idx2 in zip([_col_n2, _col_b2, _col_s2], _OFS_INDICES):
                    _ic2  = _OFS_COLORS[_idx2]
                    _exp2 = _oat_all_exp.get(_idx2, {})
                    _on2  = _exp2.get('opt_near', ('—', '—'))
                    _om2  = _exp2.get('opt_mid',  ('—', '—'))
                    _of2  = _exp2.get('opt_far',  ('—', '—'))
                    _fn2  = _exp2.get('fut_near', ('—', '—'))
                    _ff2  = _exp2.get('fut_far',  ('—', '—'))

                    with _icol2:
                        st.markdown(
                            f'<div style="background:{_ic2};border-radius:10px;padding:10px 14px;margin-bottom:8px">'
                            f'<div style="color:white;font-size:15px;font-weight:800;font-family:JetBrains Mono">{_idx2}</div>'
                            f'</div>', unsafe_allow_html=True)
                        # Expiry badges
                        st.markdown(
                            f'<div style="font-family:JetBrains Mono;font-size:10px;margin-bottom:8px">'
                            f'<div style="background:#dbeafe;border-radius:6px;padding:4px 8px;margin-bottom:3px">'
                            f'<b>Opt 1:</b> {_on2[0]} <span style="color:#64748b">(DTE {_on2[1]})</span></div>'
                            f'<div style="background:#ede9fe;border-radius:6px;padding:4px 8px;margin-bottom:3px">'
                            f'<b>Opt 2:</b> {_om2[0]} <span style="color:#64748b">(DTE {_om2[1]})</span></div>'
                            f'<div style="background:#e0e7ff;border-radius:6px;padding:4px 8px;margin-bottom:3px">'
                            f'<b>Opt 3:</b> {_of2[0]} <span style="color:#64748b">(DTE {_of2[1]})</span></div>'
                            f'<div style="background:#fef3c7;border-radius:6px;padding:4px 8px;margin-bottom:3px">'
                            f'<b>Fut 1:</b> {_fn2[0]} <span style="color:#64748b">(DTE {_fn2[1]})</span></div>'
                            f'<div style="background:#fed7aa;border-radius:6px;padding:4px 8px">'
                            f'<b>Fut 2:</b> {_ff2[0]} <span style="color:#64748b">(DTE {_ff2[1]})</span></div>'
                            f'</div>', unsafe_allow_html=True)
                        # Signal cards per TF
                        _idx2_sigs = _oat_all_sigs.get(_idx2, {})
                        for _tf2 in (_oat_tf_choices or ["5m", "15m", "1h"]):
                            tf2_signals = [s for s in _idx2_sigs.get(_tf2, [])
                                           if isinstance(s, dict) and 'strategy' in s]
                            if not tf2_signals:
                                continue
                            st.markdown(
                                f'<div style="color:{_ic2};font-size:10px;font-weight:700;'
                                f'font-family:JetBrains Mono;margin:8px 0 4px">{_tf2.upper()}</div>',
                                unsafe_allow_html=True)
                            for _sig2 in tf2_signals:
                                st.markdown(_ofs_sig_card(_sig2), unsafe_allow_html=True)
            else:
                st.info("Click 'Refresh Signals' to load current signals for all 3 indices.")

        # ── Tab 3: All Option Strategies (live pricing, all 3 indices) ──────────
        with _oat_sub[2]:
            st.markdown("#### All Option Strategies — Live Pricing")
            st.caption("10 structural strategies × 3 expiries, priced via Black-Scholes. Click Load for each index.")

            _oat_as_itabs = st.tabs([f"🟢 {_idx}" for _idx in _OFS_INDICES])
            for _oat_asi, (_oat_as_itab, _oat_as_idx) in enumerate(zip(_oat_as_itabs, _OFS_INDICES)):
                with _oat_as_itab:
                    _oat_as_load = st.button(f"&#128260; Load {_oat_as_idx}", key=f"oat_as_load_{_oat_as_idx}")
                    if _oat_as_load or f"oat_as_data_{_oat_as_idx}" not in st.session_state:
                        with st.spinner(f"Computing strategies for {_oat_as_idx}..."):
                            try:
                                st.session_state[f"oat_as_data_{_oat_as_idx}"] = _ofs_all_strategies(_oat_as_idx)
                            except Exception as _oat_ase:
                                st.error(f"Error: {_oat_ase}")
                                st.session_state[f"oat_as_data_{_oat_as_idx}"] = []

                    _oat_as_data = st.session_state.get(f"oat_as_data_{_oat_as_idx}", [])
                    if _oat_as_data:
                        _oat_strategy_order = [
                            'Long Call', 'Long Put', 'Long Straddle', 'Long Strangle',
                            'Covered Call', 'Protective Put', 'Collar',
                            'Synthetic Long', 'Call Backspread', 'Put Backspread'
                        ]
                        _oat_strat_map = {s: {} for s in _oat_strategy_order}
                        _oat_expiries_seen = []
                        for _oat_setup in _oat_as_data:
                            _oat_sn  = _oat_setup['strategy_name']
                            _oat_exp = _oat_setup['expiry']
                            if _oat_sn in _oat_strat_map:
                                _oat_strat_map[_oat_sn][_oat_exp] = _oat_setup
                            if _oat_exp not in _oat_expiries_seen:
                                _oat_expiries_seen.append(_oat_exp)

                        _oat_gh = ['Strategy', 'Market View'] + [f"Expiry {i+1}" for i in range(len(_oat_expiries_seen))]
                        _oat_hcols = st.columns([2, 1.5] + [2] * len(_oat_expiries_seen))
                        for _oat_hc, _oat_ht in zip(_oat_hcols, _oat_gh):
                            _oat_hc.markdown(
                                f'<div style="font-family:JetBrains Mono;font-size:10px;'
                                f'font-weight:700;color:#475569;padding:4px 0">{_oat_ht}</div>',
                                unsafe_allow_html=True)

                        _oat_hcols2 = st.columns([2, 1.5] + [2] * len(_oat_expiries_seen))
                        _oat_hcols2[0].markdown("")
                        _oat_hcols2[1].markdown("")
                        for _oat_ei, _oat_ev in enumerate(_oat_expiries_seen):
                            _oat_dte_v = next((s['dte'] for s in _oat_as_data if s['expiry'] == _oat_ev), 0)
                            _oat_hcols2[2 + _oat_ei].markdown(
                                f'<div style="font-family:JetBrains Mono;font-size:9px;'
                                f'color:#64748b;padding:2px 0">{_oat_ev}<br>DTE {_oat_dte_v}</div>',
                                unsafe_allow_html=True)

                        st.markdown('<hr style="margin:4px 0;border-color:#e2e8f0">', unsafe_allow_html=True)

                        _oat_view_colors = {
                            'Bullish': '#16a34a', 'Bearish': '#dc2626',
                            'High Volatility': '#7c3aed', 'Extreme Move': '#6d28d9',
                            'Neutral/Mild Bull': '#0284c7', 'Bullish + Hedged': '#059669',
                            'Neutral/Capped': '#0891b2', 'Futures-Like Risk': '#1d4ed8',
                            'Explosive Up': '#d97706', 'Explosive Down': '#b91c1c',
                        }

                        for _oat_sname in _oat_strategy_order:
                            _oat_row_data = _oat_strat_map.get(_oat_sname, {})
                            if not _oat_row_data:
                                continue
                            _oat_first = next(iter(_oat_row_data.values()))
                            _oat_view  = _oat_first.get('market_view', '')
                            _oat_vc    = _oat_view_colors.get(_oat_view, '#374151')
                            _oat_rcols = st.columns([2, 1.5] + [2] * len(_oat_expiries_seen))
                            _oat_rcols[0].markdown(
                                f'<div style="font-family:JetBrains Mono;font-size:10px;'
                                f'font-weight:700;color:#1e293b;padding:6px 0">{_oat_sname}</div>',
                                unsafe_allow_html=True)
                            _oat_rcols[1].markdown(
                                f'<div style="font-family:JetBrains Mono;font-size:9px;'
                                f'color:{_oat_vc};font-weight:600;padding:6px 0">{_oat_view}</div>',
                                unsafe_allow_html=True)
                            for _oat_ei, _oat_ev in enumerate(_oat_expiries_seen):
                                _oat_cell = _oat_row_data.get(_oat_ev)
                                if _oat_cell is None:
                                    _oat_rcols[2 + _oat_ei].markdown("—")
                                    continue
                                _oat_sign = _oat_cell.get('net_sign', 'DEBIT')
                                _oat_bg   = "#f0fdf4" if _oat_sign == 'DEBIT' else "#fff1f2"
                                _oat_bc   = "#16a34a" if _oat_sign == 'DEBIT' else "#dc2626"
                                _oat_net  = _oat_cell.get('net_premium', 0)
                                _oat_blo  = _oat_cell.get('breakeven_lo', 0)
                                _oat_bhi  = _oat_cell.get('breakeven_hi', 0)
                                _oat_legs = _oat_cell.get('legs', [])
                                _oat_leg_str = ' | '.join(
                                    f'{"BUY" if l["type"]=="BUY" else ("SELL" if l["type"]=="SELL" else l["type"])}'
                                    f' {l["option_type"]}'
                                    + (f'×{l.get("qty",1)}' if l.get('qty', 1) > 1 else '')
                                    + (f' {l["strike"]}' if l.get('strike') else '')
                                    for l in _oat_legs if l['option_type'] != 'FUT'
                                )
                                _oat_rcols[2 + _oat_ei].markdown(
                                    f'<div style="background:{_oat_bg};border:1px solid {_oat_bc}44;'
                                    f'border-radius:6px;padding:5px 7px;font-family:JetBrains Mono;font-size:8px">'
                                    f'<div style="color:{_oat_bc};font-weight:700">{_oat_sign} &#8377;{_oat_net:.0f}</div>'
                                    f'<div style="color:#374151;margin-top:2px">{_oat_leg_str[:40]}</div>'
                                    f'<div style="color:#64748b;margin-top:2px">'
                                    f'BEP: {_oat_blo:.0f}{"/" + str(round(_oat_bhi)) if _oat_blo != _oat_bhi else ""}'
                                    f'</div></div>',
                                    unsafe_allow_html=True)
                            st.markdown('<hr style="margin:2px 0;border-color:#f1f5f9">', unsafe_allow_html=True)
                    elif not _oat_as_load:
                        st.info(f"Click 'Load {_oat_as_idx}' to compute all 10 strategies × 3 expiries")

        # ── Tab 4: P&L Dashboard ─────────────────────────────────────────────
        with _oat_sub[3]:
            if _oat_trader_obj and _oat_trader_obj.closed:
                _closed2 = _oat_trader_obj.closed
                _total   = _oat_trader_obj.total_pnl

                # Metrics row
                _wins2   = [t for t in _closed2 if t.pnl_inr > 0]
                _losses2 = [t for t in _closed2 if t.pnl_inr <= 0]
                _gross_w = sum(t.pnl_inr for t in _wins2)
                _gross_l = abs(sum(t.pnl_inr for t in _losses2))
                _pf      = round(_gross_w / _gross_l, 2) if _gross_l > 0 else 99.0
                _wr      = round(len(_wins2) / len(_closed2) * 100, 1) if _closed2 else 0

                _m1, _m2, _m3, _m4, _m5 = st.columns(5)
                _m1.metric("Trades", len(_closed2))
                _m2.metric("Win Rate", f"{_wr:.1f}%")
                _m3.metric("Profit Factor", f"{_pf:.2f}")
                _m4.metric("Wins / Losses", f"{len(_wins2)} / {len(_losses2)}")
                _pnl_delta = f"₹{_total:+,.0f}"
                _m5.metric("Net P&L", _pnl_delta)

                # Equity curve
                import plotly.graph_objects as go
                _eq_vals = [t.pnl_inr for t in _closed2]
                _cumul_eq = pd.Series(_eq_vals).cumsum()
                _trade_labels = [f"{t.direction} {t.instrument} {t.strategy[:8]}"
                                 for t in _closed2]
                _eq_fig = go.Figure()
                _eq_fig.add_trace(go.Scatter(
                    x=list(range(len(_cumul_eq))),
                    y=_cumul_eq.tolist(),
                    mode='lines+markers',
                    text=_trade_labels,
                    hovertemplate='Trade %{x}: %{text}<br>Cumul P&L: ₹%{y:,.0f}',
                    line={'color': '#1d4ed8', 'width': 2},
                    marker={'color': ['#16a34a' if v >= 0 else '#dc2626' for v in _eq_vals],
                            'size': 7}))
                _eq_fig.update_layout(
                    title="Cumulative P&L — Today",
                    yaxis_title="P&L (₹)", xaxis_title="Trade #",
                    height=280, margin=dict(l=10, r=10, t=35, b=10),
                    paper_bgcolor='#f8fafc', plot_bgcolor='#f8fafc',
                    shapes=[{'type':'line','x0':0,'x1':len(_cumul_eq),
                             'y0':0,'y1':0,'line':{'color':'#94a3b8','dash':'dot'}}])
                st.plotly_chart(_eq_fig, use_container_width=True)

                # Per-strategy breakdown
                st.markdown("##### By Strategy")
                for _strat2 in _OFS_STRATEGIES:
                    _st_trades = [t for t in _closed2 if t.strategy == _strat2]
                    if not _st_trades:
                        continue
                    _sp = sum(t.pnl_inr for t in _st_trades)
                    _sc2 = "#16a34a" if _sp >= 0 else "#dc2626"
                    st.markdown(
                        f'<div style="background:#f8fafc;border:1px solid #e2e8f0;'
                        f'border-radius:8px;padding:8px 12px;margin-bottom:6px;'
                        f'font-family:JetBrains Mono;display:flex;justify-content:space-between">'
                        f'<span style="font-size:11px;font-weight:700">{_strat2.replace("_"," ")}</span>'
                        f'<span style="font-size:11px;font-weight:700;color:{_sc2}">₹{_sp:+,.0f}</span>'
                        f'<span style="font-size:10px;color:#64748b">{len(_st_trades)} trades | '
                        f'{sum(1 for t in _st_trades if t.pnl_inr>0)}/{len(_st_trades)} W</span>'
                        f'</div>', unsafe_allow_html=True)
            else:
                st.info("No closed trades yet. Start the auto-trader and wait for trades to complete.")

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 5 — P&L SUMMARY
# ═══════════════════════════════════════════════════════════════════════════════
with _tabs[5]:
    from collections import defaultdict as _ps_dd
    import numpy as _ps_np

    st.markdown('<div style="color:#1d4ed8;font-size:20px;font-weight:700;font-family:JetBrains Mono">P&L Summary</div>', unsafe_allow_html=True)
    st.caption("Paper trade results · Date-wise · Capital tracker · Auto-saved")

    # ── Capital input ─────────────────────────────────────────────────────────
    _ps_c1, _ps_c2 = st.columns([2,3])
    with _ps_c1:
        _ps_capital = st.number_input("Starting Capital (Rs)", value=100000, step=10000,
                                      min_value=10000, key="ps_capital",
                                      help="Your paper trading starting capital")
    with _ps_c2:
        _ps_paper_lbl = "PAPER TRADE MODE" if st.session_state.get('sb_paper_mode', True) else "LIVE TRADE MODE"
        _ps_paper_c   = "#166534" if st.session_state.get('sb_paper_mode', True) else "#dc2626"
        st.markdown(
            f'<div style="background:#f0fdf4;border:1px solid #86efac;border-radius:8px;'
            f'padding:10px 14px;margin-top:22px;font-family:JetBrains Mono">'
            f'<span style="color:{_ps_paper_c};font-size:13px;font-weight:700">{_ps_paper_lbl}</span>'
            f'<span style="color:#5a72a0;font-size:10px;margin-left:10px">All trades simulated · No real money</span>'
            f'</div>', unsafe_allow_html=True)

    # ── Load + enrich all trades ──────────────────────────────────────────────
    def _ps_seg(t):
        seg = str(t.get('Segment', t.get('segment', 'Equity'))).upper()
        if any(x in seg for x in ('OPT','CE','PE')): return 'Options'
        if any(x in seg for x in ('FUT','FUTURE')):  return 'Futures'
        return 'Equity'

    def _ps_inv(t):
        seg = _ps_seg(t)
        qty = float(t.get('Qty', t.get('qty', t.get('contracts', 1))) or 1)
        ep  = _trade_entry(t)
        # qty is already in contract units (lot size already baked in by the bot)
        if 'FUT' in seg.upper(): return round(ep * qty * 0.1, 2)   # 10% margin
        if 'OPT' in seg.upper(): return round(ep * qty, 2)
        return round(ep * qty, 2)

    def _ps_charges(t):
        stored = float(t.get('Charges Rs', t.get('Charges ₹', t.get('charges', -1))) or -1)
        seg = _ps_seg(t)
        qty = float(t.get('Qty', t.get('qty', t.get('contracts', 1))) or 1)
        ep  = _trade_entry(t); xp = _trade_exit(t)
        # qty is already total units — do NOT multiply by lot again
        bt = ep * qty; st_ = xp * qty; to = bt + st_; brok = 40.0
        if 'OPT' in seg.upper():
            stt = st_ * 0.0005; trans = to * 0.00053; sebi = to * 0.000001; stamp = bt * 0.00003
        elif 'FUT' in seg.upper():
            stt = st_ * 0.0001; trans = to * 0.0000193; sebi = to * 0.000001; stamp = bt * 0.00002
        else:
            stt = st_ * 0.00025; trans = to * 0.0000325; sebi = to * 0.000001; stamp = bt * 0.00015
        gst = (brok + trans + sebi) * 0.18
        calc = round(brok + stt + trans + sebi + stamp + gst, 2)
        return stored if (stored > 0 and stored < calc * 10) else calc

    _ps_all_raw = _all_trades()
    _ps_all = []
    for _pt in _ps_all_raw:
        _pt2 = dict(_pt)
        _pt2['_seg']  = _ps_seg(_pt)
        _pt2['_net']  = _trade_net_pnl(_pt)
        _pt2['_gross']= _trade_gross_pnl(_pt)
        _pt2['_chg']  = _ps_charges(_pt)
        _pt2['_inv']  = _ps_inv(_pt)
        _pt2['_date'] = _pt.get('Date', _pt.get('date', str(__import__('datetime').date.today())))
        _ps_all.append(_pt2)

    # ── Grand totals ──────────────────────────────────────────────────────────
    _ps_tot_tr  = len(_ps_all)
    _ps_tot_wins= sum(1 for t in _ps_all if t['_net'] > 0)
    _ps_tot_inv = sum(t['_inv']   for t in _ps_all)
    _ps_tot_gross=sum(t['_gross'] for t in _ps_all)
    _ps_tot_chg = sum(t['_chg']   for t in _ps_all)
    _ps_tot_net = sum(t['_net']   for t in _ps_all)
    _ps_wr      = _ps_tot_wins / _ps_tot_tr * 100 if _ps_tot_tr else 0
    _ps_balance = _ps_capital + _ps_tot_net          # capital remaining
    _ps_ret_pct = _ps_tot_net / _ps_capital * 100 if _ps_capital else 0

    _ps_nc  = "#166534" if _ps_tot_net  >= 0 else "#dc2626"
    _ps_bc  = "#166534" if _ps_balance >= _ps_capital else "#dc2626"

    # ── Big summary banner ────────────────────────────────────────────────────
    _ps_bg  = "#f0fdf4" if _ps_tot_net >= 0 else "#fff1f2"
    st.markdown(
        f'<div style="background:{_ps_bg};border:2px solid {_ps_nc}33;border-radius:14px;'
        f'padding:16px 22px;margin-bottom:14px;font-family:JetBrains Mono">'
        f'<div style="font-size:11px;font-weight:700;color:{_ps_nc};margin-bottom:4px">'
        f'PAPER TRADING RESULTS — EQUITY * OPTIONS * FUTURES</div>'
        f'<div style="display:flex;gap:32px;flex-wrap:wrap;margin-top:6px">'
        f'<div><div style="font-size:10px;color:#64748b">Total Invested</div>'
        f'<div style="font-size:22px;font-weight:700;color:#1e293b">Rs{_ps_tot_inv:,.0f}</div></div>'
        f'<div><div style="font-size:10px;color:#64748b">Total P&L</div>'
        f'<div style="font-size:22px;font-weight:700;color:{_ps_nc}">Rs{_ps_tot_net:+,.0f}</div></div>'
        f'<div><div style="font-size:10px;color:#64748b">Charges Paid</div>'
        f'<div style="font-size:22px;font-weight:700;color:#dc2626">Rs{_ps_tot_chg:,.0f}</div></div>'
        f'<div><div style="font-size:10px;color:#64748b">Amount Left</div>'
        f'<div style="font-size:22px;font-weight:700;color:{_ps_bc}">Rs{_ps_balance:,.0f}</div>'
        f'<div style="font-size:10px;color:{_ps_bc}">{_ps_ret_pct:+.2f}% return</div></div>'
        f'</div></div>', unsafe_allow_html=True)

    # ── Segment metric cards ──────────────────────────────────────────────────
    _ps_m1, _ps_m2, _ps_m3, _ps_m4, _ps_m5, _ps_m6 = st.columns(6)
    _ps_m1.metric("Trades",    str(_ps_tot_tr))
    _ps_m2.metric("Win Rate",  f"{_ps_wr:.1f}%")
    _ps_m3.metric("Equity Net",f"Rs{sum(t['_net'] for t in _ps_all if t['_seg']=='Equity'):+,.0f}")
    _ps_m4.metric("Futures Net",f"Rs{sum(t['_net'] for t in _ps_all if t['_seg']=='Futures'):+,.0f}")
    _ps_m5.metric("Options Net",f"Rs{sum(t['_net'] for t in _ps_all if t['_seg']=='Options'):+,.0f}")
    _ps_m6.metric("Capital Left",f"Rs{_ps_balance:,.0f}")

    st.markdown("---")

    if not _ps_all:
        st.info("No paper trades yet. Start the bots and trades will appear here automatically.")
    else:
        # ── Filters ───────────────────────────────────────────────────────────
        _psf1, _psf2, _psf3, _psf4 = st.columns(4)
        with _psf1: _ps_period = st.selectbox("Period", ["Today","Last 7 Days","Last 30 Days","All Time"], key="ps_period")
        with _psf2: _ps_fseg   = st.selectbox("Segment",["All","Equity","Futures","Options"], key="ps_fseg")
        with _psf3: _ps_fdir   = st.selectbox("Direction",["All","LONG","SHORT"], key="ps_fdir")
        with _psf4: _ps_fres   = st.selectbox("Result",  ["All","WIN","LOSS"], key="ps_fres")

        _today_str = str(__import__('datetime').date.today())
        _ps_filt = _ps_all[:]
        if _ps_period == "Today":
            _ps_filt = [t for t in _ps_filt if t['_date'] == _today_str]
        elif _ps_period == "Last 7 Days":
            _cut3 = (__import__('datetime').date.today()-__import__('datetime').timedelta(days=7)).isoformat()
            _ps_filt = [t for t in _ps_filt if t['_date'] >= _cut3]
        elif _ps_period == "Last 30 Days":
            _cut3 = (__import__('datetime').date.today()-__import__('datetime').timedelta(days=30)).isoformat()
            _ps_filt = [t for t in _ps_filt if t['_date'] >= _cut3]
        if _ps_fseg != "All": _ps_filt = [t for t in _ps_filt if t['_seg'] == _ps_fseg]
        if _ps_fdir != "All": _ps_filt = [t for t in _ps_filt if t.get('Direction', t.get('direction','')) == _ps_fdir]
        if _ps_fres == "WIN":  _ps_filt = [t for t in _ps_filt if t['_net'] > 0]
        elif _ps_fres == "LOSS": _ps_filt = [t for t in _ps_filt if t['_net'] <= 0]

        # ── DATE-WISE SUMMARY TABLE ───────────────────────────────────────────
        st.markdown('<div style="color:#1d4ed8;font-weight:700;font-family:JetBrains Mono;font-size:14px;margin-bottom:6px">Date-wise P&L</div>', unsafe_allow_html=True)

        _ps_daily = _ps_dd(lambda: {"trades":0,"wins":0,"invested":0.0,"gross":0.0,"charges":0.0,"net":0.0})
        for _pt3 in _ps_filt:
            _d3 = _pt3['_date']
            _ps_daily[_d3]["trades"]   += 1
            _ps_daily[_d3]["wins"]     += 1 if _pt3['_net'] > 0 else 0
            _ps_daily[_d3]["invested"] += _pt3['_inv']
            _ps_daily[_d3]["gross"]    += _pt3['_gross']
            _ps_daily[_d3]["charges"]  += _pt3['_chg']
            _ps_daily[_d3]["net"]      += _pt3['_net']

        _ps_cum = 0.0
        _ps_bal = float(_ps_capital)
        _ps_daily_sorted = sorted(_ps_daily.items())

        _ps_tbl = (
            '<div style="overflow-x:auto"><table style="width:100%;border-collapse:collapse;'
            'font-family:JetBrains Mono;font-size:11px">'
            '<thead><tr style="background:#1e293b;color:#fff">'
            '<th style="padding:8px 10px;text-align:left">Date</th>'
            '<th style="padding:8px 10px;text-align:center">Trades</th>'
            '<th style="padding:8px 10px;text-align:center">W/L</th>'
            '<th style="padding:8px 10px;text-align:right">Invested Rs</th>'
            '<th style="padding:8px 10px;text-align:right">Gross P&L</th>'
            '<th style="padding:8px 10px;text-align:right">Charges</th>'
            '<th style="padding:8px 10px;text-align:right">Net P&L</th>'
            '<th style="padding:8px 10px;text-align:right">Win %</th>'
            '<th style="padding:8px 10px;text-align:right">Return %</th>'
            '<th style="padding:8px 10px;text-align:right">Cumulative P&L</th>'
            '<th style="padding:8px 10px;text-align:right">Amount Left</th>'
            '</tr></thead><tbody>'
        )
        _ps_tot2 = {"tr":0,"wi":0,"inv":0.0,"gr":0.0,"ch":0.0,"net":0.0}
        for _ddate2, _dv2 in _ps_daily_sorted:
            _ps_cum  += _dv2["net"]
            _ps_bal  = _ps_capital + _ps_cum
            _dwr2    = _dv2["wins"] / _dv2["trades"] * 100 if _dv2["trades"] else 0
            _dret2   = _dv2["net"] / _dv2["invested"] * 100 if _dv2["invested"] > 0 else 0.0
            _dnc2    = "#166534" if _dv2["net"] >= 0 else "#dc2626"
            _dbc2    = "#166534" if _ps_bal >= _ps_capital else "#dc2626"
            _dcnc2   = "#166534" if _ps_cum >= 0 else "#dc2626"
            try: _dlbl2 = __import__('datetime').datetime.strptime(_ddate2, '%Y-%m-%d').strftime('%d %b %Y')
            except: _dlbl2 = _ddate2
            _drow_bg = "#f0fdf4" if _dv2["net"] >= 0 else "#fff1f2"
            _ps_tbl += (
                f'<tr style="background:{_drow_bg};border-bottom:1px solid #e2e8f0">'
                f'<td style="padding:6px 10px;font-weight:700">{_dlbl2}</td>'
                f'<td style="padding:6px 10px;text-align:center">{_dv2["trades"]}</td>'
                f'<td style="padding:6px 10px;text-align:center">'
                f'<span style="color:#166534">{_dv2["wins"]}</span>/<span style="color:#dc2626">{_dv2["trades"]-_dv2["wins"]}</span></td>'
                f'<td style="padding:6px 10px;text-align:right;color:#5a72a0">Rs{_dv2["invested"]:,.0f}</td>'
                f'<td style="padding:6px 10px;text-align:right">Rs{_dv2["gross"]:+,.2f}</td>'
                f'<td style="padding:6px 10px;text-align:right;color:#dc2626">Rs{_dv2["charges"]:,.2f}</td>'
                f'<td style="padding:6px 10px;text-align:right;font-weight:700;color:{_dnc2}">Rs{_dv2["net"]:+,.2f}</td>'
                f'<td style="padding:6px 10px;text-align:right">{_dwr2:.0f}%</td>'
                f'<td style="padding:6px 10px;text-align:right;color:{_dnc2}">{_dret2:+.2f}%</td>'
                f'<td style="padding:6px 10px;text-align:right;font-weight:700;color:{_dcnc2}">Rs{_ps_cum:+,.2f}</td>'
                f'<td style="padding:6px 10px;text-align:right;font-weight:700;color:{_dbc2}">Rs{_ps_bal:,.0f}</td>'
                f'</tr>'
            )
            _ps_tot2["tr"] += _dv2["trades"]; _ps_tot2["wi"] += _dv2["wins"]
            _ps_tot2["inv"] += _dv2["invested"]; _ps_tot2["gr"] += _dv2["gross"]
            _ps_tot2["ch"] += _dv2["charges"]; _ps_tot2["net"] += _dv2["net"]

        _ps_tot_nc2 = "#86efac" if _ps_tot2["net"] >= 0 else "#f87171"
        _ps_tot_ret2 = _ps_tot2["net"] / _ps_tot2["inv"] * 100 if _ps_tot2["inv"] > 0 else 0
        _ps_tot_bal2 = _ps_capital + _ps_tot2["net"]
        _ps_tbl += (
            f'<tr style="background:#1e293b;color:#fff;font-weight:700">'
            f'<td style="padding:8px 10px">TOTAL</td>'
            f'<td style="padding:8px 10px;text-align:center">{_ps_tot2["tr"]}</td>'
            f'<td style="padding:8px 10px;text-align:center">{_ps_tot2["wi"]}/{_ps_tot2["tr"]-_ps_tot2["wi"]}</td>'
            f'<td style="padding:8px 10px;text-align:right">Rs{_ps_tot2["inv"]:,.0f}</td>'
            f'<td style="padding:8px 10px;text-align:right">Rs{_ps_tot2["gr"]:+,.2f}</td>'
            f'<td style="padding:8px 10px;text-align:right">Rs{_ps_tot2["ch"]:,.2f}</td>'
            f'<td style="padding:8px 10px;text-align:right;color:{_ps_tot_nc2}">Rs{_ps_tot2["net"]:+,.2f}</td>'
            f'<td style="padding:8px 10px;text-align:right">{_ps_tot2["wi"]/_ps_tot2["tr"]*100 if _ps_tot2["tr"] else 0:.0f}%</td>'
            f'<td style="padding:8px 10px;text-align:right;color:{_ps_tot_nc2}">{_ps_tot_ret2:+.2f}%</td>'
            f'<td style="padding:8px 10px;text-align:right;color:{_ps_tot_nc2}">Rs{_ps_tot2["net"]:+,.2f}</td>'
            f'<td style="padding:8px 10px;text-align:right;color:{_ps_tot_nc2}">Rs{_ps_tot_bal2:,.0f}</td>'
            f'</tr></tbody></table></div>'
        )
        st.markdown(_ps_tbl, unsafe_allow_html=True)

        st.markdown("---")

        # ── Equity curve chart ────────────────────────────────────────────────
        if len(_ps_daily_sorted) > 0:
            _ps_dates_plt = [d for d,_ in _ps_daily_sorted]
            _ps_cum_plt   = []
            _ps_bal_plt   = []
            _running_cum  = 0.0
            for _,_dv_p in _ps_daily_sorted:
                _running_cum += _dv_p["net"]
                _ps_cum_plt.append(_running_cum)
                _ps_bal_plt.append(_ps_capital + _running_cum)

            _ps_efig = go.Figure()
            _ps_efig.add_trace(go.Bar(
                x=_ps_dates_plt, y=[v["net"] for _,v in _ps_daily_sorted],
                name="Daily Net P&L",
                marker_color=["#16a34a" if v["net"]>=0 else "#dc2626" for _,v in _ps_daily_sorted],
            ))
            _ps_efig.add_trace(go.Scatter(
                x=_ps_dates_plt, y=_ps_cum_plt, name="Cumulative P&L",
                mode="lines+markers", yaxis="y2",
                line=dict(color="#1d4ed8", width=2), marker=dict(size=6),
            ))
            _ps_efig.update_layout(
                title="Daily P&L + Cumulative",
                paper_bgcolor="#f0f6ff", plot_bgcolor="#f8faff",
                height=300, font=dict(family="JetBrains Mono", color="#1e293b", size=11),
                yaxis=dict(title="Daily Net (Rs)", tickprefix="Rs", gridcolor="#dde9ff"),
                yaxis2=dict(title="Cumulative (Rs)", tickprefix="Rs", overlaying="y", side="right"),
                legend=dict(orientation="h", y=1.1),
                barmode="relative",
            )
            st.plotly_chart(_ps_efig, width='stretch')

        # ── Trade-by-trade table ──────────────────────────────────────────────
        with st.expander(f"All Trades ({len(_ps_filt)} records)", expanded=False):
            _render_pnl_table(_ps_filt, "ps_tbt")

        # ── Export ────────────────────────────────────────────────────────────
        _pse1, _pse2, _pse3 = st.columns(3)
        with _pse1:
            if _ps_daily_sorted:
                _ps_exp_df = pd.DataFrame([{
                    "Date":d,"Trades":v["trades"],"W/L":f'{v["wins"]}/{v["trades"]-v["wins"]}',
                    "Invested Rs":round(v["invested"],2),"Gross P&L":round(v["gross"],2),
                    "Charges Rs":round(v["charges"],2),"Net P&L":round(v["net"],2),
                    "Win%":round(v["wins"]/v["trades"]*100,1) if v["trades"] else 0,
                    "Return%":round(v["net"]/v["invested"]*100,2) if v["invested"] else 0
                } for d,v in _ps_daily_sorted])
                st.download_button("Export Summary CSV", _ps_exp_df.to_csv(index=False),
                    f"india_summary_{str(__import__('datetime').date.today())}.csv", "text/csv", key="ps_exp_sum")
        with _pse2:
            if _ps_filt:
                _ps_all_df = pd.DataFrame([{k:v for k,v in t.items() if not k.startswith('_')} for t in _ps_filt])
                st.download_button("Export All Trades CSV", _ps_all_df.to_csv(index=False),
                    f"india_trades_all_{str(__import__('datetime').date.today())}.csv", "text/csv", key="ps_exp_all")
        with _pse3:
            st.markdown(
                f'<div style="background:#eff6ff;border:1px solid #1d4ed8;border-radius:8px;'
                f'padding:8px 12px;font-family:JetBrains Mono;font-size:10px;color:#1d4ed8">'
                f'Auto-saved daily to:<br>results/trades_YYYY-MM-DD.json + .csv</div>',
                unsafe_allow_html=True)


with _tabs[6]:
    st.markdown(f'<div style="color:#1d4ed8;font-size:20px;font-weight:700;font-family:JetBrains Mono">Settings</div>', unsafe_allow_html=True)

    _stg1, _stg2 = st.columns(2)

    with _stg1:
        st.markdown("**Angel One API Configuration**")
        st.caption("Store credentials in .env file for security")
        st.markdown(
            '<div style="background:#fff;border:1px solid #dde9ff;border-radius:8px;padding:12px;font-family:JetBrains Mono;font-size:11px">'
            f'<div style="color:#5a72a0">API Key: {"✅ Set" if os.getenv("ANGEL_API_KEY") else "❌ Not set"}</div>'
            f'<div style="color:#5a72a0">Client ID: {"✅ Set" if os.getenv("ANGEL_CLIENT_ID") else "❌ Not set"}</div>'
            f'<div style="color:#5a72a0">MPIN: {"✅ Set" if os.getenv("ANGEL_MPIN") else "❌ Not set"}</div>'
            f'<div style="color:#5a72a0">TOTP Key: {"✅ Set" if os.getenv("ANGEL_TOTP_KEY") else "❌ Not set"}</div>'
            '</div>',
            unsafe_allow_html=True)
        st.caption("Edit the `.env` file in the India trader folder to update credentials.")

        st.markdown("---")
        st.markdown("**Angel One Status**")
        if st.session_state.angel_connected:
            _funds2 = st.session_state.angel_funds
            st.success("Connected to Angel One")
            if _funds2:
                _fa2, _fb2 = st.columns(2)
                with _fa2:
                    st.metric("Available Cash", f"₹{float(_funds2.get('availablecash',0)):,.0f}")
                with _fb2:
                    st.metric("Net", f"₹{float(_funds2.get('net',0)):,.0f}")
        else:
            st.warning("Not connected to Angel One")

    with _stg2:
        st.markdown("**Market Configuration**")
        st.markdown(
            '<div style="background:#fff;border:1px solid #dde9ff;border-radius:8px;padding:12px;font-family:JetBrains Mono;font-size:11px">'
            '<div style="color:#1d4ed8;font-weight:700;margin-bottom:6px">Market Hours (IST)</div>'
            '<div style="color:#5a72a0">Market Opens: 9:15 AM IST</div>'
            '<div style="color:#5a72a0">Market Closes: 3:30 PM IST</div>'
            '<div style="color:#5a72a0">Bot Trading: 9:20 AM - 3:15 PM IST</div>'
            '<div style="color:#5a72a0">Auto Square Off: 3:15 PM IST</div>'
            '</div>',
            unsafe_allow_html=True)

        st.markdown("**Instrument Lot Sizes**")
        _ls_data = [{"Instrument": k, "Lot Size": v, "Margin (approx ₹)": f"₹{v*24500*0.06:,.0f}"} for k,v in NSE_LOT_SIZE.items()]
        st.dataframe(pd.DataFrame(_ls_data), width='stretch', hide_index=True)

        st.markdown("**Results Folder**")
        _n_files = len(_glob_mod.glob(os.path.join(_RESULTS_DIR, "trades_*.json")))
        _n_trades = len(_all_trades())
        st.markdown(
            f'<div style="background:#f0f6ff;border:1px solid #dde9ff;border-radius:8px;padding:10px;font-family:JetBrains Mono;font-size:11px">'
            f'<div style="color:#1d4ed8;font-weight:700">{_n_files} result files · {_n_trades} total trades</div>'
            f'<div style="color:#5a72a0;margin-top:4px">{_RESULTS_DIR}</div>'
            f'</div>',
            unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 7 — TESTING  (DLE Smart Money Concepts Strategy)
# ═══════════════════════════════════════════════════════════════════════════════
with _tabs[7]:
    st.markdown(
        '<div style="background:linear-gradient(135deg,#0f172a,#1e3a5f);border-radius:12px;'
        'padding:16px 20px;margin-bottom:14px">'
        '<div style="color:#60a5fa;font-size:20px;font-weight:700;font-family:JetBrains Mono">'
        'DLE Strategy — Direction · Location · Execution</div>'
        '<div style="color:#94a3b8;font-size:11px;margin-top:4px;font-family:JetBrains Mono">'
        'Smart Money Concepts · HTF=4H Bias → MTF=1H POI (Order Blocks) → LTF=15M Execution</div>'
        '</div>',
        unsafe_allow_html=True)

    # ── Config ──────────────────────────────────────────────────────────────
    _dl_c1, _dl_c2, _dl_c3, _dl_c4 = st.columns(4)
    with _dl_c1:
        _dl_idx    = st.selectbox("Index", ["NIFTY","BANKNIFTY","SENSEX"], key="dl_idx")
    with _dl_c2:
        _dl_days   = st.selectbox("Period", [30,45,60,90], format_func=lambda x:f"{x} days", key="dl_days")
    with _dl_c3:
        _dl_swing  = st.selectbox("Swing Lookback", [3,5,7], index=1, key="dl_swing",
                                   help="Candles each side to confirm swing H/L")
    with _dl_c4:
        _dl_imp    = st.selectbox("Min Impulse Candles", [2,3,4], index=0, key="dl_imp",
                                   help="Consecutive candles to confirm OB impulse")

    # ── Architecture explainer ───────────────────────────────────────────────
    with st.expander("Strategy Architecture — 3-Step Pipeline", expanded=False):
        st.markdown("""
**Step 1 — 4H Directional Bias (BOS)**
- Detect swing highs/lows using a rolling `N`-candle window.
- **Bullish BOS**: candle closes above previous swing high → Bias = LONG only.
- **Bearish BOS**: candle closes below previous swing low → Bias = SHORT only.
- Defines the Swing Range (Low → High for BULL, High → Low for BEAR) and equilibrium (50%).

**Step 2 — 1H Location (Order Block POI)**
- **Discount zone** (below 50%): valid for LONG demand OBs.
- **Premium zone** (above 50%): valid for SHORT supply OBs.
- **Demand OB**: last *bearish* 1H candle before a bullish impulse move, located below equilibrium, unmitigated.
- **Supply OB**: last *bullish* 1H candle before a bearish impulse move, located above equilibrium, unmitigated.

**Step 3 — 15M Execution (3 Confluences required)**
1. **Rejection**: Bullish/Bearish Engulfing OR Pin Bar inside the OB zone.
2. **MSS**: 15M candle closes beyond the recent local swing (breaks micro-structure in bias direction).
3. **Failed Counter-Move**: The immediately following opposite candle fails to close past the rejection candle's extreme.

**Risk Management**: SL = below rejection low (LONG) / above rejection high (SHORT). TP = 1:3 R:R fixed.
        """)

    if st.button("▶ Run DLE Backtest", key="dl_run", type="primary"):
        try:
            from dle_strategy import DLEStrategyEngine
        except ImportError as _dle_err:
            st.error(f"dle_strategy.py not found: {_dle_err}")
            st.stop()

        _dl_spin = st.spinner(f"Fetching 1H (→4H resample) / 1H / 15M data for {_dl_idx} ({_dl_days} days)…")
        _dl_err_detail = []
        with _dl_spin:
            _dl_eng = DLEStrategyEngine(
                index=_dl_idx, days=_dl_days,
                swing_n=_dl_swing, min_impulse=_dl_imp)
            try:
                _dl_res = _dl_eng.run()
            except Exception as _dl_ex:
                _dl_res = None
                _dl_err_detail.append(str(_dl_ex))

        if _dl_res is None:
            st.error("Could not fetch data — yfinance returned empty results for one or more timeframes.")
            # Quick diagnostic
            import yfinance as _dl_yf
            _dl_sym = {"NIFTY":"^NSEI","BANKNIFTY":"^NSEBANK","SENSEX":"^BSESN"}.get(_dl_idx,"^NSEI")
            try:
                _dl_test = _dl_yf.download(_dl_sym, period="5d", interval="1h", progress=False, auto_adjust=True)
                if _dl_test.empty:
                    st.error(f"yfinance returned empty data for {_dl_sym} (1h). Check symbol or internet.")
                else:
                    st.info(f"yfinance OK for {_dl_sym} — got {len(_dl_test)} 1H bars. "
                            f"Likely the 15M data is unavailable (>60 days selected). "
                            f"Try 30 or 45 days.")
            except Exception as _dl_te:
                st.error(f"yfinance connection failed: {_dl_te}")
            if _dl_err_detail:
                st.code('\n'.join(_dl_err_detail))
        elif _dl_res.total_trades == 0:
            st.warning("No DLE setups found for this period. Try increasing the period or reducing swing lookback.")
        else:
            # ── Metrics row ─────────────────────────────────────────────────
            _m1,_m2,_m3,_m4,_m5,_m6 = st.columns(6)
            _m1.metric("Trades",      _dl_res.total_trades)
            _m2.metric("Win Rate",    f"{_dl_res.win_rate}%")
            _m3.metric("Total PnL",   f"{_dl_res.total_pnl_pts} pts")
            _m4.metric("Profit Factor", _dl_res.profit_factor)
            _m5.metric("Avg Win",     f"{_dl_res.avg_win_pts} pts")
            _m6.metric("Max DD",      f"{_dl_res.max_drawdown_pts} pts")

            # ── Current bias banner ─────────────────────────────────────────
            _bias_col = "#16a34a" if _dl_res.current_bias == "BULL" else "#dc2626" if _dl_res.current_bias == "BEAR" else "#64748b"
            _bias_lbl = "BULLISH (LONG only)" if _dl_res.current_bias == "BULL" else "BEARISH (SHORT only)" if _dl_res.current_bias == "BEAR" else "NEUTRAL"
            st.markdown(
                f'<div style="background:{_bias_col}22;border:1px solid {_bias_col};border-radius:8px;'
                f'padding:8px 16px;margin:8px 0;font-family:JetBrains Mono;font-size:13px">'
                f'<b style="color:{_bias_col}">4H Bias: {_bias_lbl}</b> &nbsp;|&nbsp; '
                f'Equilibrium: <b>{_dl_res.current_equil}</b></div>',
                unsafe_allow_html=True)

            # ── Equity curve ────────────────────────────────────────────────
            import plotly.graph_objects as _dl_go
            _eq_fig = _dl_go.Figure()
            _eq_fig.add_trace(_dl_go.Scatter(
                y=_dl_res.equity_curve, mode='lines+markers',
                line=dict(color='#3b82f6', width=2),
                marker=dict(size=4),
                name='Equity'))
            _eq_fig.update_layout(
                title="Equity Curve (₹ paper capital)",
                height=260, margin=dict(l=0,r=0,t=30,b=0),
                paper_bgcolor='white', plot_bgcolor='#f8faff',
                font=dict(family='JetBrains Mono', size=10))
            st.plotly_chart(_eq_fig, width='stretch')

            # ── Trades table ────────────────────────────────────────────────
            _dl_rows = []
            for _t in _dl_res.trades:
                _dl_rows.append({
                    "Date":       _t.date,
                    "Time":       _t.time,
                    "Bias":       _t.htf_bias,
                    "OB Zone":    f"{_t.ob_zone_low}–{_t.ob_zone_high}",
                    "Equil":      _t.equil,
                    "Entry":      _t.entry,
                    "SL":         _t.stop,
                    "TP":         _t.target,
                    "Risk pts":   _t.risk_pts,
                    "Exit":       _t.exit_price,
                    "Exit Time":  _t.exit_time,
                    "Reason":     _t.exit_reason,
                    "PnL pts":    _t.pnl_pts,
                    "Won":        "✓" if _t.won else "✗",
                    "C1":         _t.c1,
                    "C2":         _t.c2,
                    "C3":         _t.c3,
                })
            _dl_df = pd.DataFrame(_dl_rows)

            # Colour won/lost rows
            def _dl_style(row):
                clr = "#d1fae5" if row["Won"] == "✓" else "#fee2e2"
                return [f"background-color:{clr}"] * len(row)

            st.markdown("**Trade Log**")
            st.dataframe(_dl_df.style.apply(_dl_style, axis=1),
                         width='stretch', hide_index=True)

            # ── Excel export ────────────────────────────────────────────────
            import io as _dl_io
            _dl_xl = _dl_io.BytesIO()
            try:
                with pd.ExcelWriter(_dl_xl, engine='openpyxl') as _dl_xw:
                    _dl_df.to_excel(_dl_xw, sheet_name='DLE_Trades', index=False)
                    pd.DataFrame([{
                        "Symbol": _dl_res.symbol,
                        "Period (days)": _dl_days,
                        "Total Trades": _dl_res.total_trades,
                        "Wins": _dl_res.winning_trades,
                        "Losses": _dl_res.losing_trades,
                        "Win Rate %": _dl_res.win_rate,
                        "Total PnL pts": _dl_res.total_pnl_pts,
                        "Avg Win pts": _dl_res.avg_win_pts,
                        "Avg Loss pts": _dl_res.avg_loss_pts,
                        "Profit Factor": _dl_res.profit_factor,
                        "Max DD pts": _dl_res.max_drawdown_pts,
                        "Current 4H Bias": _dl_res.current_bias,
                        "Equilibrium": _dl_res.current_equil,
                    }]).to_excel(_dl_xw, sheet_name='Summary', index=False)
                _dl_xl.seek(0)
                _dl_fn = f"DLE_{_dl_idx}_{_dl_days}d_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
                st.download_button(
                    label=f"⬇ Download Excel — {_dl_fn}",
                    data=_dl_xl.getvalue(), file_name=_dl_fn,
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    key="dl_dl_btn")
            except ImportError:
                st.info("Install openpyxl for Excel export: pip install openpyxl")
